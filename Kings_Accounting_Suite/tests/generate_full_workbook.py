"""
GENERATE FULL WORKBOOK
======================
Comprehensive Excel workbook for Aureon Creator
Tax Years 2024/25 and 2025/26

Sheets:
1. Dashboard — Summary of both tax years side-by-side
2. P&L 2024/25 — Profit & Loss for tax year 2024/25
3. P&L 2025/26 — Profit & Loss for tax year 2025/26
4. SA103 2024/25 — HMRC Self Assessment boxes
5. SA103 2025/26 — HMRC Self Assessment boxes
6. Transaction Ledger — All classified transactions
7. Tax Estimate — Full tax calculation for both years
8. Private Advisory — Notes for Gary only

Aureon Creator / Aureon Research — April 2026
"""

import sys, os, json, re
from datetime import datetime, date
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from core.hnc_import import HNCImportEngine
from core.hnc_soup import HNCSoup, SA103_CATEGORIES
from core.hnc_soup_kitchen import HNCSoupKitchen
from core.cis_reconciliation import CISReconciliation
from core.tax_strategy import TaxStrategy
from core.hnc_metacognition import HNCMetacognition
from core.hnc_intelligence_registry import HNCIntelligenceRegistry

# ========================================================================
# BRAND PALETTE
# ========================================================================
NAVY = "1B2A4A"
BLUE = "2E5090"
GOLD = "C49A2E"
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"
RED = "CC0000"
GREEN = "006600"

header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=NAVY)
sub_header_font = Font(name="Arial", bold=True, color=NAVY, size=10)
sub_header_fill = PatternFill("solid", fgColor=LIGHT_GREY)
gold_font = Font(name="Arial", bold=True, color=GOLD, size=12)
navy_font = Font(name="Arial", bold=True, color=NAVY, size=10)
normal_font = Font(name="Arial", size=10)
input_font = Font(name="Arial", size=10, color="0000FF")
formula_font = Font(name="Arial", size=10, color="000000")
red_font = Font(name="Arial", bold=True, color=RED, size=10)
green_font = Font(name="Arial", bold=True, color=GREEN, size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
gbp_format = '£#,##0.00'
pct_format = '0.0%'

# ========================================================================
# DATA PROCESSING (same as build_full_picture.py)
# ========================================================================
SEARCH_DIRS = [
    os.path.join(os.path.dirname(__file__), "..", "..", "uploads"),
    "/sessions/upbeat-stoic-hamilton/mnt/uploads",
]

TAX_YEAR_2425_START = date(2024, 4, 6)
TAX_YEAR_2425_END = date(2025, 4, 5)
TAX_YEAR_2526_START = date(2025, 4, 6)
TAX_YEAR_2526_END = date(2026, 4, 5)

def find_csv_files():
    seen_names = set()
    csv_files = []
    for d in SEARCH_DIRS:
        if not os.path.isdir(d):
            continue
        for f in sorted(os.listdir(d)):
            if not f.endswith(".csv") or "Statement" not in f:
                continue
            match = re.search(r'Statement_(\d{2}_\w{3}_\d{4}_\d{2}_\w{3}_\d{4})', f)
            date_key = match.group(1) if match else f
            if date_key not in seen_names:
                seen_names.add(date_key)
                csv_files.append(os.path.join(d, f))
    csv_files.sort()
    return csv_files

def read_utf16_csv(filepath):
    for enc in ("utf-16", "utf-16-le", "utf-16-be", "utf-8-sig", "latin-1"):
        try:
            with open(filepath, "r", encoding=enc) as f:
                text = f.read()
            text = text.replace("\x00", "")
            if "Date" in text:
                return text
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise RuntimeError(f"Could not decode {filepath}")

def parse_date_str(date_str):
    if not date_str:
        return None
    for fmt in ("%d %b %Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(date_str.strip(), fmt).date()
        except (ValueError, AttributeError):
            continue
    return None

def get_tax_year(d):
    if d is None:
        return "unknown"
    if d < TAX_YEAR_2425_START:
        return "pre-2024/25"
    elif d <= TAX_YEAR_2425_END:
        return "2024/25"
    elif d <= TAX_YEAR_2526_END:
        return "2025/26"
    else:
        return "post-2025/26"

# Import and process
csv_files = find_csv_files()
importer = HNCImportEngine()
for filepath in csv_files:
    try:
        csv_text = read_utf16_csv(filepath)
        importer.import_csv_string(csv_text, os.path.basename(filepath))
    except:
        pass

all_txns = importer.get_bank_transactions()

# Dedup
seen = set()
unique_txns = []
for t in all_txns:
    key = (t.get("date", ""), t.get("amount", 0), t.get("description", "")[:40])
    if key not in seen:
        seen.add(key)
        unique_txns.append(t)

# Split by tax year
by_year = defaultdict(list)
for t in unique_txns:
    d = parse_date_str(t.get("date", ""))
    ty = get_tax_year(d)
    by_year[ty].append(t)

# Run through Soup, then Kitchen
year_data = {}
for year_key in ["2024/25", "2025/26"]:
    txns = by_year.get(year_key, [])
    if not txns:
        continue
    soup = HNCSoup(entity_name="Aureon Creator",
                   trades=["construction", "food", "consulting", "digital", "property"])
    results = soup.classify_all(txns)

    # Run the Kitchen — auto-correct before output
    ty_start = int(year_key.split("/")[0])
    kitchen = HNCSoupKitchen(soup, tax_year=ty_start)
    kitchen_report = kitchen.audit_and_correct()

    year_data[year_key] = {
        "soup": soup, "results": results, "txns": txns,
        "kitchen": kitchen, "kitchen_report": kitchen_report,
    }
    print(f"  {year_key}: {kitchen_report.corrections_made} Kitchen corrections, "
          f"{len([r for r in kitchen_report.risks_found if r.severity == 'CRITICAL'])} critical risks")

print("Data processed. Building workbook...")

# ========================================================================
# BUILD WORKBOOK
# ========================================================================

wb = Workbook()

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

def style_sub_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
        cell.border = thin_border

def write_currency(ws, row, col, value, font=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = gbp_format
    cell.font = font or formula_font
    cell.border = thin_border
    return cell

# ========================================================================
# SHEET 1: DASHBOARD
# ========================================================================
ws = wb.active
ws.title = "Dashboard"
ws.sheet_properties.tabColor = NAVY

# Title
ws.merge_cells("A1:G1")
ws["A1"] = "THE HNC ACCOUNTANT — FULL FINANCIAL PICTURE"
ws["A1"].font = Font(name="Arial", bold=True, color=GOLD, size=16)
ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
ws["A1"].alignment = Alignment(horizontal="center")

ws.merge_cells("A2:G2")
ws["A2"] = "Aureon Creator | Self-Employed Trader | Sep 2024 — Apr 2026"
ws["A2"].font = Font(name="Arial", color=WHITE, size=11)
ws["A2"].fill = PatternFill("solid", fgColor=BLUE)
ws["A2"].alignment = Alignment(horizontal="center")

# Column widths
for col, w in enumerate([30, 18, 18, 5, 18, 18, 5], 1):
    ws.column_dimensions[get_column_letter(col)].width = w

# Headers row 4
row = 4
ws.cell(row=row, column=1, value="").font = navy_font
ws.cell(row=row, column=2, value="2024/25").font = header_font
ws.cell(row=row, column=2).fill = header_fill
ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
ws.cell(row=row, column=3, value="Notes").font = header_font
ws.cell(row=row, column=3).fill = header_fill
ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
ws.cell(row=row, column=5, value="2025/26").font = header_font
ws.cell(row=row, column=5).fill = header_fill
ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
ws.cell(row=row, column=6, value="Notes").font = header_font
ws.cell(row=row, column=6).fill = header_fill
ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")

# Summary data
for yr_key, col_start in [("2024/25", 2), ("2025/26", 5)]:
    data = year_data.get(yr_key)
    if not data:
        continue
    soup = data["soup"]
    income = soup.get_income_by_trade()
    expenses = soup.get_sa103_summary()
    transfers = soup.get_transfer_summary()
    total_income = sum(income.values())
    total_expenses = sum(expenses.values())
    net_profit = total_income - total_expenses
    non_allowable = soup.get_non_allowable()
    na_total = sum(r.original.get("amount", 0) for r in non_allowable)

    # CIS and drawings data from Kitchen
    kr = data["kitchen_report"]
    cis_gross = kr.cis_gross_income
    cis_deducted = kr.cis_deductions_credit
    cis_citb = kr.cis_citb_expense
    drawings = kr.drawings_total
    adjusted_income = max(total_income, cis_gross) if cis_gross > 0 else total_income

    rows_data = [
        ("Transactions Processed", len(data["txns"]), ""),
        ("Transfers Excluded", transfers["count"], "Internal movements"),
        ("", "", ""),
        ("BANK TURNOVER (NET)", total_income, "As per bank statements"),
    ]
    if cis_gross > 0:
        rows_data += [
            ("CIS GROSS INCOME", cis_gross, "Invoice gross (before deductions)"),
            ("  CIS Tax Deducted (20%)", cis_deducted, "Reclaimable tax credit"),
            ("  CITB Levy (0.7%)", cis_citb, "Allowable expense"),
        ]
    rows_data += [
        ("  Construction", income.get("construction", 0), "Grove/Heron/AMC/CSR"),
        ("  Food", income.get("food", 0), "Food Venture"),
        ("  Consulting", income.get("consulting", 0), "Aureon Consulting Entity"),
        ("  Property", income.get("property", 0), "16 Howard St"),
        ("  Other", income.get("general", 0) + income.get("digital", 0), "Unclassified income"),
        ("", "", ""),
        ("ALLOWABLE EXPENSES", total_expenses, ""),
        ("  Cost of Sales (Box 10)", expenses.get("cost_of_sales", 0), "Food stock, materials"),
        ("  Other Direct Costs (Box 12)", expenses.get("other_direct_costs", 0), "Labour, subcontractors"),
        ("  Premises (Box 14)", expenses.get("premises", 0), "Rent, rates"),
        ("  Admin (Box 15)", expenses.get("admin", 0), "Software, subscriptions"),
        ("  Advertising (Box 16)", expenses.get("advertising", 0), "Marketing"),
        ("  Interest (Box 17)", expenses.get("interest", 0), "Bank charges"),
        ("  Other Expenses (Box 19)", expenses.get("other_expenses", 0), "Sundry"),
        ("  Motor (Box 20)", expenses.get("motor", 0), "Fuel, vehicle, travel"),
        ("", "", ""),
        ("PERSONAL DRAWINGS", drawings, "Below profit line"),
        ("NON-ALLOWABLE", na_total, "Personal spending"),
        ("", "", ""),
        ("NET PROFIT", net_profit, "Box 21"),
    ]

    for i, (label, value, note) in enumerate(rows_data):
        r = 5 + i
        ws.cell(row=r, column=1, value=label).font = navy_font if label.isupper() or label.startswith("NET") else normal_font
        if isinstance(value, (int, float)) and value != 0:
            cell = ws.cell(row=r, column=col_start, value=value)
            cell.number_format = gbp_format if isinstance(value, float) else '#,##0'
            cell.font = gold_font if "NET PROFIT" in label or "GROSS TURNOVER" in label else input_font
            cell.border = thin_border
        elif isinstance(value, int) and value == 0:
            pass
        ws.cell(row=r, column=col_start + 1, value=note).font = Font(name="Arial", size=9, color="666666")

# Tax estimate section
tax_row = 30
ws.cell(row=tax_row, column=1, value="TAX ESTIMATES").font = Font(name="Arial", bold=True, color=GOLD, size=12)
ws.cell(row=tax_row, column=1).fill = PatternFill("solid", fgColor=NAVY)
ws.merge_cells(f"A{tax_row}:G{tax_row}")

for yr_key, col_start in [("2024/25", 2), ("2025/26", 5)]:
    data = year_data.get(yr_key)
    if not data:
        continue
    soup = data["soup"]
    income = soup.get_income_by_trade()
    expenses = soup.get_sa103_summary()
    total_income = sum(income.values())
    total_expenses = sum(expenses.values())
    net_profit = total_income - total_expenses

    pa = 12_570
    taxable = max(0, net_profit - pa)
    if taxable <= 37_700:
        tax = taxable * 0.20
    elif taxable <= 125_140:
        tax = 37_700 * 0.20 + (taxable - 37_700) * 0.40
    else:
        tax = 37_700 * 0.20 + 87_440 * 0.40 + (taxable - 125_140) * 0.45

    ni_class2 = 179.40 if net_profit >= 12_570 else 0
    ni_class4 = 0
    if net_profit > 12_570:
        ni_band1 = min(net_profit, 50_270) - 12_570
        ni_class4 = ni_band1 * 0.06
        if net_profit > 50_270:
            ni_class4 += (net_profit - 50_270) * 0.02

    total_tax = tax + ni_class2 + ni_class4

    # Get CIS credit for this tax year
    kr = data["kitchen_report"]
    cis_credit = kr.cis_deductions_credit

    tax_items = [
        ("Net Profit", net_profit),
        ("Personal Allowance", pa),
        ("Taxable Income", taxable),
        ("Income Tax (20%/40%)", tax),
        ("NI Class 2", ni_class2),
        ("NI Class 4 (6%/2%)", ni_class4),
        ("TOTAL TAX LIABILITY", total_tax),
    ]
    if cis_credit > 0:
        tax_items.append(("Less CIS Tax Credit", -cis_credit))
        tax_items.append(("TAX REMAINING TO PAY", max(0, total_tax - cis_credit)))
    tax_items.append(("TAKE-HOME PROFIT", net_profit - total_tax))

    for i, (label, value) in enumerate(tax_items):
        r = tax_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = navy_font if "TOTAL" in label or "TAKE-HOME" in label else normal_font
        cell = ws.cell(row=r, column=col_start, value=value)
        cell.number_format = gbp_format
        cell.font = gold_font if "TAKE-HOME" in label else (red_font if "TOTAL TAX" in label else input_font)
        cell.border = thin_border


# ========================================================================
# SHEET 2 & 3: P&L by Tax Year
# ========================================================================
for yr_key in ["2024/25", "2025/26"]:
    data = year_data.get(yr_key)
    if not data:
        continue

    ws = wb.create_sheet(f"P&L {yr_key.replace('/', '-')}")
    ws.sheet_properties.tabColor = BLUE

    soup = data["soup"]
    income = soup.get_income_by_trade()
    expenses = soup.get_sa103_summary()

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12

    ws.merge_cells("A1:C1")
    ws["A1"] = f"PROFIT & LOSS — Tax Year {yr_key}"
    ws["A1"].font = Font(name="Arial", bold=True, color=GOLD, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)

    ws["A2"] = "Aureon Creator — Self-Employed Trader"
    ws["A2"].font = Font(name="Arial", color=NAVY, size=10)

    # Income section
    row = 4
    ws.cell(row=row, column=1, value="INCOME").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2, value="Amount").font = header_font
    ws.cell(row=row, column=2).fill = header_fill
    ws.cell(row=row, column=3, value="% of Total").font = header_font
    ws.cell(row=row, column=3).fill = header_fill

    total_income = sum(income.values())
    income_start = row + 1
    for trade, amount in sorted(income.items(), key=lambda x: -x[1]):
        row += 1
        ws.cell(row=row, column=1, value=f"  {trade.title()} Income").font = normal_font
        write_currency(ws, row, 2, amount, input_font)
        pct = amount / total_income if total_income else 0
        cell = ws.cell(row=row, column=3, value=pct)
        cell.number_format = pct_format
        cell.font = normal_font
    income_end = row

    row += 1
    ws.cell(row=row, column=1, value="TOTAL INCOME").font = navy_font
    ws.cell(row=row, column=2).font = gold_font
    ws.cell(row=row, column=2).number_format = gbp_format
    ws.cell(row=row, column=2, value=f"=SUM(B{income_start}:B{income_end})")
    ws.cell(row=row, column=2).border = thin_border
    total_income_row = row

    # Expenses section
    row += 2
    ws.cell(row=row, column=1, value="EXPENSES").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2, value="Amount").font = header_font
    ws.cell(row=row, column=2).fill = header_fill
    ws.cell(row=row, column=3, value="SA103 Box").font = header_font
    ws.cell(row=row, column=3).fill = header_fill

    total_expenses = sum(expenses.values())
    exp_start = row + 1
    for cat in ["cost_of_sales", "other_direct_costs", "premises", "admin",
                "advertising", "interest", "other_expenses", "motor"]:
        amount = expenses.get(cat, 0)
        if amount == 0:
            continue
        row += 1
        info = SA103_CATEGORIES.get(cat, {})
        ws.cell(row=row, column=1, value=f"  {info.get('label', cat)[:45]}").font = normal_font
        write_currency(ws, row, 2, amount, input_font)
        ws.cell(row=row, column=3, value=info.get("box", "")).font = normal_font
    exp_end = row

    row += 1
    ws.cell(row=row, column=1, value="TOTAL EXPENSES").font = navy_font
    ws.cell(row=row, column=2).font = red_font
    ws.cell(row=row, column=2).number_format = gbp_format
    ws.cell(row=row, column=2, value=f"=SUM(B{exp_start}:B{exp_end})")
    ws.cell(row=row, column=2).border = thin_border
    total_exp_row = row

    # Net Profit
    row += 2
    ws.cell(row=row, column=1, value="NET PROFIT").font = Font(name="Arial", bold=True, color=GOLD, size=14)
    ws.cell(row=row, column=2).font = gold_font
    ws.cell(row=row, column=2).number_format = gbp_format
    ws.cell(row=row, column=2, value=f"=B{total_income_row}-B{total_exp_row}")
    ws.cell(row=row, column=2).border = thin_border


# ========================================================================
# SHEET 4 & 5: SA103 SUMMARY
# ========================================================================
for yr_key in ["2024/25", "2025/26"]:
    data = year_data.get(yr_key)
    if not data:
        continue

    ws = wb.create_sheet(f"SA103 {yr_key.replace('/', '-')}")
    ws.sheet_properties.tabColor = GOLD

    soup = data["soup"]
    income = soup.get_income_by_trade()
    expenses = soup.get_sa103_summary()
    total_income = sum(income.values())
    total_expenses = sum(expenses.values())

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 18

    ws.merge_cells("A1:C1")
    ws["A1"] = f"SA103S — Self Employment (Short) — {yr_key}"
    ws["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=12)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)

    ws["A2"] = "HMRC Self Assessment Tax Return"
    ws["A2"].font = Font(name="Arial", size=10, color=NAVY)

    row = 4
    headers = ["Box", "Description", "Amount"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i+1, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border

    # Income
    row += 1
    ws.cell(row=row, column=1, value="Box 9").font = navy_font
    ws.cell(row=row, column=2, value="Turnover — the annual takings before expenses").font = navy_font
    write_currency(ws, row, 3, total_income, gold_font)
    income_row = row

    # Expenses
    box_map = [
        ("Box 10", "cost_of_sales", "Cost of goods bought for resale or goods used"),
        ("Box 11", "construction_costs", "Construction industry costs"),
        ("Box 12", "other_direct_costs", "Other direct costs"),
        ("Box 13", None, "Staff costs (NOT USED — no PAYE employees)"),
        ("Box 14", "premises", "Premises costs"),
        ("Box 15", "admin", "Admin costs"),
        ("Box 16", "advertising", "Advertising and business entertainment costs"),
        ("Box 17", "interest", "Interest and alternative finance payments"),
        ("Box 18", "phone", "Phone, fax, stationery and other office costs"),
        ("Box 19", "other_expenses", "Other business expenses"),
        ("Box 20", "motor", "Car, van and travel expenses"),
    ]

    exp_start = row + 1
    for box, cat, desc in box_map:
        row += 1
        amount = expenses.get(cat, 0) if cat else 0
        ws.cell(row=row, column=1, value=box).font = normal_font
        ws.cell(row=row, column=2, value=desc).font = normal_font
        write_currency(ws, row, 3, amount, input_font if amount > 0 else normal_font)
    exp_end = row

    # Total expenses
    row += 1
    ws.cell(row=row, column=1, value="Box 21").font = navy_font
    ws.cell(row=row, column=2, value="Total allowable expenses").font = navy_font
    ws.cell(row=row, column=3, value=f"=SUM(C{exp_start}:C{exp_end})").font = red_font
    ws.cell(row=row, column=3).number_format = gbp_format
    ws.cell(row=row, column=3).border = thin_border
    total_exp_row = row

    # Net profit
    row += 1
    ws.cell(row=row, column=1, value="Box 22").font = Font(name="Arial", bold=True, color=GOLD, size=11)
    ws.cell(row=row, column=2, value="Net profit (turnover minus expenses)").font = Font(name="Arial", bold=True, color=GOLD, size=11)
    ws.cell(row=row, column=3, value=f"=C{income_row}-C{total_exp_row}").font = gold_font
    ws.cell(row=row, column=3).number_format = gbp_format
    ws.cell(row=row, column=3).border = thin_border


# ========================================================================
# SHEET 6: TRANSACTION LEDGER
# ========================================================================
ws = wb.create_sheet("Transaction Ledger")
ws.sheet_properties.tabColor = BLUE

ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 45
ws.column_dimensions["D"].width = 6
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 20
ws.column_dimensions["G"].width = 10
ws.column_dimensions["H"].width = 25
ws.column_dimensions["I"].width = 15

row = 1
headers = ["Date", "Tax Year", "Description", "Dir", "Amount", "HMRC Category", "SA103 Box", "Label", "Trade"]
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h)
    c.font = header_font
    c.fill = header_fill
    c.border = thin_border

ws.auto_filter.ref = f"A1:I1"

# Write all results
all_sorted_results = []
for yr_key in ["2024/25", "2025/26"]:
    data = year_data.get(yr_key)
    if not data:
        continue
    for r in data["results"]:
        if not r.is_transfer:
            all_sorted_results.append((yr_key, r))

all_sorted_results.sort(key=lambda x: x[1].original.get("date", ""))

for yr_key, r in all_sorted_results:
    row += 1
    t = r.original
    ws.cell(row=row, column=1, value=t.get("date", "")).font = normal_font
    ws.cell(row=row, column=2, value=yr_key).font = normal_font
    ws.cell(row=row, column=3, value=t.get("description", "")[:50]).font = normal_font
    ws.cell(row=row, column=4, value=t.get("direction", "")).font = normal_font
    write_currency(ws, row, 5, t.get("amount", 0), green_font if r.is_income else red_font)
    ws.cell(row=row, column=6, value=r.hmrc_category).font = normal_font
    ws.cell(row=row, column=7, value=r.sa103_box or ("Income" if r.is_income else "")).font = normal_font
    ws.cell(row=row, column=8, value=r.hmrc_label[:30]).font = normal_font
    ws.cell(row=row, column=9, value=r.trade).font = normal_font

ledger_end = row
print(f"  Transaction Ledger: {ledger_end - 1} rows")


# ========================================================================
# SHEET 7: TAX ESTIMATE
# ========================================================================
ws = wb.create_sheet("Tax Estimate")
ws.sheet_properties.tabColor = RED

ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 18

ws.merge_cells("A1:C1")
ws["A1"] = "TAX ESTIMATE — BOTH TAX YEARS"
ws["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
ws["A1"].fill = PatternFill("solid", fgColor=RED)

row = 3
ws.cell(row=row, column=1, value="").font = navy_font
ws.cell(row=row, column=2, value="2024/25").font = header_font
ws.cell(row=row, column=2).fill = header_fill
ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
ws.cell(row=row, column=3, value="2025/26").font = header_font
ws.cell(row=row, column=3).fill = header_fill
ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

for yr_key, col in [("2024/25", 2), ("2025/26", 3)]:
    data = year_data.get(yr_key)
    if not data:
        continue
    soup = data["soup"]
    income = soup.get_income_by_trade()
    expenses = soup.get_sa103_summary()
    total_income = sum(income.values())
    total_expenses = sum(expenses.values())
    net_profit = total_income - total_expenses
    pa = 12_570
    taxable = max(0, net_profit - pa)
    if taxable <= 37_700:
        tax = taxable * 0.20
    elif taxable <= 125_140:
        tax = 37_700 * 0.20 + (taxable - 37_700) * 0.40
    else:
        tax = 37_700 * 0.20 + 87_440 * 0.40 + (taxable - 125_140) * 0.45

    ni_class2 = 179.40 if net_profit >= 12_570 else 0
    ni_class4 = 0
    if net_profit > 12_570:
        ni_band1 = min(net_profit, 50_270) - 12_570
        ni_class4 = ni_band1 * 0.06
        if net_profit > 50_270:
            ni_class4 += (net_profit - 50_270) * 0.02

    total_tax = tax + ni_class2 + ni_class4

    items = [
        ("Gross Turnover", total_income),
        ("Allowable Expenses", total_expenses),
        ("Net Profit", net_profit),
        ("", ""),
        ("Personal Allowance (2024/25)", pa),
        ("Taxable Income", taxable),
        ("", ""),
        ("Income Tax — Basic Rate (20%)", min(taxable, 37_700) * 0.20 if taxable > 0 else 0),
        ("Income Tax — Higher Rate (40%)", max(0, taxable - 37_700) * 0.40 if taxable > 37_700 else 0),
        ("Total Income Tax", tax),
        ("", ""),
        ("NI Class 2 (£3.45/week)", ni_class2),
        ("NI Class 4 — Main (6%)", min(net_profit, 50_270) - 12_570 if net_profit > 12_570 else 0),
        ("NI Class 4 — Additional (2%)", max(0, net_profit - 50_270) * 0.02 if net_profit > 50_270 else 0),
        ("Total NI", ni_class2 + ni_class4),
        ("", ""),
        ("TOTAL TAX LIABILITY", total_tax),
        ("TAKE-HOME PROFIT", net_profit - total_tax),
        ("", ""),
        ("Effective Tax Rate", total_tax / net_profit if net_profit else 0),
    ]

    for i, (label, value) in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).font = navy_font if any(x in label for x in ["TOTAL", "TAKE", "Net Profit", "Gross"]) else normal_font
        if isinstance(value, str) and value == "":
            continue
        cell = ws.cell(row=r, column=col, value=value)
        if "Rate" in label and "Tax Rate" not in label:
            cell.number_format = gbp_format
        elif "Effective" in label:
            cell.number_format = pct_format
        else:
            cell.number_format = gbp_format
        cell.font = gold_font if "TAKE-HOME" in label else (red_font if "TOTAL TAX" in label else input_font)
        cell.border = thin_border


# ========================================================================
# SHEET 8: RISK ASSESSMENT (Kitchen Report)
# ========================================================================
ws = wb.create_sheet("Risk Assessment")
ws.sheet_properties.tabColor = "CC0000"

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 65
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 45

ws.merge_cells("A1:E1")
ws["A1"] = "SOUP KITCHEN — HMRC RISK ASSESSMENT"
ws["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
ws["A1"].fill = PatternFill("solid", fgColor="CC0000")

ws.merge_cells("A2:E2")
ws["A2"] = "Auto-generated stress test. Every risk has been corrected where possible."
ws["A2"].font = Font(name="Arial", size=10, color=RED, italic=True)

row = 4
headers_risk = ["Tax Year", "Severity", "Description", "Amount", "Action Taken"]
for i, h in enumerate(headers_risk):
    c = ws.cell(row=row, column=i+1, value=h)
    c.font = header_font
    c.fill = header_fill
    c.border = thin_border

severity_colors = {
    "CRITICAL": PatternFill("solid", fgColor="FFCCCC"),
    "HIGH": PatternFill("solid", fgColor="FFE0CC"),
    "MEDIUM": PatternFill("solid", fgColor="FFFFCC"),
    "LOW": PatternFill("solid", fgColor="E0FFE0"),
    "INFO": PatternFill("solid", fgColor="CCE0FF"),
}

for yr_key in ["2024/25", "2025/26"]:
    data = year_data.get(yr_key)
    if not data or "kitchen_report" not in data:
        continue
    report = data["kitchen_report"]

    for risk in sorted(report.risks_found, key=lambda x: {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}.get(x.severity, 5)):
        row += 1
        ws.cell(row=row, column=1, value=yr_key).font = normal_font
        sev_cell = ws.cell(row=row, column=2, value=risk.severity)
        sev_cell.font = Font(name="Arial", bold=True, size=10,
                            color=RED if risk.severity == "CRITICAL" else "000000")
        sev_cell.fill = severity_colors.get(risk.severity, PatternFill())
        ws.cell(row=row, column=3, value=risk.description[:80]).font = normal_font
        if risk.amount:
            write_currency(ws, row, 4, risk.amount)
        ws.cell(row=row, column=5, value=risk.action[:55]).font = normal_font
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border

# Defence narrative
row += 3
ws.cell(row=row, column=1, value="DEFENCE NARRATIVE").font = Font(name="Arial", bold=True, color=GOLD, size=12)
ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=NAVY)
ws.merge_cells(f"A{row}:E{row}")
row += 1

for yr_key in ["2024/25", "2025/26"]:
    data = year_data.get(yr_key)
    if not data or "kitchen_report" not in data:
        continue
    ws.cell(row=row, column=1, value=f"--- {yr_key} ---").font = navy_font
    row += 1
    for line in data["kitchen_report"].defence_narrative:
        if line.strip():
            ws.merge_cells(f"A{row}:E{row}")
            ws.cell(row=row, column=1, value=line.strip()).font = normal_font
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        row += 1
    row += 1


# ========================================================================
# SHEET 9: PRIVATE ADVISORY
# ========================================================================
ws = wb.create_sheet("Private Advisory")
ws.sheet_properties.tabColor = "FF0000"

ws.column_dimensions["A"].width = 80

ws.merge_cells("A1:A1")
ws["A1"] = "PRIVATE ADVISORY — FOR YOUR EYES ONLY"
ws["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
ws["A1"].fill = PatternFill("solid", fgColor="FF0000")

ws["A2"] = "These notes are NEVER sent to HMRC. They are strategic recommendations."
ws["A2"].font = Font(name="Arial", color=RED, size=10, italic=True)

row = 4
for yr_key in ["2024/25", "2025/26"]:
    data = year_data.get(yr_key)
    if not data:
        continue
    ws.cell(row=row, column=1, value=f"--- TAX YEAR {yr_key} ---").font = navy_font
    row += 1

    advisory = data["soup"].get_private_advisory()
    seen_notes = set()
    for note in advisory:
        short = note.strip()[:80]
        if short in seen_notes:
            continue
        seen_notes.add(short)
        if note.startswith("\n["):
            row += 1
            ws.cell(row=row, column=1, value=note.strip()).font = Font(name="Arial", size=10, color=RED)
        else:
            ws.cell(row=row, column=1, value=note.strip()).font = normal_font
        row += 1
    row += 1


# ========================================================================
# SHEET 10: CIS RECONCILIATION
# ========================================================================
ws = wb.create_sheet("CIS Schedule")
ws.sheet_properties.tabColor = "006600"

ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 28
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 14
ws.column_dimensions["G"].width = 12
ws.column_dimensions["H"].width = 12
ws.column_dimensions["I"].width = 10
ws.column_dimensions["J"].width = 10
ws.column_dimensions["K"].width = 35

ws.merge_cells("A1:K1")
ws["A1"] = "CIS RECONCILIATION — Aureon Consulting Entity / Construction Client Alpha"
ws["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
ws["A1"].fill = PatternFill("solid", fgColor="006600")

ws.merge_cells("A2:K2")
ws["A2"] = "Construction Industry Scheme — Invoice Schedule, CIS Deductions & Tax Credits"
ws["A2"].font = Font(name="Arial", size=10, color="006600")

cis = CISReconciliation()
schedule = cis.get_full_schedule()

row = 4
cis_headers = ["Invoice", "Date", "Service Period", "Gross", "CITB", "Gross-CITB",
               "CIS 20%", "Net Paid", "Tax Year", "CIS Stmt", "Notes"]
for i, h in enumerate(cis_headers):
    c = ws.cell(row=row, column=i+1, value=h)
    c.font = header_font
    c.fill = PatternFill("solid", fgColor="006600")
    c.border = thin_border

cis_start = row + 1
for entry in schedule:
    row += 1
    ws.cell(row=row, column=1, value=entry["invoice"]).font = normal_font
    ws.cell(row=row, column=2, value=entry["date"]).font = normal_font
    ws.cell(row=row, column=3, value=entry["period"]).font = normal_font
    write_currency(ws, row, 4, entry["gross"], input_font)
    write_currency(ws, row, 5, entry["citb"])
    write_currency(ws, row, 6, entry["gross_less_citb"])
    write_currency(ws, row, 7, entry["cis_20pct"], red_font)
    write_currency(ws, row, 8, entry["net"], green_font)
    ws.cell(row=row, column=9, value=entry["tax_year"]).font = normal_font
    ws.cell(row=row, column=10, value="Yes" if entry["has_cis_statement"] else "No").font = (
        Font(name="Arial", size=10, color="006600") if entry["has_cis_statement"]
        else Font(name="Arial", size=10, color="CC0000"))
    ws.cell(row=row, column=11, value=entry["notes"][:40]).font = Font(name="Arial", size=9, color="666666")
    for col in range(1, 12):
        ws.cell(row=row, column=col).border = thin_border
cis_end = row

# Totals row
row += 1
ws.cell(row=row, column=1, value="TOTALS").font = navy_font
ws.cell(row=row, column=4, value=f"=SUM(D{cis_start}:D{cis_end})").font = gold_font
ws.cell(row=row, column=4).number_format = gbp_format
ws.cell(row=row, column=5, value=f"=SUM(E{cis_start}:E{cis_end})").font = navy_font
ws.cell(row=row, column=5).number_format = gbp_format
ws.cell(row=row, column=7, value=f"=SUM(G{cis_start}:G{cis_end})").font = red_font
ws.cell(row=row, column=7).number_format = gbp_format
ws.cell(row=row, column=8, value=f"=SUM(H{cis_start}:H{cis_end})").font = green_font
ws.cell(row=row, column=8).number_format = gbp_format
for col in range(1, 12):
    ws.cell(row=row, column=col).border = Border(top=Side(style="double"))

# CIS Summary
row += 3
ws.cell(row=row, column=1, value="CIS TAX CREDIT SUMMARY").font = Font(name="Arial", bold=True, color=WHITE, size=12)
ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="006600")
ws.merge_cells(f"A{row}:D{row}")

for ty_start in [2024, 2025]:
    ty_totals = cis.get_tax_year_totals(ty_start)
    if ty_totals["invoice_count"] == 0:
        continue
    row += 2
    ws.cell(row=row, column=1, value=f"Tax Year {ty_totals['tax_year']}").font = navy_font
    row += 1
    items = [
        ("Total Gross Invoiced", ty_totals["total_gross"]),
        ("CITB Levy Paid (allowable expense)", ty_totals["total_citb"]),
        ("CIS Tax Deducted (reclaimable)", ty_totals["total_cis_deducted"]),
        ("Net Received in Bank", ty_totals["total_net_received"]),
        ("CIS TAX CREDIT TO CLAIM", ty_totals["cis_tax_credit"]),
    ]
    for label, val in items:
        ws.cell(row=row, column=1, value=label).font = (
            Font(name="Arial", bold=True, color="006600", size=11) if "CREDIT" in label else normal_font)
        write_currency(ws, row, 2, val,
                      Font(name="Arial", bold=True, color="006600", size=11) if "CREDIT" in label else input_font)
        row += 1

# Known gaps
if cis.gaps:
    row += 2
    ws.cell(row=row, column=1, value="KNOWN GAPS").font = Font(name="Arial", bold=True, color="CC0000", size=11)
    for gap in cis.gaps:
        row += 1
        ws.cell(row=row, column=1, value=gap["period"]).font = normal_font
        ws.cell(row=row, column=2, value=gap["status"]).font = Font(name="Arial", color="CC0000", size=10)
        ws.cell(row=row, column=3, value=gap.get("notes", "")).font = Font(name="Arial", size=9, color="666666")

print(f"  CIS Schedule: {len(schedule)} invoices")


# ========================================================================
# SHEET 11: TAX STRATEGY — THE COOKING
# ========================================================================
ws = wb.create_sheet("Tax Strategy")
ws.sheet_properties.tabColor = "C49A2E"

ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 55
ws.column_dimensions["E"].width = 55
ws.column_dimensions["F"].width = 18

ws.merge_cells("A1:F1")
ws["A1"] = "TAX STRATEGY — WHAT THE SOUP SAVES YOU"
ws["A1"].font = Font(name="Arial", bold=True, color=GOLD, size=16)
ws["A1"].fill = PatternFill("solid", fgColor=NAVY)

ws.merge_cells("A2:F2")
ws["A2"] = "Every legitimate optimisation under UK tax law. Same rules. Same HMRC. Different result."
ws["A2"].font = Font(name="Arial", color=WHITE, size=11)
ws["A2"].fill = PatternFill("solid", fgColor=BLUE)

# Run strategy for 2025/26 (the main year with CIS)
data_2526 = year_data.get("2025/26")
if data_2526:
    soup = data_2526["soup"]
    income = soup.get_income_by_trade()
    expenses = soup.get_sa103_summary()
    total_income = sum(income.values())
    total_expenses = sum(expenses.values())
    net_profit = total_income - total_expenses
    kr = data_2526["kitchen_report"]

    strategy = TaxStrategy(
        net_profit=net_profit,
        total_income=total_income,
        total_expenses=total_expenses,
        motor_expenses=expenses.get("motor", 0),
        cis_deducted=kr.cis_deductions_credit,
        cis_citb=kr.cis_citb_expense,
        drawings=kr.drawings_total,
        spouse_income=0,
        business_miles_estimate=15_000,
        home_hours_per_month=80,
    )
    savings = strategy.run_all_strategies()
    report = strategy.get_strategy_report()

    # === THE HEADLINE ===
    row = 4
    ws.cell(row=row, column=1, value="THE BOTTOM LINE — TAX YEAR 2025/26").font = Font(name="Arial", bold=True, color=WHITE, size=12)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(f"A{row}:F{row}")

    row = 6
    ws.cell(row=row, column=1, value="Normal accountant files this:").font = Font(name="Arial", size=11, color="666666")
    write_currency(ws, row, 2, report["naive_tax"], Font(name="Arial", size=14, color="CC0000", bold=True))

    row = 7
    ws.cell(row=row, column=1, value="The Soup files this:").font = Font(name="Arial", size=11, color=NAVY, bold=True)
    write_currency(ws, row, 2, report["optimised_tax"], Font(name="Arial", size=14, color="006600", bold=True))

    row = 8
    ws.cell(row=row, column=1, value="ALREADY SAVED (auto-applied):").font = Font(name="Arial", size=11, color=GOLD, bold=True)
    write_currency(ws, row, 2, report["naive_tax"] - report["optimised_tax"],
                  Font(name="Arial", size=14, color=GOLD, bold=True))

    row = 9
    ws.cell(row=row, column=1, value="ADDITIONAL SAVINGS AVAILABLE:").font = Font(name="Arial", size=11, color="006600", bold=True)
    write_currency(ws, row, 2, report["additional_savings_available"],
                  Font(name="Arial", size=14, color="006600", bold=True))

    row = 10
    ws.cell(row=row, column=1, value="TOTAL POTENTIAL SAVING:").font = Font(name="Arial", size=12, color=GOLD, bold=True)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=NAVY)
    write_currency(ws, row, 2, report["total_potential_saving"],
                  Font(name="Arial", size=16, color=GOLD, bold=True))
    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(f"A{row}:A{row}")

    # === STRATEGY BREAKDOWN ===
    row = 12
    ws.cell(row=row, column=1, value="STRATEGY BREAKDOWN").font = Font(name="Arial", bold=True, color=WHITE, size=12)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(f"A{row}:F{row}")

    row = 13
    strat_headers = ["Strategy", "Annual Saving", "Risk", "What It Is", "What You Need To Do", "Legal Basis"]
    for i, h in enumerate(strat_headers):
        c = ws.cell(row=row, column=i+1, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border

    auto_fill = PatternFill("solid", fgColor="E0FFE0")
    action_fill = PatternFill("solid", fgColor="FFFFCC")

    strat_start = row + 1
    for s in savings:
        row += 1
        ws.cell(row=row, column=1, value=s.name).font = Font(name="Arial", bold=True, size=10)
        write_currency(ws, row, 2, s.annual_saving,
                      Font(name="Arial", bold=True, size=10, color="006600") if s.annual_saving > 0
                      else Font(name="Arial", size=10, color="666666"))
        ws.cell(row=row, column=3, value=s.risk_level).font = (
            Font(name="Arial", size=10, color="006600") if s.risk_level == "NONE"
            else Font(name="Arial", size=10, color="CC6600"))
        ws.cell(row=row, column=4, value=s.description[:120]).font = Font(name="Arial", size=9)
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=5, value=s.action_required[:120]).font = Font(name="Arial", size=9)
        ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=6, value=s.legal_basis[:80]).font = Font(name="Arial", size=8, color="666666")
        ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True)

        fill = auto_fill if s.auto_applied else action_fill
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = thin_border

        ws.row_dimensions[row].height = 45
    strat_end = row

    # Totals
    row += 1
    ws.cell(row=row, column=1, value="TOTAL ALL STRATEGIES").font = Font(name="Arial", bold=True, color=GOLD, size=11)
    ws.cell(row=row, column=2, value=f"=SUM(B{strat_start}:B{strat_end})").font = Font(name="Arial", bold=True, color=GOLD, size=12)
    ws.cell(row=row, column=2).number_format = gbp_format
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = Border(top=Side(style="double"))

    # === LEGEND ===
    row += 2
    ws.cell(row=row, column=1, value="KEY:").font = navy_font
    row += 1
    ws.cell(row=row, column=1, value="Green rows = Already applied by the Soup (automatic)").font = Font(name="Arial", size=10, color="006600")
    ws.cell(row=row, column=1).fill = auto_fill
    row += 1
    ws.cell(row=row, column=1, value="Yellow rows = Action needed from you (instructions in column E)").font = Font(name="Arial", size=10, color="CC6600")
    ws.cell(row=row, column=1).fill = action_fill

    # === MARRIAGE ALLOWANCE BACKDATE NOTE ===
    row += 2
    ws.cell(row=row, column=1, value="QUICK WINS — DO THIS TODAY").font = Font(name="Arial", bold=True, color=WHITE, size=12)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="006600")
    ws.merge_cells(f"A{row}:F{row}")

    quick_wins = [
        ("Marriage Allowance", "£252/year + up to £1,008 backdate", "gov.uk/marriage-allowance — Tina applies, takes 5 min", "FREE MONEY — zero risk"),
        ("Mileage Log", "Start recording business miles NOW", "Date, destination, purpose, miles — use a phone app", "Needed to claim 45p/mile (if better than actual costs)"),
        ("Home Office Hours", "Log hours worked from home each month", "Keep a simple diary/spreadsheet", "Unlocks £312-1,359/year deduction"),
        ("PPE Receipts", "Keep ALL receipts for safety gear", "Boots, hi-vis, hard hats, gloves", "100% allowable — often paid cash so invisible to bank"),
    ]
    for item in quick_wins:
        row += 1
        ws.cell(row=row, column=1, value=item[0]).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=row, column=2, value=item[1]).font = Font(name="Arial", size=10, color="006600")
        ws.cell(row=row, column=3, value=item[2]).font = Font(name="Arial", size=9)
        ws.cell(row=row, column=4, value=item[3]).font = Font(name="Arial", size=9, color="006600", italic=True)
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border

    print(f"  Tax Strategy: {len(savings)} strategies, "
          f"£{report['total_potential_saving']:,.0f} total potential saving")

# Also run for 2024/25 if data exists
data_2425 = year_data.get("2024/25")
if data_2425:
    # Add a summary section for 2024/25 below the 2025/26 strategies
    row += 3
    ws.cell(row=row, column=1, value="TAX YEAR 2024/25 — COMPARISON").font = Font(name="Arial", bold=True, color=WHITE, size=12)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(f"A{row}:F{row}")

    soup24 = data_2425["soup"]
    income24 = soup24.get_income_by_trade()
    expenses24 = soup24.get_sa103_summary()
    ti24 = sum(income24.values())
    te24 = sum(expenses24.values())
    np24 = ti24 - te24

    strat24 = TaxStrategy(
        net_profit=np24,
        total_income=ti24,
        total_expenses=te24,
        motor_expenses=expenses24.get("motor", 0),
        cis_deducted=0,  # No CIS in 2024/25
        cis_citb=0,
        drawings=0,
        spouse_income=0,
        business_miles_estimate=8_000,  # Partial year
        home_hours_per_month=60,
    )
    savings24 = strat24.run_all_strategies()
    report24 = strat24.get_strategy_report()

    row += 2
    ws.cell(row=row, column=1, value="Normal accountant:").font = Font(name="Arial", size=11, color="666666")
    write_currency(ws, row, 2, report24["naive_tax"], Font(name="Arial", size=12, color="CC0000", bold=True))
    row += 1
    ws.cell(row=row, column=1, value="With available optimisations:").font = Font(name="Arial", size=11, color=NAVY, bold=True)
    write_currency(ws, row, 2, report24["naive_tax"] - report24["additional_savings_available"],
                  Font(name="Arial", size=12, color="006600", bold=True))
    row += 1
    ws.cell(row=row, column=1, value="Potential saving:").font = Font(name="Arial", size=11, color=GOLD, bold=True)
    write_currency(ws, row, 2, report24["additional_savings_available"],
                  Font(name="Arial", size=12, color=GOLD, bold=True))

    print(f"  Tax Strategy 2024/25: £{report24['additional_savings_available']:,.0f} potential saving")


# ========================================================================
# SHEET 12: METACOGNITION ANALYSIS
# ========================================================================
ws = wb.create_sheet("Metacognition")
ws.sheet_properties.tabColor = "8B0000"

ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 70
ws.column_dimensions["F"].width = 40

ws.merge_cells("A1:F1")
ws["A1"] = "METACOGNITION ENGINE — DEEP TAX ANALYSIS"
ws["A1"].font = Font(name="Arial", bold=True, color=GOLD, size=16)
ws["A1"].fill = PatternFill("solid", fgColor="1B2A4A")

ws.merge_cells("A2:F2")
ws["A2"] = "Adapted from Aureon Sentience Architecture — every payment analysed through 9 reasoning layers"
ws["A2"].font = Font(name="Arial", color=WHITE, size=11)
ws["A2"].fill = PatternFill("solid", fgColor=BLUE)

# Run metacognition for 2025/26
data_2526 = year_data.get("2025/26")
if data_2526:
    soup = data_2526["soup"]
    income = soup.get_income_by_trade()
    expenses = soup.get_sa103_summary()
    total_income = sum(income.values())
    total_expenses = sum(expenses.values())
    net_profit = total_income - total_expenses
    kr = data_2526["kitchen_report"]

    # Calculate food P&L for loss relief
    food_income = income.get("food", 0)
    food_expenses = 0
    for r in data_2526["results"]:
        if not r.is_transfer and not r.is_income and r.trade == "food":
            food_expenses += abs(r.original.get("amount", 0))
    food_pnl = food_income - food_expenses

    meta = HNCMetacognition(tax_year=2025)
    meta_results = meta.analyse_full_position(
        net_profit=net_profit,
        total_income=total_income,
        total_expenses=total_expenses,
        motor_expenses=expenses.get("motor", 0),
        cis_deducted=kr.cis_deductions_credit,
        cis_citb=kr.cis_citb_expense,
        drawings=kr.drawings_total,
        food_pnl=food_pnl,
    )

    summary = meta.get_strategy_summary()

    # Headline
    row = 4
    ws.cell(row=row, column=1, value="METACOGNITION SUMMARY — 2025/26").font = Font(name="Arial", bold=True, color=WHITE, size=12)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(f"A{row}:F{row}")

    row = 6
    ws.cell(row=row, column=1, value="Total analyses performed:").font = normal_font
    ws.cell(row=row, column=2, value=len(meta_results)).font = Font(name="Arial", bold=True, size=11)

    row = 7
    ws.cell(row=row, column=1, value="Total thoughts generated:").font = normal_font
    total_thoughts = sum(len(r.thoughts) for r in meta_results)
    ws.cell(row=row, column=2, value=total_thoughts).font = Font(name="Arial", bold=True, size=11)

    row = 8
    ws.cell(row=row, column=1, value="Tax savings identified:").font = Font(name="Arial", bold=True, color=GOLD, size=11)
    total_meta_saving = sum(r.tax_saving for r in meta_results)
    write_currency(ws, row, 2, total_meta_saving, Font(name="Arial", bold=True, color=GOLD, size=14))

    # Each analysis area
    row = 10
    ws.cell(row=row, column=1, value="DETAILED ANALYSIS BY AREA").font = Font(name="Arial", bold=True, color=WHITE, size=12)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(f"A{row}:F{row}")

    row = 11
    meta_headers = ["Analysis Area", "Tax Saving", "Confidence", "Risk", "Key Insight", "Action Required"]
    for i, h in enumerate(meta_headers):
        c = ws.cell(row=row, column=i+1, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border

    conf_fills = {
        "high": PatternFill("solid", fgColor="E0FFE0"),
        "medium": PatternFill("solid", fgColor="FFFFCC"),
        "low": PatternFill("solid", fgColor="FFCCCC"),
    }

    for mr in meta_results:
        row += 1
        ws.cell(row=row, column=1, value=mr.subject).font = Font(name="Arial", bold=True, size=10)
        write_currency(ws, row, 2, mr.tax_saving,
                      Font(name="Arial", bold=True, size=10, color="006600") if mr.tax_saving > 0
                      else Font(name="Arial", size=10, color="666666"))
        ws.cell(row=row, column=3, value=f"{mr.confidence:.0%}").font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=row, column=4, value=mr.risk_level).font = normal_font

        # Get the best insight
        insights = [t for t in mr.thoughts if t.thought_type in ["INSIGHT", "ANALYSIS"]]
        best = insights[0].content[:120] if insights else ""
        ws.cell(row=row, column=5, value=best).font = Font(name="Arial", size=9)
        ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)

        # Get action items
        actions = [t.action for t in mr.thoughts if t.action]
        action_text = actions[0][:80] if actions else (mr.action_required[:80] if mr.action_required else "")
        ws.cell(row=row, column=6, value=action_text).font = Font(name="Arial", size=9)
        ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True)

        # Colour by confidence
        conf_level = "high" if mr.confidence >= 0.8 else ("medium" if mr.confidence >= 0.6 else "low")
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = conf_fills[conf_level]
            ws.cell(row=row, column=col).border = thin_border
        ws.row_dimensions[row].height = 50

    # Thought Stream
    row += 3
    ws.cell(row=row, column=1, value="FULL THOUGHT STREAM").font = Font(name="Arial", bold=True, color=WHITE, size=12)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="8B0000")
    ws.merge_cells(f"A{row}:F{row}")

    row += 1
    thought_headers = ["Area", "Type", "Thought", "Legal Basis", "Action", "Saving"]
    for i, h in enumerate(thought_headers):
        c = ws.cell(row=row, column=i+1, value=h)
        c.font = Font(name="Arial", bold=True, color=WHITE, size=9)
        c.fill = PatternFill("solid", fgColor="8B0000")
        c.border = thin_border

    type_colors = {
        "INSIGHT": PatternFill("solid", fgColor="E8F5E9"),
        "ANALYSIS": PatternFill("solid", fgColor="E3F2FD"),
        "CURIOSITY": PatternFill("solid", fgColor="FFF3E0"),
        "MEMORY": PatternFill("solid", fgColor="F3E5F5"),
        "DOUBT": PatternFill("solid", fgColor="FFEBEE"),
        "VETO": PatternFill("solid", fgColor="FFCDD2"),
        "CONVICTION": PatternFill("solid", fgColor="C8E6C9"),
        "REFLECTION": PatternFill("solid", fgColor="E0E0E0"),
        "OBSERVATION": PatternFill("solid", fgColor="F5F5F5"),
    }

    for mr in meta_results:
        for t in mr.thoughts:
            row += 1
            ws.cell(row=row, column=1, value=mr.subject[:25]).font = Font(name="Arial", size=8)
            ws.cell(row=row, column=2, value=t.thought_type).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=row, column=3, value=t.content[:120]).font = Font(name="Arial", size=8)
            ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=4, value=t.legal_basis[:60] if t.legal_basis else "").font = Font(name="Arial", size=8, color="666666")
            ws.cell(row=row, column=5, value=t.action[:60] if t.action else "").font = Font(name="Arial", size=8)
            if t.saving_impact > 0:
                write_currency(ws, row, 6, t.saving_impact, Font(name="Arial", size=8, color="006600"))
            fill = type_colors.get(t.thought_type, PatternFill())
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = fill
                ws.cell(row=row, column=col).border = thin_border
            ws.row_dimensions[row].height = 35

    # Aureon Architecture Credit
    row += 3
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1,
            value="Architecture: Adapted from Aureon Sentience Engine — "
                  "Inner Dialogue, Curiosity Loop, Reflection Loop, Conscience Gate, "
                  "Cascade Amplification, Memory/Mirrors"
    ).font = Font(name="Arial", size=9, color="666666", italic=True)

    print(f"  Metacognition: {len(meta_results)} analyses, "
          f"{total_thoughts} thoughts, £{total_meta_saving:,.0f} savings identified")


# ========================================================================
# SHEET 13: INTELLIGENCE REGISTRY — Government Moves & Our Counter-Moves
# ========================================================================
ws = wb.create_sheet("Intelligence")
ws.sheet_properties.tabColor = "4A0080"

ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 65
ws.column_dimensions["F"].width = 50

ws.merge_cells("A1:F1")
ws["A1"] = "INTELLIGENCE REGISTRY — GOVERNMENT MOVES vs OUR COUNTER-MOVES"
ws["A1"].font = Font(name="Arial", bold=True, color=GOLD, size=16)
ws["A1"].fill = PatternFill("solid", fgColor="4A0080")

ws.merge_cells("A2:F2")
ws["A2"] = "Adapted from Aureon Unified Intelligence Registry — 3 systems, 20+ signals, real-time policy tracking"
ws["A2"].font = Font(name="Arial", color=WHITE, size=11)
ws["A2"].fill = PatternFill("solid", fgColor="6A00B0")

data_2526 = year_data.get("2025/26")
if data_2526:
    soup = data_2526["soup"]
    income = soup.get_income_by_trade()
    expenses = soup.get_sa103_summary()
    total_income = sum(income.values())
    total_expenses = sum(expenses.values())
    net_profit = total_income - total_expenses
    kr = data_2526["kitchen_report"]

    registry = HNCIntelligenceRegistry(
        net_profit=net_profit,
        total_income=total_income,
        total_expenses=total_expenses,
        motor_expenses=expenses.get("motor", 0),
        cis_deducted=kr.cis_deductions_credit,
        months_elapsed=12,
        has_spouse=True,
    )
    verdict = registry.run_all_systems()

    # === VERDICT HEADLINE ===
    row = 4
    ws.cell(row=row, column=1, value="UNIFIED VERDICT — 2025/26").font = Font(name="Arial", bold=True, color=WHITE, size=12)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="4A0080")
    ws.merge_cells(f"A{row}:F{row}")

    row = 6
    items = [
        ("Total intelligence signals:", verdict.total_signals, ""),
        ("Total tax saving identified:", verdict.total_tax_saving, gbp_format),
        ("Auto-applied (already done):", verdict.auto_applied_saving, gbp_format),
        ("Action required (do this):", verdict.action_required_saving, gbp_format),
        ("Confidence-weighted saving:", verdict.confidence_weighted_saving, gbp_format),
        ("Risk assessment:", 0, ""),
    ]
    for label, val, fmt in items:
        ws.cell(row=row, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        if fmt == gbp_format and val > 0:
            write_currency(ws, row, 2, val, Font(name="Arial", bold=True, size=12, color=GOLD))
        elif isinstance(val, int) and val > 0:
            ws.cell(row=row, column=2, value=val).font = Font(name="Arial", bold=True, size=12)
        row += 1
    # Risk assessment text
    ws.cell(row=row-1, column=2, value=verdict.risk_assessment).font = Font(
        name="Arial", bold=True, size=10,
        color="CC0000" if "HIGH" in verdict.risk_assessment else "CC6600")

    # === GOVERNMENT MOVES ===
    row += 2
    ws.cell(row=row, column=1, value="GOVERNMENT MOVES (POLICY CHANGES)").font = Font(name="Arial", bold=True, color=WHITE, size=12)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="CC0000")
    ws.merge_cells(f"A{row}:F{row}")

    row += 1
    gov_headers = ["Source", "Impact", "Risk", "Priority", "What They Did", "Our Counter-Move"]
    for i, h in enumerate(gov_headers):
        c = ws.cell(row=row, column=i+1, value=h)
        c.font = header_font
        c.fill = PatternFill("solid", fgColor="CC0000")
        c.border = thin_border

    policy_signals = [s for s in verdict.signals if s.signal_type in ["POLICY", "WARNING"]]
    for s in sorted(policy_signals, key=lambda x: x.priority, reverse=True):
        row += 1
        ws.cell(row=row, column=1, value=s.source[:20]).font = Font(name="Arial", size=9)
        if s.tax_impact > 0:
            write_currency(ws, row, 2, s.tax_impact, Font(name="Arial", size=9, color="006600"))
        else:
            ws.cell(row=row, column=2, value="—").font = Font(name="Arial", size=9, color="666666")
        ws.cell(row=row, column=3, value=s.risk_level).font = Font(name="Arial", size=9,
            color="CC0000" if s.risk_level in ["HIGH", "MEDIUM"] else "006600")
        ws.cell(row=row, column=4, value=s.priority).font = Font(name="Arial", size=9, bold=True)
        ws.cell(row=row, column=5, value=s.description[:120]).font = Font(name="Arial", size=8)
        ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=6, value=s.action_required[:100]).font = Font(name="Arial", size=8)
        ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True)

        if s.risk_level in ["HIGH", "MEDIUM"]:
            fill = PatternFill("solid", fgColor="FFCCCC")
        elif "POSITIVE" in s.description:
            fill = PatternFill("solid", fgColor="E0FFE0")
        else:
            fill = PatternFill("solid", fgColor="FFFFCC")
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = thin_border
        ws.row_dimensions[row].height = 40

    # === TAX WAR KILL LIST ===
    row += 3
    ws.cell(row=row, column=1, value="TAX WAR STRATEGY — KILL LIST").font = Font(name="Arial", bold=True, color=WHITE, size=12)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="006600")
    ws.merge_cells(f"A{row}:F{row}")

    row += 1
    kill_headers = ["Strategy", "Tax Saving", "Confidence", "Priority", "Description", "Action"]
    for i, h in enumerate(kill_headers):
        c = ws.cell(row=row, column=i+1, value=h)
        c.font = Font(name="Arial", bold=True, color=WHITE, size=9)
        c.fill = PatternFill("solid", fgColor="006600")
        c.border = thin_border

    kill_signals = [s for s in verdict.signals
                   if s.signal_type in ["CREDIT", "DEDUCTION"] and s.tax_impact > 0]
    kill_start = row + 1
    for s in sorted(kill_signals, key=lambda x: x.tax_impact, reverse=True):
        row += 1
        ws.cell(row=row, column=1, value=s.source[:20]).font = Font(name="Arial", bold=True, size=9)
        write_currency(ws, row, 2, s.tax_impact, Font(name="Arial", bold=True, size=10, color="006600"))
        ws.cell(row=row, column=3, value=f"{s.confidence:.0%}").font = Font(name="Arial", size=9)
        ws.cell(row=row, column=4, value=s.priority).font = Font(name="Arial", size=9, bold=True)
        ws.cell(row=row, column=5, value=s.description[:120]).font = Font(name="Arial", size=8)
        ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=6, value=s.action_required[:100]).font = Font(name="Arial", size=8)
        ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True)

        fill = PatternFill("solid", fgColor="E0FFE0") if s.auto_apply else PatternFill("solid", fgColor="FFFFCC")
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = thin_border
        ws.row_dimensions[row].height = 40
    kill_end = row

    # Totals
    row += 1
    ws.cell(row=row, column=1, value="TOTAL KILL LIST").font = Font(name="Arial", bold=True, color=GOLD, size=11)
    ws.cell(row=row, column=2, value=f"=SUM(B{kill_start}:B{kill_end})").font = Font(name="Arial", bold=True, color=GOLD, size=12)
    ws.cell(row=row, column=2).number_format = gbp_format
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = Border(top=Side(style="double"))

    # === FORECAST ===
    row += 3
    ws.cell(row=row, column=1, value="TAX RESERVE FORECAST").font = Font(name="Arial", bold=True, color=WHITE, size=12)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=BLUE)
    ws.merge_cells(f"A{row}:F{row}")

    forecast_signals = [s for s in verdict.signals if s.signal_type in ["FORECAST", "TIMING"]]
    for s in forecast_signals:
        row += 1
        ws.cell(row=row, column=1, value=s.source[:20]).font = Font(name="Arial", bold=True, size=9)
        ws.cell(row=row, column=2, value=s.description[:120]).font = Font(name="Arial", size=9)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"B{row}:E{row}")
        if s.action_required:
            ws.cell(row=row, column=6, value=s.action_required[:80]).font = Font(name="Arial", size=9, color="006600")
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border
        ws.row_dimensions[row].height = 35

    # === SYSTEM ARCHITECTURE CREDIT ===
    row += 3
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1,
            value="Architecture: Adapted from Aureon Unified Intelligence Registry — "
                  "BattlefieldIntel, War Strategy, Tax Reserve Forecaster, "
                  "CompoundKing Monte Carlo, DeepMoneyFlowAnalyzer"
    ).font = Font(name="Arial", size=9, color="666666", italic=True)

    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1,
            value="Legal authority: IRC v Duke of Westminster [1936] — "
                  "'Every man is entitled if he can to order his affairs so that the "
                  "tax attaching under the appropriate Acts is less than it otherwise would be.'"
    ).font = Font(name="Arial", size=9, color="4A0080", italic=True)

    print(f"  Intelligence: {verdict.total_signals} signals, "
          f"£{verdict.total_tax_saving:,.0f} total saving, "
          f"risk: {verdict.risk_assessment[:20]}")


# ========================================================================
# SAVE
# ========================================================================
output_dir = os.path.join(os.path.dirname(__file__), "..", "output", "final")
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, "Gary_Leckey_Full_Accounts_2024-2026.xlsx")
wb.save(output_path)
print(f"\n  Workbook saved to: {output_path}")
print(f"  Sheets: {wb.sheetnames}")
print("  Done.")
