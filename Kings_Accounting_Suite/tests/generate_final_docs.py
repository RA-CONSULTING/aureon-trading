"""
FINAL DOCUMENT GENERATION — Soup-Classified Real Data
======================================================
Generates the complete document pack using the Soup's intelligent
classification of Aureon Creator's Q1 2026 bank data.

Aureon Creator / Aureon Research — April 2026
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from core.hnc_import import HNCImportEngine
from core.hnc_soup import HNCSoup, SA103_CATEGORIES

# ========================================================================
# CONSTANTS
# ========================================================================

NAVY = "1B2A4A"
BLUE = "2E5090"
GOLD = "C49A2E"
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"

TITLE_FONT = Font(name="Arial", size=14, bold=True, color=WHITE)
HEADER_FONT = Font(name="Arial", size=10, bold=True, color=WHITE)
SUBHEAD_FONT = Font(name="Arial", size=10, bold=True, color=NAVY)
BODY_FONT = Font(name="Arial", size=10, color="333333")
MONEY_FONT = Font(name="Arial", size=10, color="333333")
TOTAL_FONT = Font(name="Arial", size=10, bold=True, color=NAVY)
BLUE_FONT = Font(name="Arial", size=10, color="0000FF")

NAVY_FILL = PatternFill("solid", fgColor=NAVY)
BLUE_FILL = PatternFill("solid", fgColor=BLUE)
GOLD_FILL = PatternFill("solid", fgColor=GOLD)
GREY_FILL = PatternFill("solid", fgColor=LIGHT_GREY)
WHITE_FILL = PatternFill("solid", fgColor=WHITE)

THIN_BORDER = Border(bottom=Side(style="thin", color="999999"))
TOTAL_BORDER = Border(
    top=Side(style="thin", color=NAVY),
    bottom=Side(style="double", color=NAVY),
)

GBP_FORMAT = '£#,##0.00;(£#,##0.00);"-"'
PCT_FORMAT = '0.0%'

# ========================================================================
# FILE HANDLING
# ========================================================================

SEARCH_DIRS = [
    os.path.join(os.path.dirname(__file__), "..", "..", "uploads"),
    "/sessions/upbeat-stoic-hamilton/mnt/uploads",
]

def find_csv_files():
    seen = set()
    csv_files = []
    for d in SEARCH_DIRS:
        if not os.path.isdir(d):
            continue
        for f in sorted(os.listdir(d)):
            if f.endswith(".csv") and "Statement" in f and f not in seen:
                seen.add(f)
                csv_files.append(os.path.join(d, f))
    csv_files.sort()
    return csv_files

def read_utf16_csv(filepath):
    for enc in ("utf-16", "utf-16-le", "utf-16-be", "utf-8-sig", "latin-1"):
        try:
            with open(filepath, "r", encoding=enc) as f:
                text = f.read()
            text = text.replace("\x00", "")
            if "Date" in text:
                return text
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise RuntimeError(f"Could not decode {filepath}")

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ========================================================================
# WORKBOOK 1: MANAGEMENT ACCOUNTS
# ========================================================================

def build_management_accounts(soup, output_path):
    wb = Workbook()

    # ---- Sheet 1: P&L ----
    ws = wb.active
    ws.title = "P&L"
    set_col_widths(ws, [40, 15, 15, 15, 15])

    # Title block
    ws.merge_cells("A1:E1")
    ws["A1"] = "Aureon Creator — Profit & Loss Account"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = NAVY_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:E2")
    ws["A2"] = "Quarter 1: 01 January 2026 — 31 March 2026"
    ws["A2"].font = Font(name="Arial", size=10, color=WHITE)
    ws["A2"].fill = BLUE_FILL
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:E3")
    ws["A3"] = "Prepared in accordance with FRS 102 Section 1A"
    ws["A3"].font = Font(name="Arial", size=8, italic=True, color="666666")
    ws["A3"].alignment = Alignment(horizontal="center")

    # Column headers
    headers = ["", "Construction", "Food (Food Venture)", "Consulting", "TOTAL"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = BLUE_FILL
        c.alignment = Alignment(horizontal="right" if i > 1 else "left")

    # Income
    income = soup.get_income_by_trade()
    r = 7
    ws.cell(row=r, column=1, value="TURNOVER").font = SUBHEAD_FONT
    r += 1
    ws.cell(row=r, column=1, value="Trading income").font = BODY_FONT
    ws.cell(row=r, column=2, value=income.get("construction", 0)).number_format = GBP_FORMAT
    ws.cell(row=r, column=3, value=income.get("food", 0)).number_format = GBP_FORMAT
    ws.cell(row=r, column=4, value=income.get("consulting", 0)).number_format = GBP_FORMAT
    ws.cell(row=r, column=5).value = f"=SUM(B{r}:D{r})"
    ws.cell(row=r, column=5).number_format = GBP_FORMAT
    ws.cell(row=r, column=5).font = TOTAL_FONT

    income_row = r
    r += 2

    # Expenses by SA103 box
    ws.cell(row=r, column=1, value="ALLOWABLE EXPENSES").font = SUBHEAD_FONT
    r += 1

    sa103 = soup.get_sa103_summary()
    expense_rows = []
    for cat, amount in sorted(sa103.items(), key=lambda x: -x[1]):
        info = SA103_CATEGORIES.get(cat, {})
        label = info.get("label", cat)
        box = info.get("box", "")
        ws.cell(row=r, column=1, value=f"{label} ({box})").font = BODY_FONT
        ws.cell(row=r, column=5, value=amount).number_format = GBP_FORMAT
        ws.cell(row=r, column=5).font = MONEY_FONT
        # Blue font for hardcoded inputs
        ws.cell(row=r, column=5).font = BLUE_FONT
        expense_rows.append(r)
        r += 1

    # Total expenses
    r += 1
    ws.cell(row=r, column=1, value="Total Allowable Expenses").font = TOTAL_FONT
    first_exp = expense_rows[0] if expense_rows else r
    last_exp = expense_rows[-1] if expense_rows else r
    ws.cell(row=r, column=5).value = f"=SUM(E{first_exp}:E{last_exp})"
    ws.cell(row=r, column=5).number_format = GBP_FORMAT
    ws.cell(row=r, column=5).font = TOTAL_FONT
    ws.cell(row=r, column=5).border = THIN_BORDER
    total_exp_row = r

    # Net profit
    r += 2
    ws.cell(row=r, column=1, value="NET PROFIT").font = Font(name="Arial", size=12, bold=True, color=NAVY)
    ws.cell(row=r, column=5).value = f"=E{income_row}-E{total_exp_row}"
    ws.cell(row=r, column=5).number_format = GBP_FORMAT
    ws.cell(row=r, column=5).font = Font(name="Arial", size=12, bold=True, color=NAVY)
    ws.cell(row=r, column=5).border = TOTAL_BORDER

    # Non-allowable note
    non_allowable = soup.get_non_allowable()
    if non_allowable:
        r += 2
        na_total = sum(x.original.get("amount", 0) for x in non_allowable)
        ws.cell(row=r, column=1,
                value=f"Note: £{na_total:,.2f} personal expenditure excluded (non-allowable)").font = Font(
            name="Arial", size=8, italic=True, color="999999")

    # ---- Sheet 2: Transaction Ledger ----
    ws2 = wb.create_sheet("Transaction Ledger")
    set_col_widths(ws2, [12, 5, 12, 35, 25, 8, 12])

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "Aureon Creator — Transaction Ledger Q1 2026"
    ws2["A1"].font = TITLE_FONT
    ws2["A1"].fill = NAVY_FILL

    headers2 = ["Date", "Dir", "Amount", "Description", "HMRC Category", "Box", "Trade"]
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(row=3, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = BLUE_FILL

    r = 4
    for sr in sorted(soup.results, key=lambda x: x.original.get("date", "")):
        t = sr.original
        if sr.is_transfer:
            continue  # Skip transfers from the ledger

        ws2.cell(row=r, column=1, value=t.get("date", "")).font = BODY_FONT
        ws2.cell(row=r, column=2, value=t.get("direction", "")).font = BODY_FONT
        ws2.cell(row=r, column=3, value=t.get("amount", 0)).number_format = GBP_FORMAT
        ws2.cell(row=r, column=3).font = MONEY_FONT
        ws2.cell(row=r, column=4, value=t.get("description", "")[:35]).font = BODY_FONT
        ws2.cell(row=r, column=5, value=sr.hmrc_label[:25]).font = BODY_FONT
        ws2.cell(row=r, column=6, value=sr.sa103_box or ("Income" if sr.is_income else "—")).font = BODY_FONT
        ws2.cell(row=r, column=7, value=sr.trade).font = BODY_FONT

        if r % 2 == 0:
            for col in range(1, 8):
                ws2.cell(row=r, column=col).fill = GREY_FILL
        r += 1

    # Totals
    r += 1
    ws2.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws2.cell(row=r, column=3).value = f"=SUM(C4:C{r-2})"
    ws2.cell(row=r, column=3).number_format = GBP_FORMAT
    ws2.cell(row=r, column=3).font = TOTAL_FONT
    ws2.cell(row=r, column=3).border = TOTAL_BORDER

    # ---- Sheet 3: SA103 Summary ----
    ws3 = wb.create_sheet("SA103 Summary")
    set_col_widths(ws3, [10, 45, 15])

    ws3.merge_cells("A1:C1")
    ws3["A1"] = "SA103S Self-Employment (Short) — Expense Summary"
    ws3["A1"].font = TITLE_FONT
    ws3["A1"].fill = NAVY_FILL

    ws3.merge_cells("A2:C2")
    ws3["A2"] = "Tax Year 2025/26 — Q1 Only (pro-rata)"
    ws3["A2"].font = Font(name="Arial", size=10, color=WHITE)
    ws3["A2"].fill = BLUE_FILL

    headers3 = ["Box", "Description", "Amount"]
    for i, h in enumerate(headers3, 1):
        c = ws3.cell(row=4, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = BLUE_FILL

    # Income
    r = 6
    total_income = sum(soup.get_income_by_trade().values())
    ws3.cell(row=r, column=1, value="Box 9").font = BODY_FONT
    ws3.cell(row=r, column=2, value="Turnover — the income from your business").font = BODY_FONT
    ws3.cell(row=r, column=3, value=total_income).number_format = GBP_FORMAT
    ws3.cell(row=r, column=3).font = BLUE_FONT
    turnover_row = r

    r += 2
    ws3.cell(row=r, column=1).font = SUBHEAD_FONT
    ws3.cell(row=r, column=2, value="Allowable business expenses").font = SUBHEAD_FONT
    r += 1

    exp_start = r
    for cat, amount in sorted(sa103.items(), key=lambda x: -x[1]):
        info = SA103_CATEGORIES.get(cat, {})
        ws3.cell(row=r, column=1, value=info.get("box", "")).font = BODY_FONT
        ws3.cell(row=r, column=2, value=info.get("label", cat)).font = BODY_FONT
        ws3.cell(row=r, column=3, value=amount).number_format = GBP_FORMAT
        ws3.cell(row=r, column=3).font = BLUE_FONT
        r += 1
    exp_end = r - 1

    r += 1
    ws3.cell(row=r, column=1, value="Box 20").font = TOTAL_FONT
    ws3.cell(row=r, column=2, value="Total allowable expenses").font = TOTAL_FONT
    ws3.cell(row=r, column=3).value = f"=SUM(C{exp_start}:C{exp_end})"
    ws3.cell(row=r, column=3).number_format = GBP_FORMAT
    ws3.cell(row=r, column=3).font = TOTAL_FONT
    ws3.cell(row=r, column=3).border = THIN_BORDER
    total_exp_row2 = r

    r += 2
    ws3.cell(row=r, column=1, value="Box 21").font = Font(name="Arial", size=12, bold=True, color=NAVY)
    ws3.cell(row=r, column=2, value="Net profit (loss)").font = Font(name="Arial", size=12, bold=True, color=NAVY)
    ws3.cell(row=r, column=3).value = f"=C{turnover_row}-C{total_exp_row2}"
    ws3.cell(row=r, column=3).number_format = GBP_FORMAT
    ws3.cell(row=r, column=3).font = Font(name="Arial", size=12, bold=True, color=NAVY)
    ws3.cell(row=r, column=3).border = TOTAL_BORDER

    # ---- Sheet 4: Tax Estimate ----
    ws4 = wb.create_sheet("Tax Estimate")
    set_col_widths(ws4, [35, 15, 20])

    ws4.merge_cells("A1:C1")
    ws4["A1"] = "Estimated Tax Position — 2025/26"
    ws4["A1"].font = TITLE_FONT
    ws4["A1"].fill = NAVY_FILL

    ws4.merge_cells("A2:C2")
    ws4["A2"] = "Based on Q1 data annualised (×4)"
    ws4["A2"].font = Font(name="Arial", size=10, color=WHITE, italic=True)
    ws4["A2"].fill = BLUE_FILL

    data_rows = [
        (4, "Q1 Net Profit", f"='{ws3.title}'!C{r}", None),
        (5, "Annualised Profit (×4)", "=B4*4", None),
        (6, "Personal Allowance", 12570, "Source: HMRC 2025/26"),
        (7, "Taxable Income", "=MAX(0,B5-B6)", None),
        (9, "Income Tax (Basic Rate 20%)", "=MIN(B7,37700)*0.2", None),
        (10, "Income Tax (Higher Rate 40%)", "=MAX(0,B7-37700)*0.4", None),
        (11, "Total Income Tax", "=B9+B10", None),
        (13, "NI Class 2 (annual)", 179.40, "Source: HMRC 2025/26 £3.45/wk"),
        (14, "NI Class 4 (6% on profits £12,570-£50,270)", "=MAX(0,MIN(B5,50270)-12570)*0.06", None),
        (15, "NI Class 4 (2% above £50,270)", "=MAX(0,B5-50270)*0.02", None),
        (16, "Total NI", "=SUM(B13:B15)", None),
        (18, "TOTAL ESTIMATED TAX", "=B11+B16", None),
    ]

    for row_num, label, value, comment in data_rows:
        ws4.cell(row=row_num, column=1, value=label).font = BODY_FONT
        c = ws4.cell(row=row_num, column=2, value=value)
        c.number_format = GBP_FORMAT

        if isinstance(value, str) and value.startswith("="):
            c.font = Font(name="Arial", size=10, color="000000")
        else:
            c.font = BLUE_FONT

        if comment:
            ws4.cell(row=row_num, column=3, value=comment).font = Font(
                name="Arial", size=8, italic=True, color="999999")

    # Bold the total
    ws4.cell(row=18, column=1).font = Font(name="Arial", size=12, bold=True, color=NAVY)
    ws4.cell(row=18, column=2).font = Font(name="Arial", size=12, bold=True, color=NAVY)
    ws4.cell(row=18, column=2).border = TOTAL_BORDER

    # VAT check
    ws4.cell(row=20, column=1, value="VAT THRESHOLD CHECK").font = SUBHEAD_FONT
    ws4.cell(row=21, column=1, value="Estimated Annual Turnover").font = BODY_FONT
    ws4.cell(row=21, column=2).value = f"='{ws3.title}'!C{turnover_row}*4"
    ws4.cell(row=21, column=2).number_format = GBP_FORMAT
    ws4.cell(row=22, column=1, value="VAT Registration Threshold").font = BODY_FONT
    ws4.cell(row=22, column=2, value=90000).number_format = GBP_FORMAT
    ws4.cell(row=22, column=2).font = BLUE_FONT
    ws4.cell(row=22, column=3, value="Source: HMRC 2025/26").font = Font(
        name="Arial", size=8, italic=True, color="999999")
    ws4.cell(row=23, column=1, value="Headroom").font = BODY_FONT
    ws4.cell(row=23, column=2).value = "=B22-B21"
    ws4.cell(row=23, column=2).number_format = GBP_FORMAT

    # Save
    wb.save(output_path)
    print(f"  [OK] Management Accounts: {output_path}")
    return output_path


# ========================================================================
# MAIN
# ========================================================================

def main():
    print("=" * 70)
    print("  HNC FINAL DOCUMENT GENERATION")
    print("  Aureon Creator | Q1 2026 | Soup-Classified")
    print("=" * 70)

    # Import data
    csv_files = find_csv_files()
    importer = HNCImportEngine()
    for filepath in csv_files:
        csv_text = read_utf16_csv(filepath)
        importer.import_csv_string(csv_text, os.path.basename(filepath))

    all_txns = importer.get_bank_transactions()
    print(f"\n  Imported {len(all_txns)} transactions")

    # Run through Soup
    soup = HNCSoup(entity_name="Aureon Creator", trades=["construction", "food", "consulting"])
    soup.classify_all(all_txns)

    # Quick summary
    income = soup.get_income_by_trade()
    sa103 = soup.get_sa103_summary()
    total_income = sum(income.values())
    total_expenses = sum(sa103.values())
    print(f"  Income:   £{total_income:>10,.2f}")
    print(f"  Expenses: £{total_expenses:>10,.2f}")
    print(f"  Profit:   £{total_income - total_expenses:>10,.2f}")

    # Output directory
    output_dir = os.path.join(os.path.dirname(__file__), "..", "output", "final")
    os.makedirs(output_dir, exist_ok=True)

    # Generate workbook
    xlsx_path = os.path.join(output_dir, "Gary_Leckey_Q1_2026_Accounts.xlsx")
    build_management_accounts(soup, xlsx_path)

    # Recalculate formulas
    recalc_script = os.path.join(os.path.dirname(__file__), "..", "..", ".claude", "skills", "xlsx", "scripts", "recalc.py")
    if os.path.exists(recalc_script):
        print(f"\n  Recalculating formulas...")
        import subprocess
        result = subprocess.run(
            [sys.executable, recalc_script, xlsx_path],
            capture_output=True, text=True, timeout=60
        )
        print(f"  {result.stdout[:200]}")
        if result.returncode != 0:
            print(f"  [WARN] Recalc: {result.stderr[:200]}")
    else:
        print(f"  [INFO] Recalc script not found — formulas will calculate on open")

    # List output
    print(f"\n  Output: {output_dir}")
    for f in sorted(os.listdir(output_dir)):
        fpath = os.path.join(output_dir, f)
        print(f"    {f}  ({os.path.getsize(fpath):,} bytes)")

    print(f"\n{'='*70}")
    print(f"  DONE — Documents ready")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
