"""Wave-based raw-data scan and confirmation layer for Aureon accounting.

This module turns the accounting intake into an auditable swarm pass:

1. sweep every file in the raw-data roots, not only supported accounting files;
2. run provider and evidence lookups;
3. compare the sweep with the normalised bank feed and evidence inventory;
4. build a Phi-cadenced consensus score and checklist;
5. write machine and human artifacts for the end user and the organism.

It is local-only. It does not file with Companies House, submit to HMRC, make
payments, create external evidence, or touch trading/exchange state.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import shutil
import sys
import textwrap
import time
import zipfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Sequence
from xml.etree import ElementTree as ET


KAS_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = KAS_DIR.parent
TOOLS_DIR = KAS_DIR / "tools"
sys.path.insert(0, str(REPO_ROOT))
sys.path.insert(0, str(KAS_DIR))
sys.path.insert(0, str(TOOLS_DIR))

from company_house_tax_audit import (  # noqa: E402
    DEFAULT_COMPANY_NAME,
    DEFAULT_COMPANY_NUMBER,
    DEFAULT_PERIOD_END,
    DEFAULT_PERIOD_START,
)
from company_raw_data_intake import (  # noqa: E402
    RAW_EVIDENCE_SUFFIXES,
    SKIP_DIR_NAMES,
    build_company_raw_data_manifest,
    classify_raw_data_file,
    is_skipped,
    resolve_raw_roots,
)


SAFE_BOUNDARIES = {
    "official_companies_house_filing": "manual_only",
    "official_hmrc_submission": "manual_only",
    "tax_or_penalty_payment": "manual_only",
    "external_receipt_or_invoice_creation": "blocked",
    "exchange_or_trading_mutation": "blocked_from_accounting_swarm_scan",
}

PHI = (1.0 + math.sqrt(5.0)) / 2.0
PHI_MINOR = 1.0 / PHI
SCHEMA_VERSION = "accounting-swarm-raw-data-wave-scan-v1"

PROVIDER_PATTERNS = {
    "zempler": ("zempler", "statement__gbp", "business_gbp_monthly"),
    "revolut": ("revolut", "transactions report", "account-statement", "statement_", "msagm9ys"),
    "sumup": ("sumup",),
    "hmrc": ("hmrc", "government gateway", "corporation tax", "ct600"),
    "companies_house": ("companies house", "company house", "confirmation statement", "cs01"),
}

FILING_REQUIREMENT_KEYWORDS = {
    "companies_house_annual_accounts": ("accounts", "balance sheet", "directors report", "company house", "companies house"),
    "hmrc_company_tax_return": ("ct600", "corporation tax", "hmrc", "tax computation"),
    "confirmation_statement": ("confirmation statement", "cs01", "officer", "psc", "shareholder"),
    "source_evidence": ("receipt", "invoice", "statement", "bank", "ledger", "expense", "sales", "cash"),
}

EXTERNAL_MUTATION_WORDS = (
    "submit",
    "payment",
    "pay tax",
    "file with companies house",
    "hmrc api",
    "place order",
    "kraken order",
)

READABLE_STATUSES = {
    "readable_text",
    "readable_csv",
    "readable_json",
    "readable_workbook",
    "readable_pdf_text",
    "readable_docx_text",
    "readable_archive_index",
    "readable_image_ocr",
    "readable_empty_file",
}


def apply_safe_environment() -> None:
    os.environ["AUREON_ACCOUNTING_ON_DEMAND"] = "1"
    os.environ["AUREON_DISABLE_GOVERNMENT_SUBMISSION"] = "1"
    os.environ["AUREON_DISABLE_REAL_ORDERS"] = os.environ.get("AUREON_DISABLE_REAL_ORDERS", "1")
    os.environ["AUREON_LIVE_TRADING"] = os.environ.get("AUREON_LIVE_TRADING", "0")


def run_accounting_swarm_data_wave_scan(
    *,
    repo_root: str | Path = REPO_ROOT,
    company_name: str = DEFAULT_COMPANY_NAME,
    company_number: str = DEFAULT_COMPANY_NUMBER,
    period_start: str = DEFAULT_PERIOD_START,
    period_end: str = DEFAULT_PERIOD_END,
    raw_data_roots: Sequence[str | Path] | None = None,
    include_default_roots: bool = True,
    output_dir: str | Path | None = None,
) -> dict[str, Any]:
    """Run the local raw-data swarm scan and write JSON/Markdown artifacts."""
    apply_safe_environment()
    root = Path(repo_root).resolve()
    out_dir = Path(output_dir) if output_dir else (
        root
        / "Kings_Accounting_Suite"
        / "output"
        / "end_user_accounts"
        / company_number
        / f"{period_start}_to_{period_end}"
    )
    out_dir.mkdir(parents=True, exist_ok=True)

    benchmark: list[dict[str, Any]] = []
    started = time.perf_counter()

    wave_start = time.perf_counter()
    raw_roots = resolve_raw_roots(root, raw_data_roots, include_default_roots=include_default_roots)
    source_inventory = discover_raw_source_inventory(raw_roots, root)
    files = source_inventory["included_paths"]
    filesystem_records = [filesystem_record(path, root, raw_roots) for path in files]
    record_benchmark(
        benchmark,
        "wave_1_filesystem_discovery",
        wave_start,
        file_count=len(files),
        excluded_file_count=len(source_inventory["closure_records"]) - len(files),
        byte_count=sum(item["bytes"] for item in filesystem_records),
    )

    wave_start = time.perf_counter()
    readability_records = [file_readability_record(path, root) for path in files]
    readability_by_path = {
        normalise_path(item.get("path")): item
        for item in readability_records
    }
    filesystem_records = [
        {
            **item,
            "readability": readability_by_path.get(normalise_path(item.get("path")), {}),
        }
        for item in filesystem_records
    ]
    record_benchmark(
        benchmark,
        "wave_1b_file_readability_extraction",
        wave_start,
        file_count=len(files),
        byte_count=sum(item["bytes"] for item in filesystem_records),
    )

    wave_start = time.perf_counter()
    lookup_records = [add_lookup_hints(item, path) for item, path in zip(filesystem_records, files)]
    record_benchmark(benchmark, "wave_2_provider_lookup", wave_start, file_count=len(files))

    wave_start = time.perf_counter()
    raw_manifest = build_company_raw_data_manifest(
        root,
        company_number=company_number,
        period_start=period_start,
        period_end=period_end,
        raw_roots=raw_data_roots,
        include_default_roots=include_default_roots,
    ).to_dict()
    combined_summary = load_combined_bank_summary(root, company_number, period_start, period_end)
    combined_summary["payer_provenance_summary"] = load_payer_provenance_summary(root, company_number, period_start, period_end)
    ingestion_records = add_ingestion_matches(lookup_records, raw_manifest, combined_summary)
    record_benchmark(
        benchmark,
        "wave_3_ingestion_match",
        wave_start,
        file_count=len(files),
        transaction_sources=combined_summary.get("transaction_source_count", combined_summary.get("csv_source_count", 0)),
    )

    wave_start = time.perf_counter()
    folding = build_evidence_folding(ingestion_records, combined_summary)
    record_benchmark(benchmark, "wave_4_evidence_folding", wave_start, file_count=len(files))

    wave_start = time.perf_counter()
    consensus = build_phi_swarm_consensus(ingestion_records, folding, combined_summary)
    readability_summary = summarise_file_readability(readability_records)
    source_closure = summarise_source_closure(
        source_inventory["closure_records"],
        readability_summary,
        raw_roots,
        source_inventory["missing_roots"],
    )
    checklist = build_internal_logic_chain_checklist(
        company_number=company_number,
        period_start=period_start,
        period_end=period_end,
        raw_roots=raw_roots,
        records=ingestion_records,
        combined_summary=combined_summary,
        folding=folding,
        consensus=consensus,
        source_closure=source_closure,
    )
    record_benchmark(benchmark, "wave_5_phi_swarm_consensus", wave_start, file_count=len(files))

    total_duration = max(time.perf_counter() - started, 0.000001)
    benchmark_summary = {
        "total_duration_seconds": round(total_duration, 6),
        "files_scanned": len(files),
        "bytes_scanned": sum(item["bytes"] for item in filesystem_records),
        "files_per_second": round(len(files) / total_duration, 3),
        "waves": benchmark,
    }

    manifest_status = "completed"
    if source_closure.get("status") != "closed_no_silent_skips":
        manifest_status = "blocked_no_data_skip_contract"
    if (
        int(readability_summary.get("file_count", 0) or 0) == 0
        or int(readability_summary.get("readable_file_count", 0) or 0) != int(readability_summary.get("file_count", 0) or 0)
        or int(readability_summary.get("metadata_only_file_count", 0) or 0) > 0
        or int(readability_summary.get("blocked_file_count", 0) or 0) > 0
    ):
        manifest_status = "blocked_no_data_skip_contract"

    manifest = {
        "schema_version": SCHEMA_VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "status": manifest_status,
        "company_name": company_name,
        "company_number": company_number,
        "period_start": period_start,
        "period_end": period_end,
        "repo_root": str(root),
        "raw_roots": [str(path) for path in raw_roots],
        "source_closure": source_closure,
        "source_closure_records": source_inventory["closure_records"],
        "safe_boundaries": dict(SAFE_BOUNDARIES),
        "phi_bridge": {
            "phi": PHI,
            "phi_minor": PHI_MINOR,
            "cadence": "deterministic_accounting_wave_scan",
            "external_peer_required": False,
        },
        "swarm_agents": build_swarm_agent_roster(),
        "waves": {
            "filesystem_discovery": summarise_filesystem_records(filesystem_records),
            "source_closure": source_closure,
            "file_readability": readability_summary,
            "provider_lookup": summarise_lookup_records(lookup_records),
            "payer_provenance_lookup": combined_summary.get("payer_provenance_summary") or {},
            "ingestion_match": summarise_ingestion_records(ingestion_records),
            "evidence_folding": folding,
            "phi_swarm_consensus": consensus,
        },
        "file_readability": readability_summary,
        "file_readability_records": readability_records,
        "files": ingestion_records,
        "raw_data_intake_summary": raw_manifest.get("summary") or {},
        "combined_bank_data": combined_summary,
        "payer_provenance_summary": combined_summary.get("payer_provenance_summary") or {},
        "internal_logic_chain_checklist": checklist,
        "end_user_confirmation": build_end_user_confirmation_summary(
            records=ingestion_records,
            combined_summary=combined_summary,
            checklist=checklist,
            consensus=consensus,
            manifest_status=manifest_status,
        ),
        "benchmark": benchmark_summary,
        "manual_next_step": (
            "Review the confirmation and checklist, then manually approve, sign, upload, enter, submit, and pay through official channels."
        ),
    }

    unsafe_hits = detect_external_mutation_words(manifest)
    if unsafe_hits:
        manifest["status"] = "blocked_unsafe_external_mutation_language"
        manifest["unsafe_hits"] = unsafe_hits

    source_closure_outputs = write_source_closure_artifacts(manifest, output_dir=out_dir)
    readability_outputs = write_file_readability_artifacts(manifest, output_dir=out_dir)
    json_path = out_dir / "swarm_raw_data_wave_scan.json"
    md_path = out_dir / "swarm_raw_data_wave_scan.md"
    json_path.write_text(json.dumps(manifest, indent=2, sort_keys=True, default=str), encoding="utf-8")
    md_path.write_text(render_swarm_wave_scan_markdown(manifest), encoding="utf-8")
    manifest["outputs"] = {
        "swarm_raw_data_wave_scan_json": str(json_path),
        "swarm_raw_data_wave_scan_markdown": str(md_path),
        **source_closure_outputs,
        **readability_outputs,
    }
    json_path.write_text(json.dumps(manifest, indent=2, sort_keys=True, default=str), encoding="utf-8")
    return manifest


def discover_all_raw_files(raw_roots: Sequence[Path]) -> list[Path]:
    return discover_raw_source_inventory(raw_roots, REPO_ROOT)["included_paths"]


def discover_raw_source_inventory(raw_roots: Sequence[Path], repo_root: Path) -> dict[str, Any]:
    included: list[Path] = []
    records: list[dict[str, Any]] = []
    seen: set[str] = set()
    missing_roots: list[str] = []
    for raw_root in raw_roots:
        if not raw_root.exists():
            missing_roots.append(str(raw_root))
            records.append(
                {
                    "decision": "missing_root",
                    "reason": "raw_data_root_not_found",
                    "path": str(raw_root),
                    "repo_relative_path": relative_to_repo(raw_root, repo_root),
                    "raw_root": str(raw_root),
                    "bytes": 0,
                    "suffix": "",
                    "autonomous_next_action": "Create the folder or correct the raw-data root path before calling the pack complete.",
                }
            )
            continue
        for path in sorted(raw_root.rglob("*"), key=lambda item: str(item).lower()):
            if not path.is_file():
                continue
            key = str(path.resolve()).lower()
            if is_skipped(path):
                records.append(source_closure_record(path, repo_root, raw_roots, "excluded", skip_reason(path)))
                continue
            if key in seen:
                records.append(
                    source_closure_record(
                        path,
                        repo_root,
                        raw_roots,
                        "excluded",
                        "duplicate_path_already_scanned_from_overlapping_root",
                    )
                )
                continue
            seen.add(key)
            included.append(path)
            records.append(source_closure_record(path, repo_root, raw_roots, "included", "included_for_readability_and_accounting_wave"))
    return {
        "included_paths": included,
        "closure_records": records,
        "missing_roots": missing_roots,
    }


def skip_reason(path: Path) -> str:
    matched = [part for part in path.parts if part in SKIP_DIR_NAMES]
    if matched:
        return f"excluded_generated_cache_or_runtime_folder:{matched[0]}"
    return "excluded_by_accounting_raw_data_skip_rule"


def source_closure_record(
    path: Path,
    repo_root: Path,
    raw_roots: Sequence[Path],
    decision: str,
    reason: str,
) -> dict[str, Any]:
    try:
        stat = path.stat()
        size = stat.st_size
        modified_at = datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat()
    except Exception:
        size = 0
        modified_at = ""
    raw_root_label = ""
    relative_to_root = ""
    for raw_root in raw_roots:
        try:
            relative_to_root = str(path.relative_to(raw_root)).replace("\\", "/")
            raw_root_label = str(raw_root)
            break
        except ValueError:
            continue
    action = "Read this file and fold it into accounting evidence."
    if decision == "excluded":
        action = "Excluded from accounting source ingestion with a recorded reason; move it out of generated/cache/runtime folders if it is real evidence."
    return {
        "decision": decision,
        "reason": reason,
        "path": str(path),
        "repo_relative_path": relative_to_repo(path, repo_root),
        "raw_root": raw_root_label,
        "relative_to_raw_root": relative_to_root,
        "suffix": path.suffix.lower(),
        "bytes": size,
        "modified_at": modified_at,
        "autonomous_next_action": action,
    }


def filesystem_record(path: Path, repo_root: Path, raw_roots: Sequence[Path]) -> dict[str, Any]:
    stat = path.stat()
    root_label = ""
    rel_to_root = ""
    for raw_root in raw_roots:
        try:
            rel_to_root = str(path.relative_to(raw_root)).replace("\\", "/")
            root_label = str(raw_root)
            break
        except ValueError:
            continue
    try:
        repo_rel = str(path.relative_to(repo_root)).replace("\\", "/")
    except ValueError:
        repo_rel = str(path)
    return {
        "path": str(path),
        "repo_relative_path": repo_rel,
        "raw_root": root_label,
        "relative_to_raw_root": rel_to_root,
        "name": path.name,
        "suffix": path.suffix.lower(),
        "bytes": stat.st_size,
        "modified_at": datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat(),
        "sha256": sha256_file(path),
        "supported_by_standard_intake": path.suffix.lower() in RAW_EVIDENCE_SUFFIXES,
    }


def add_lookup_hints(record: dict[str, Any], path: Path) -> dict[str, Any]:
    out = dict(record)
    readability = out.get("readability") or {}
    sample = str(readability.get("sample_text") or "") or content_probe(path)
    text = " ".join(
        str(value).lower()
        for value in (
            out.get("name"),
            out.get("repo_relative_path"),
            out.get("relative_to_raw_root"),
            sample,
        )
    )
    provider = "unknown"
    for candidate, patterns in PROVIDER_PATTERNS.items():
        if any(pattern in text for pattern in patterns):
            provider = candidate
            break

    if out["supported_by_standard_intake"]:
        category, intake_provider, ingestion_path, evidence_role, warnings = classify_raw_data_file(path)
        if provider == "unknown" and intake_provider:
            provider = intake_provider
    else:
        category, ingestion_path, evidence_role, warnings = classify_nonstandard_file(path, text)

    out.update(
        {
            "provider_hint": provider,
            "category_hint": category,
            "ingestion_path_hint": ingestion_path,
            "evidence_role_hint": evidence_role,
            "lookup_warnings": warnings,
            "file_read_status": readability.get("status", "unknown"),
            "readable_by_aureon": bool(readability.get("readable_by_aureon")),
            "readability_action": readability.get("autonomous_next_action", ""),
            "extracted_chars": int(readability.get("chars_extracted") or 0),
            "extracted_rows": int(readability.get("rows_extracted") or 0),
            "lookup_agents": [
                "raw_lookup_agent",
                "provider_lookup_agent",
                "evidence_role_agent",
                "file_readability_agent",
            ],
        }
    )
    return out


def classify_nonstandard_file(path: Path, text: str) -> tuple[str, str, str, list[str]]:
    suffix = path.suffix.lower()
    warnings = ["unsupported_suffix_catalogued_by_swarm"]
    if any(word in text for word in ("bank", "statement", "transaction", "ledger")):
        return "nonstandard_financial_data", "manual_review_or_converter_needed", "financial_source_candidate", warnings
    if any(word in text for word in ("invoice", "receipt", "supplier", "expense")):
        return "nonstandard_evidence", "evidence_review", "supporting_evidence", warnings
    if suffix in {".zip", ".7z", ".rar"}:
        return "archive_needs_unpacking", "manual_review_or_unpack", "raw_archive", warnings
    return "unclassified_raw_file", "catalogue_only", "raw_context", warnings


def add_ingestion_matches(
    records: list[dict[str, Any]],
    raw_manifest: dict[str, Any],
    combined_summary: dict[str, Any],
) -> list[dict[str, Any]]:
    intake_paths = {normalise_path(item.get("path")) for item in raw_manifest.get("files") or []}
    source_map: dict[str, dict[str, Any]] = {}
    evidence_paths: set[str] = set()
    for key in ("csv_sources", "pdf_sources"):
        for item in combined_summary.get(key) or []:
            path_key = normalise_path(item.get("path"))
            if path_key:
                source_map[path_key] = item
    for item in combined_summary.get("evidence_files") or []:
        if isinstance(item, dict):
            path_key = normalise_path(item.get("path"))
        else:
            path_key = normalise_path(item)
        if path_key:
            evidence_paths.add(path_key)

    out: list[dict[str, Any]] = []
    for record in records:
        path_key = normalise_path(record.get("path"))
        source = source_map.get(path_key) or {}
        intake_seen = path_key in intake_paths
        combined_seen = bool(source)
        evidence_seen = path_key in evidence_paths
        if combined_seen and int(source.get("rows_imported", source.get("rows_in_period", 0)) or 0) > 0:
            fold_status = "folded_transaction_feed"
        elif combined_seen:
            fold_status = "seen_transaction_source_outside_period"
        elif evidence_seen:
            fold_status = "folded_evidence_inventory"
        elif intake_seen:
            fold_status = "catalogued_for_review"
        else:
            fold_status = "catalogued_by_swarm_only"
        enriched = dict(record)
        enriched.update(
            {
                "standard_intake_seen": intake_seen,
                "combined_feed_seen": combined_seen,
                "evidence_inventory_seen": evidence_seen,
                "fold_status": fold_status,
                "rows_read": int(source.get("rows_read") or 0),
                "rows_in_period": int(source.get("rows_in_period") or 0),
                "rows_imported": int(source.get("rows_imported") or 0),
                "source_account": source.get("source_account") or "",
                "source_kind": source.get("source_kind") or "",
                "ingestion_agents": [
                    "ingestion_match_agent",
                    "dedupe_evidence_agent",
                    "source_account_agent",
                ],
            }
        )
        out.append(enriched)
    return out


def build_evidence_folding(records: list[dict[str, Any]], combined_summary: dict[str, Any]) -> dict[str, Any]:
    by_status = Counter(item.get("fold_status", "unknown") for item in records)
    by_requirement: dict[str, list[str]] = {key: [] for key in FILING_REQUIREMENT_KEYWORDS}
    for item in records:
        haystack = " ".join(
            str(item.get(key, "")).lower()
            for key in ("name", "repo_relative_path", "category_hint", "evidence_role_hint", "ingestion_path_hint")
        )
        for requirement, keywords in FILING_REQUIREMENT_KEYWORDS.items():
            if any(keyword in haystack for keyword in keywords):
                by_requirement[requirement].append(item.get("repo_relative_path") or item.get("path") or "")

    folded_count = sum(
        count
        for status, count in by_status.items()
        if status in {"folded_transaction_feed", "seen_transaction_source_outside_period", "folded_evidence_inventory", "catalogued_for_review"}
    )
    file_count = len(records)
    coverage_ratio = folded_count / file_count if file_count else 1.0
    return {
        "agents": [
            "evidence_folding_agent",
            "filing_requirement_agent",
            "source_evidence_agent",
        ],
        "fold_status_counts": dict(sorted(by_status.items())),
        "folded_or_catalogued_count": folded_count,
        "file_count": file_count,
        "coverage_ratio": round(coverage_ratio, 6),
        "requirement_file_counts": {
            key: len(value)
            for key, value in sorted(by_requirement.items())
        },
        "requirement_examples": {
            key: value[:8]
            for key, value in sorted(by_requirement.items())
            if value
        },
        "combined_unique_rows": combined_summary.get("unique_rows_in_period", 0),
        "duplicate_rows_removed": combined_summary.get("duplicate_rows_removed", 0),
    }


def build_phi_swarm_consensus(
    records: list[dict[str, Any]],
    folding: dict[str, Any],
    combined_summary: dict[str, Any],
) -> dict[str, Any]:
    file_count = max(len(records), 1)
    standard_seen = sum(1 for item in records if item.get("standard_intake_seen"))
    provider_known = sum(1 for item in records if item.get("provider_hint") != "unknown")
    readable_files = sum(1 for item in records if item.get("readable_by_aureon"))
    transaction_sources = int(combined_summary.get("transaction_source_count", combined_summary.get("csv_source_count", 0)) or 0)
    unique_rows = int(combined_summary.get("unique_rows_in_period") or 0)
    folded_ratio = float(folding.get("coverage_ratio") or 0.0)
    standard_ratio = standard_seen / file_count
    provider_ratio = provider_known / file_count
    readability_ratio = readable_files / file_count
    transaction_score = min(1.0, transaction_sources / max(1.0, file_count * PHI_MINOR))
    row_score = 1.0 if unique_rows > 0 else 0.0
    score = (
        folded_ratio * 0.27
        + standard_ratio * 0.18
        + provider_ratio * 0.12
        + readability_ratio * 0.15
        + transaction_score * 0.18
        + row_score * 0.10
    )
    score = max(0.0, min(1.0, score))
    if score >= PHI_MINOR and folded_ratio >= PHI_MINOR and unique_rows > 0:
        status = "coherent_ready_for_end_user_confirmation"
    elif unique_rows > 0:
        status = "partial_needs_review"
    else:
        status = "blocked_no_transaction_rows"
    return {
        "agents": [
            "phi_bridge_agent",
            "swarm_consensus_agent",
            "accountant_logic_agent",
            "payer_lookup_agent",
            "cis_vat_tax_basis_agent",
            "end_user_observance_agent",
        ],
        "status": status,
        "score": round(score, 6),
        "threshold": round(PHI_MINOR, 6),
        "folded_ratio": round(folded_ratio, 6),
        "standard_intake_ratio": round(standard_ratio, 6),
        "provider_known_ratio": round(provider_ratio, 6),
        "readability_ratio": round(readability_ratio, 6),
        "transaction_score": round(transaction_score, 6),
        "unique_rows_in_period": unique_rows,
        "decision": (
            "The accounting organism has enough swept, catalogued, and ingested local evidence to present an end-user pack."
            if status == "coherent_ready_for_end_user_confirmation"
            else "The accounting organism must keep attention on the listed gaps before the user relies on the pack."
        ),
    }


def build_internal_logic_chain_checklist(
    *,
    company_number: str,
    period_start: str,
    period_end: str,
    raw_roots: Sequence[Path],
    records: list[dict[str, Any]],
    combined_summary: dict[str, Any],
    folding: dict[str, Any],
    consensus: dict[str, Any],
    source_closure: dict[str, Any],
) -> list[dict[str, Any]]:
    transaction_sources = int(combined_summary.get("transaction_source_count", combined_summary.get("csv_source_count", 0)) or 0)
    unique_rows = int(combined_summary.get("unique_rows_in_period") or 0)
    known_providers = sorted({item.get("provider_hint") for item in records if item.get("provider_hint") != "unknown"})
    swarm_only = [item for item in records if item.get("fold_status") == "catalogued_by_swarm_only"]
    unreadable = [item for item in records if not item.get("readable_by_aureon")]
    source_closed = source_closure.get("status") == "closed_no_silent_skips"
    payer_summary = combined_summary.get("payer_provenance_summary") or {}
    return [
        checklist_item(
            "resolve_raw_roots",
            True,
            f"{len(raw_roots)} roots resolved for {company_number} {period_start} to {period_end}",
            [str(path) for path in raw_roots],
        ),
        checklist_item(
            "sweep_every_file",
            True,
            f"{len(records)} files swept across raw roots, including unsupported suffixes.",
            [item.get("repo_relative_path", "") for item in records[:10]],
        ),
        checklist_item(
            "no_silent_source_skips",
            source_closed,
            (
                f"{source_closure.get('included_file_count', 0)} files included; "
                f"{source_closure.get('excluded_file_count', 0)} files explicitly excluded with reasons; "
                f"{source_closure.get('missing_root_count', 0)} missing roots; "
                f"{source_closure.get('unaccounted_file_count', 0)} unaccounted."
            ),
            [item.get("path", "") for item in source_closure.get("blocked_examples", [])],
            status="complete" if source_closed else "blocked",
            next_action=(
                "Fix missing raw roots or unresolved source decisions before generating final-ready accounts."
                if not source_closed
                else "Every candidate source file is either included or explicitly excluded with a reason."
            ),
        ),
        checklist_item(
            "read_or_account_for_every_file",
            not unreadable and bool(records),
            f"{len(records) - len(unreadable)} files content-extracted; {len(unreadable)} files need OCR/converter/manual read.",
            [
                f"{item.get('repo_relative_path', '')}: {item.get('file_read_status')} -> {item.get('readability_action')}"
                for item in unreadable[:12]
            ],
            status="blocked" if unreadable or not records else "complete",
            next_action=(
                "Run OCR/converters or provide text/CSV exports for the listed files, then rerun the workflow."
                if unreadable
                else "All swept files were content-readable by Aureon."
            ),
        ),
        checklist_item(
            "provider_lookup",
            bool(known_providers),
            f"Providers detected: {', '.join(known_providers) or 'none'}",
            known_providers,
        ),
        checklist_item(
            "combined_bank_feed",
            transaction_sources > 0 and unique_rows > 0,
            f"{transaction_sources} transaction sources, {unique_rows} unique period rows.",
            [combined_summary.get("combined_csv_path", "")],
        ),
        checklist_item(
            "duplicate_overlap_control",
            "duplicate_rows_removed" in combined_summary,
            f"duplicate_rows_removed={combined_summary.get('duplicate_rows_removed', 'unknown')}",
            [],
        ),
        checklist_item(
            "payer_provenance_lookup_control",
            bool(payer_summary) and bool(payer_summary.get("lookup_required_plus_not_required_equals_total")),
            (
                f"{payer_summary.get('incoming_rows_total', 0)} incoming rows; "
                f"{payer_summary.get('lookup_required_count', 0)} lookup-required; "
                f"{payer_summary.get('lookup_not_required_count', 0)} not-required."
            ),
            [],
            status="complete" if payer_summary.get("lookup_required_plus_not_required_equals_total") else "attention",
            next_action=(
                "Review payer provenance manifest for lookup-required rows and CIS/VAT status."
                if payer_summary
                else "Run the statutory pack so payer provenance manifest is available to the swarm scan."
            ),
        ),
        checklist_item(
            "evidence_folding",
            float(folding.get("coverage_ratio") or 0.0) >= PHI_MINOR,
            f"folded/catalogued={folding.get('folded_or_catalogued_count', 0)}/{folding.get('file_count', 0)} ratio={folding.get('coverage_ratio')}",
            [],
        ),
        checklist_item(
            "swarm_only_attention_queue",
            True,
            f"{len(swarm_only)} files were not recognised by the standard intake and are listed for review.",
            [item.get("repo_relative_path", "") for item in swarm_only[:12]],
            status="attention" if swarm_only else "complete",
            next_action="Review or add converters for files in the swarm-only queue." if swarm_only else "No swarm-only raw files found.",
        ),
        checklist_item(
            "phi_swarm_consensus",
            consensus.get("status") == "coherent_ready_for_end_user_confirmation",
            f"score={consensus.get('score')} threshold={consensus.get('threshold')} status={consensus.get('status')}",
            [],
        ),
        checklist_item(
            "manual_filing_boundary",
            True,
            "Companies House, HMRC, tax payments, and penalties remain manual end-user actions.",
            [],
        ),
    ]


def checklist_item(
    item_id: str,
    ok: bool,
    evidence: str,
    examples: list[str],
    *,
    status: str | None = None,
    next_action: str = "",
) -> dict[str, Any]:
    return {
        "id": item_id,
        "status": status or ("complete" if ok else "blocked"),
        "ok": bool(ok),
        "evidence": evidence,
        "examples": [example for example in examples if example][:12],
        "next_action": next_action or ("Continue." if ok else "Resolve before relying on the generated pack."),
    }


def build_end_user_confirmation_summary(
    *,
    records: list[dict[str, Any]],
    combined_summary: dict[str, Any],
    checklist: list[dict[str, Any]],
    consensus: dict[str, Any],
    manifest_status: str = "completed",
) -> dict[str, Any]:
    manual_items = [
        "Review and approve the generated accounts and tax computation.",
        "Sign any required director/accountant declarations.",
        "Upload or enter Companies House and HMRC information manually.",
        "Submit and pay manually through official channels.",
    ]
    readability_attention = next(
        (item for item in checklist if item.get("id") == "read_or_account_for_every_file"),
        {},
    )
    source_closure_attention = next(
        (item for item in checklist if item.get("id") == "no_silent_source_skips"),
        {},
    )
    ready = (
        manifest_status == "completed"
        and consensus.get("status") == "coherent_ready_for_end_user_confirmation"
        and not [item for item in checklist if item.get("status") == "blocked"]
    )
    return {
        "status": "ready" if ready else "needs_review",
        "what_aureon_confirmed": [
            f"Raw-data swarm scanned {len(records)} files.",
            str(source_closure_attention.get("evidence") or "No-silent-skip source closure completed."),
            str(readability_attention.get("evidence") or "File readability audit completed."),
            f"Combined feed has {combined_summary.get('transaction_source_count', combined_summary.get('csv_source_count', 0))} transaction sources.",
            f"Combined feed has {combined_summary.get('unique_rows_in_period', 0)} unique period rows.",
            f"Duplicate overlap control removed {combined_summary.get('duplicate_rows_removed', 0)} rows.",
            f"Phi swarm consensus status is {consensus.get('status')} with score {consensus.get('score')}.",
        ],
        "human_actions_left": manual_items,
        "attention_items": [item for item in checklist if item.get("status") in {"attention", "blocked"}],
        "no_external_actions_taken": [
            "No Companies House filing was submitted.",
            "No HMRC return was submitted.",
            "No tax, penalty, or filing fee was paid.",
            "No exchange or trading order was placed.",
        ],
    }


def write_logic_chain_artifacts(
    manifest: dict[str, Any],
    *,
    output_dir: str | Path,
) -> dict[str, str]:
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    checklist = manifest.get("internal_logic_chain_checklist") or []
    payload = {
        "schema_version": "accounting-internal-logic-chain-checklist-v1",
        "generated_at": manifest.get("generated_at"),
        "company_number": manifest.get("company_number"),
        "period_start": manifest.get("period_start"),
        "period_end": manifest.get("period_end"),
        "safe_boundaries": manifest.get("safe_boundaries"),
        "checklist": checklist,
        "consensus": (manifest.get("waves") or {}).get("phi_swarm_consensus") or {},
        "benchmark": manifest.get("benchmark") or {},
    }
    json_path = out_dir / "internal_logic_chain_checklist.json"
    md_path = out_dir / "internal_logic_chain_checklist.md"
    json_path.write_text(json.dumps(payload, indent=2, sort_keys=True, default=str), encoding="utf-8")
    md_path.write_text(render_logic_chain_markdown(payload), encoding="utf-8")
    return {
        "internal_logic_chain_checklist_json": str(json_path),
        "internal_logic_chain_checklist_markdown": str(md_path),
    }


def write_end_user_confirmation_artifacts(
    manifest: dict[str, Any],
    *,
    output_dir: str | Path,
) -> dict[str, str]:
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    confirmation = manifest.get("end_user_confirmation") or {}
    payload = {
        "schema_version": "end-user-accounting-confirmation-v1",
        "generated_at": manifest.get("generated_at"),
        "company_name": manifest.get("company_name"),
        "company_number": manifest.get("company_number"),
        "period_start": manifest.get("period_start"),
        "period_end": manifest.get("period_end"),
        "status": confirmation.get("status", "unknown"),
        "confirmation": confirmation,
        "outputs": manifest.get("outputs") or {},
        "safe_boundaries": manifest.get("safe_boundaries") or {},
        "manual_next_step": manifest.get("manual_next_step"),
    }
    json_path = out_dir / "END_USER_CONFIRMATION.json"
    md_path = out_dir / "END_USER_CONFIRMATION.md"
    json_path.write_text(json.dumps(payload, indent=2, sort_keys=True, default=str), encoding="utf-8")
    md_path.write_text(render_end_user_confirmation_markdown(payload), encoding="utf-8")
    return {
        "end_user_confirmation_json": str(json_path),
        "end_user_confirmation_markdown": str(md_path),
    }


def write_file_readability_artifacts(
    manifest: dict[str, Any],
    *,
    output_dir: str | Path,
) -> dict[str, str]:
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    records = manifest.get("file_readability_records") or []
    summary = manifest.get("file_readability") or {}
    payload = {
        "schema_version": "accounting-all-source-file-readability-audit-v1",
        "generated_at": manifest.get("generated_at"),
        "company_name": manifest.get("company_name"),
        "company_number": manifest.get("company_number"),
        "period_start": manifest.get("period_start"),
        "period_end": manifest.get("period_end"),
        "summary": summary,
        "records": records,
        "safe_boundaries": manifest.get("safe_boundaries") or {},
    }
    json_path = out_dir / "all_source_file_readability_audit.json"
    csv_path = out_dir / "all_source_file_readability_audit.csv"
    md_path = out_dir / "all_source_file_readability_audit.md"
    pdf_path = out_dir / "all_source_file_readability_audit.pdf"
    json_path.write_text(json.dumps(payload, indent=2, sort_keys=True, default=str), encoding="utf-8")
    write_file_readability_csv(records, csv_path)
    md_text = render_file_readability_markdown(payload)
    md_path.write_text(md_text, encoding="utf-8")
    write_readability_pdf(pdf_path, "All Source File Readability Audit", md_text)
    return {
        "all_source_file_readability_audit_json": str(json_path),
        "all_source_file_readability_audit_csv": str(csv_path),
        "all_source_file_readability_audit_markdown": str(md_path),
        "all_source_file_readability_audit_pdf": str(pdf_path),
    }


def write_source_closure_artifacts(
    manifest: dict[str, Any],
    *,
    output_dir: str | Path,
) -> dict[str, str]:
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "schema_version": "accounting-no-data-skipped-source-closure-v1",
        "generated_at": manifest.get("generated_at"),
        "company_name": manifest.get("company_name"),
        "company_number": manifest.get("company_number"),
        "period_start": manifest.get("period_start"),
        "period_end": manifest.get("period_end"),
        "summary": manifest.get("source_closure") or {},
        "records": manifest.get("source_closure_records") or [],
        "safe_boundaries": manifest.get("safe_boundaries") or {},
    }
    json_path = out_dir / "no_data_skipped_source_closure_audit.json"
    csv_path = out_dir / "no_data_skipped_source_closure_audit.csv"
    md_path = out_dir / "no_data_skipped_source_closure_audit.md"
    pdf_path = out_dir / "no_data_skipped_source_closure_audit.pdf"
    json_path.write_text(json.dumps(payload, indent=2, sort_keys=True, default=str), encoding="utf-8")
    write_source_closure_csv(payload.get("records") or [], csv_path)
    md_text = render_source_closure_markdown(payload)
    md_path.write_text(md_text, encoding="utf-8")
    write_readability_pdf(pdf_path, "No Data Skipped Source Closure Audit", md_text)
    return {
        "no_data_skipped_source_closure_audit_json": str(json_path),
        "no_data_skipped_source_closure_audit_csv": str(csv_path),
        "no_data_skipped_source_closure_audit_markdown": str(md_path),
        "no_data_skipped_source_closure_audit_pdf": str(pdf_path),
    }


def write_source_closure_csv(records: Sequence[dict[str, Any]], path: Path) -> None:
    fields = [
        "decision",
        "reason",
        "repo_relative_path",
        "raw_root",
        "relative_to_raw_root",
        "suffix",
        "bytes",
        "autonomous_next_action",
    ]
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        for record in records:
            writer.writerow({field: record.get(field, "") for field in fields})


def render_source_closure_markdown(payload: dict[str, Any]) -> str:
    summary = payload.get("summary") or {}
    lines = [
        "# No Data Skipped Source Closure Audit",
        "",
        f"- Company: `{payload.get('company_number')}` {payload.get('company_name')}",
        f"- Period: `{payload.get('period_start')}` to `{payload.get('period_end')}`",
        f"- Status: `{summary.get('status', 'unknown')}`",
        f"- Candidate files considered: `{summary.get('candidate_file_count', 0)}`",
        f"- Included for reading/accounting: `{summary.get('included_file_count', 0)}`",
        f"- Explicitly excluded with reason: `{summary.get('excluded_file_count', 0)}`",
        f"- Missing raw roots: `{summary.get('missing_root_count', 0)}`",
        f"- Unaccounted files: `{summary.get('unaccounted_file_count', 0)}`",
        f"- All candidate decisions recorded: `{summary.get('all_candidate_decisions_recorded', False)}`",
        "",
        "## Decision Counts",
        "",
    ]
    for decision, count in (summary.get("decision_counts") or {}).items():
        lines.append(f"- `{decision}`: {count}")
    lines.extend(["", "## Exclusion Reasons", ""])
    reasons = summary.get("reason_counts") or {}
    if not reasons:
        lines.append("- None.")
    for reason, count in reasons.items():
        lines.append(f"- `{reason}`: {count}")
    lines.extend(["", "## Blocking Examples", ""])
    blocked = summary.get("blocked_examples") or []
    if not blocked:
        lines.append("- None.")
    for item in blocked:
        lines.append(f"- `{item.get('path')}`: {item.get('reason')} - {item.get('next_action')}")
    lines.extend(["", "## Every Source Decision", ""])
    for record in payload.get("records") or []:
        lines.append(
            f"- `{record.get('decision')}` `{record.get('repo_relative_path') or record.get('path')}`: "
            f"{record.get('reason')}"
        )
    lines.extend(
        [
            "",
            "## Safety",
            "",
            "- This closure audit reads and classifies local accounting source files only.",
            "- Generated/cache/runtime folders are only excluded when the exclusion is recorded here.",
            "- Missing raw roots or unreadable source files block final-ready status.",
            "- It does not submit to Companies House or HMRC, pay tax, or mutate trading/exchange state.",
            "",
        ]
    )
    return "\n".join(lines)


def write_file_readability_csv(records: Sequence[dict[str, Any]], path: Path) -> None:
    fields = [
        "repo_relative_path",
        "suffix",
        "bytes",
        "status",
        "extractor",
        "readable_by_aureon",
        "chars_extracted",
        "rows_extracted",
        "pages_extracted",
        "sheets_extracted",
        "autonomous_next_action",
        "errors",
    ]
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        for record in records:
            row = {field: record.get(field, "") for field in fields}
            row["errors"] = "; ".join(str(item) for item in record.get("errors") or [])
            writer.writerow(row)


def render_file_readability_markdown(payload: dict[str, Any]) -> str:
    summary = payload.get("summary") or {}
    lines = [
        "# All Source File Readability Audit",
        "",
        f"- Company: `{payload.get('company_number')}` {payload.get('company_name')}",
        f"- Period: `{payload.get('period_start')}` to `{payload.get('period_end')}`",
        f"- Files accounted for: `{summary.get('accounted_file_count', 0)}` / `{summary.get('file_count', 0)}`",
        f"- Content-readable files: `{summary.get('readable_file_count', 0)}`",
        f"- Metadata-only files: `{summary.get('metadata_only_file_count', 0)}`",
        f"- Blocked/converter-needed files: `{summary.get('blocked_file_count', 0)}`",
        f"- Readability ratio: `{summary.get('readability_ratio', 0)}`",
        "",
        "## Status Counts",
        "",
    ]
    for status, count in (summary.get("status_counts") or {}).items():
        lines.append(f"- `{status}`: {count}")
    lines.extend(["", "## Files Needing OCR, Converter, Or Manual Read", ""])
    blocked = summary.get("blocked_examples") or []
    if not blocked:
        lines.append("- None.")
    for item in blocked:
        lines.append(f"- `{item.get('path')}`: {item.get('status')} - {item.get('next_action')}")
    lines.extend(["", "## Every File", ""])
    for record in payload.get("records") or []:
        lines.append(
            f"- `{record.get('repo_relative_path')}`: {record.get('status')} "
            f"via {record.get('extractor') or 'none'}; rows={record.get('rows_extracted', 0)} "
            f"pages={record.get('pages_extracted', 0)} sheets={record.get('sheets_extracted', 0)}"
        )
    lines.extend(
        [
            "",
            "## Safety",
            "",
            "- This audit reads local files only.",
            "- It does not invent missing evidence.",
            "- It does not submit to Companies House or HMRC.",
            "- It does not pay tax or mutate trading/exchange state.",
            "",
        ]
    )
    return "\n".join(lines)


def write_readability_pdf(path: Path, title: str, markdown_text: str) -> bool:
    try:
        from pdf_markdown_renderer import render_markdown_pdf
    except Exception:
        return False
    return render_markdown_pdf(path, title, markdown_text)


def load_combined_bank_summary(
    repo_root: Path,
    company_number: str,
    period_start: str,
    period_end: str,
) -> dict[str, Any]:
    period_manifest = (
        repo_root
        / "Kings_Accounting_Suite"
        / "output"
        / "gateway"
        / f"{period_start}_to_{period_end}"
        / "period_pack_manifest.json"
    )
    full_manifest = (
        repo_root
        / "Kings_Accounting_Suite"
        / "output"
        / "company_compliance"
        / company_number
        / "full_company_accounts_run_manifest.json"
    )
    for path in (period_manifest, full_manifest):
        data = read_json(path)
        if not data:
            continue
        if "combined_bank_data" in data:
            summary = dict(data.get("combined_bank_data") or {})
        else:
            summary = dict(((data.get("source_data_inventory") or {}).get("combined_bank_data") or {}))
        if summary:
            summary.setdefault("combined_csv_path", str(
                repo_root
                / "Kings_Accounting_Suite"
                / "output"
                / "gateway"
                / f"{period_start}_to_{period_end}"
                / f"combined_bank_transactions_{period_start}_to_{period_end}.csv"
            ))
            return summary
    return {}


def load_payer_provenance_summary(
    repo_root: Path,
    company_number: str,
    period_start: str,
    period_end: str,
) -> dict[str, Any]:
    statutory_manifest = (
        repo_root
        / "Kings_Accounting_Suite"
        / "output"
        / "statutory"
        / company_number
        / f"{period_start}_to_{period_end}"
        / "statutory_filing_pack_manifest.json"
    )
    data = read_json(statutory_manifest)
    enrichment = data.get("accounting_report_enrichment") or {}
    summary = enrichment.get("payer_provenance_summary") or {}
    if summary:
        return dict(summary)
    manifest = data.get("payer_provenance_manifest") or {}
    return dict(manifest.get("summary") or {})


def read_json(path: Path) -> dict[str, Any]:
    try:
        if not path.exists():
            return {}
        data = json.loads(path.read_text(encoding="utf-8", errors="replace"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    try:
        with path.open("rb") as fh:
            for chunk in iter(lambda: fh.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()
    except Exception:
        return ""


def file_readability_record(path: Path, repo_root: Path, *, sample_limit: int = 1200) -> dict[str, Any]:
    """Attempt safe local extraction for one raw accounting file."""
    suffix = path.suffix.lower()
    base = {
        "path": str(path),
        "repo_relative_path": relative_to_repo(path, repo_root),
        "name": path.name,
        "suffix": suffix,
        "bytes": path.stat().st_size if path.exists() else 0,
        "status": "blocked_read_error",
        "extractor": "",
        "readable_by_aureon": False,
        "chars_extracted": 0,
        "rows_extracted": 0,
        "pages_extracted": 0,
        "sheets_extracted": 0,
        "sample_text": "",
        "errors": [],
        "autonomous_next_action": "Review the file manually or add a converter.",
    }
    if base["bytes"] == 0:
        base.update(
            {
                "status": "readable_empty_file",
                "extractor": "filesystem",
                "readable_by_aureon": True,
                "autonomous_next_action": "No data to extract; keep file as empty-file evidence.",
            }
        )
        return base

    try:
        if suffix in {".txt", ".md", ".html", ".htm", ".xml", ".ofx", ".qif", ".log", ".custom", ".weird"}:
            text = path.read_text(encoding="utf-8", errors="replace")
            return readable_record(base, "readable_text", "text", text, sample_limit)
        if suffix == ".csv":
            return read_csv_file_record(path, base, sample_limit)
        if suffix == ".json":
            text = path.read_text(encoding="utf-8", errors="replace")
            parsed = json.loads(text)
            keys = json_key_sample(parsed)
            record = readable_record(base, "readable_json", "json", text, sample_limit)
            record["json_key_sample"] = keys
            return record
        if suffix in {".xlsx", ".xlsm"}:
            return read_xlsx_file_record(path, base, sample_limit)
        if suffix == ".xls":
            base.update(
                {
                    "status": "unsupported_legacy_xls_requires_converter",
                    "extractor": "legacy_xls",
                    "errors": ["Install/use an XLS converter or save as XLSX/CSV for autonomous extraction."],
                    "autonomous_next_action": "Convert legacy XLS to XLSX or CSV, then rerun the accounts workflow.",
                }
            )
            return base
        if suffix == ".pdf":
            return read_pdf_file_record(path, base, sample_limit)
        if suffix in {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"}:
            return read_image_file_record(path, base, sample_limit)
        if suffix == ".docx":
            return read_docx_file_record(path, base, sample_limit)
        if suffix == ".zip":
            return read_zip_file_record(path, base, sample_limit)
    except Exception as exc:
        base["errors"] = [f"{type(exc).__name__}: {exc}"]
        base["autonomous_next_action"] = "Fix the file or add a converter, then rerun the accounts workflow."
        return base

    raw_probe = path.read_bytes()[:sample_limit]
    text = raw_probe.decode("utf-8", errors="ignore")
    if text.strip() and looks_like_text(raw_probe, text):
        return readable_record(base, "readable_text", "binary_text_probe", text, sample_limit)
    base.update(
        {
            "status": "unsupported_binary_requires_converter",
            "extractor": "none",
            "errors": [f"No autonomous extractor registered for suffix {suffix or '<none>'}."],
            "autonomous_next_action": "Add a converter/OCR/parser for this file type or provide a CSV/PDF/text export.",
        }
    )
    return base


def looks_like_text(raw: bytes, text: str) -> bool:
    if b"\x00" in raw:
        return False
    if not text:
        return False
    printable = sum(1 for char in text if char.isprintable() or char.isspace())
    return printable / max(len(text), 1) >= 0.85


def readable_record(base: dict[str, Any], status: str, extractor: str, text: str, sample_limit: int) -> dict[str, Any]:
    clean = " ".join(str(text or "").replace("\x00", " ").split())
    base.update(
        {
            "status": status,
            "extractor": extractor,
            "readable_by_aureon": status in READABLE_STATUSES,
            "chars_extracted": len(text or ""),
            "sample_text": clean[:sample_limit],
            "autonomous_next_action": "Use extracted text/tables as accounting context and source evidence.",
        }
    )
    return base


def configure_tesseract_command(pytesseract_module: Any) -> str:
    """Point pytesseract at a local Windows Tesseract install when PATH is stale."""
    candidates = [
        os.environ.get("TESSERACT_CMD", ""),
        os.environ.get("TESSERACT_EXE", ""),
        shutil.which("tesseract") or "",
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        str(Path.home() / "AppData" / "Local" / "Programs" / "Tesseract-OCR" / "tesseract.exe"),
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            pytesseract_module.pytesseract.tesseract_cmd = str(Path(candidate))
            return str(Path(candidate))
    return ""


def read_csv_file_record(path: Path, base: dict[str, Any], sample_limit: int) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(text.splitlines(), dialect=dialect))
    header = rows[0] if rows else []
    preview = "\n".join(",".join(row[:12]) for row in rows[:8])
    record = readable_record(base, "readable_csv", "csv", preview or text, sample_limit)
    record.update(
        {
            "rows_extracted": max(len(rows) - 1, 0),
            "headers": header[:30],
            "autonomous_next_action": "Feed rows into bank/ledger import if transaction-like, otherwise keep as schedule evidence.",
        }
    )
    return record


def read_xlsx_file_record(path: Path, base: dict[str, Any], sample_limit: int) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        base.update(
            {
                "status": "blocked_missing_dependency",
                "extractor": "openpyxl",
                "errors": [f"openpyxl unavailable: {exc}"],
                "autonomous_next_action": "Install openpyxl or export workbook as CSV.",
            }
        )
        return base
    workbook = load_workbook(path, read_only=True, data_only=True)
    chunks: list[str] = []
    row_count = 0
    for sheet in workbook.worksheets:
        chunks.append(f"Sheet: {sheet.title}")
        for row in sheet.iter_rows(max_row=25, values_only=True):
            values = ["" if value is None else str(value) for value in row[:20]]
            if any(values):
                chunks.append(" | ".join(values))
                row_count += 1
    record = readable_record(base, "readable_workbook", "openpyxl", "\n".join(chunks), sample_limit)
    record.update(
        {
            "rows_extracted": row_count,
            "sheets_extracted": len(workbook.worksheets),
            "sheet_names": [sheet.title for sheet in workbook.worksheets],
            "autonomous_next_action": "Use workbook sheets as ledger/schedule evidence; convert transaction sheets to CSV if needed.",
        }
    )
    return record


def read_pdf_file_record(path: Path, base: dict[str, Any], sample_limit: int) -> dict[str, Any]:
    try:
        from pypdf import PdfReader
    except Exception as exc:
        base.update(
            {
                "status": "blocked_missing_dependency",
                "extractor": "pypdf",
                "errors": [f"pypdf unavailable: {exc}"],
                "autonomous_next_action": "Install pypdf or provide text/CSV export.",
            }
        )
        return base
    reader = PdfReader(str(path))
    chunks = [(page.extract_text() or "") for page in reader.pages]
    text = "\n".join(chunks)
    if text.strip():
        record = readable_record(base, "readable_pdf_text", "pypdf", text, sample_limit)
        record["pages_extracted"] = len(reader.pages)
        return record
    base.update(
        {
            "status": "metadata_only_pdf_no_text",
            "extractor": "pypdf",
            "pages_extracted": len(reader.pages),
            "errors": ["PDF contains no extractable text; likely scanned/image-only or encrypted."],
            "autonomous_next_action": "Run OCR or provide a text/CSV export before relying on this file's contents.",
        }
    )
    return base


def read_image_file_record(path: Path, base: dict[str, Any], sample_limit: int) -> dict[str, Any]:
    try:
        from PIL import Image
    except Exception as exc:
        base.update(
            {
                "status": "metadata_only_image_reader_unavailable",
                "extractor": "PIL",
                "errors": [f"PIL unavailable: {exc}"],
                "autonomous_next_action": "Install Pillow/OCR or provide a text/CSV export of the receipt/image.",
            }
        )
        return base
    with Image.open(path) as image:
        base["image"] = {"width": image.width, "height": image.height, "mode": image.mode}
    try:
        import pytesseract  # type: ignore

        base["ocr_command"] = configure_tesseract_command(pytesseract) or "PATH/default"
        with Image.open(path) as image:
            text = pytesseract.image_to_string(image)
        if text.strip():
            return readable_record(base, "readable_image_ocr", "pytesseract", text, sample_limit)
    except Exception as exc:
        base["errors"] = [f"OCR unavailable or failed: {type(exc).__name__}: {exc}"]
    base.update(
        {
            "status": "metadata_only_image_ocr_unavailable",
            "extractor": "PIL",
            "autonomous_next_action": "Run OCR or attach typed receipt details before treating this image as read evidence.",
        }
    )
    return base


def read_docx_file_record(path: Path, base: dict[str, Any], sample_limit: int) -> dict[str, Any]:
    with zipfile.ZipFile(path) as docx:
        xml = docx.read("word/document.xml")
    root = ET.fromstring(xml)
    text = "\n".join(node.text or "" for node in root.iter() if node.text)
    return readable_record(base, "readable_docx_text", "docx_zip_xml", text, sample_limit)


def read_zip_file_record(path: Path, base: dict[str, Any], sample_limit: int) -> dict[str, Any]:
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
    record = readable_record(base, "readable_archive_index", "zipfile", "\n".join(names), sample_limit)
    record["archive_member_count"] = len(names)
    record["autonomous_next_action"] = "Archive index was read; unpack and rerun if it contains accounting evidence."
    return record


def json_key_sample(value: Any, *, prefix: str = "", limit: int = 40) -> list[str]:
    out: list[str] = []

    def walk(item: Any, current: str) -> None:
        if len(out) >= limit:
            return
        if isinstance(item, dict):
            for key, child in item.items():
                path = f"{current}.{key}" if current else str(key)
                out.append(path)
                walk(child, path)
                if len(out) >= limit:
                    return
        elif isinstance(item, list) and item:
            walk(item[0], f"{current}[]")

    walk(value, prefix)
    return out[:limit]


def relative_to_repo(path: Path, repo_root: Path) -> str:
    try:
        return str(path.relative_to(repo_root)).replace("\\", "/")
    except ValueError:
        return str(path)


def content_probe(path: Path, limit: int = 8192) -> str:
    try:
        with path.open("rb") as fh:
            return fh.read(limit).decode("utf-8", errors="ignore").lower()
    except Exception:
        return ""


def normalise_path(value: Any) -> str:
    if not value:
        return ""
    try:
        return str(Path(str(value)).resolve()).lower()
    except Exception:
        return str(value).replace("/", "\\").lower()


def summarise_filesystem_records(records: list[dict[str, Any]]) -> dict[str, Any]:
    suffixes = Counter(item.get("suffix") or "<none>" for item in records)
    roots = Counter(item.get("raw_root") or "unknown" for item in records)
    return {
        "file_count": len(records),
        "bytes": sum(int(item.get("bytes") or 0) for item in records),
        "suffix_counts": dict(sorted(suffixes.items())),
        "root_counts": dict(sorted(roots.items())),
    }


def summarise_lookup_records(records: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "provider_counts": dict(sorted(Counter(item.get("provider_hint", "unknown") for item in records).items())),
        "category_counts": dict(sorted(Counter(item.get("category_hint", "unknown") for item in records).items())),
        "ingestion_path_counts": dict(sorted(Counter(item.get("ingestion_path_hint", "unknown") for item in records).items())),
    }


def summarise_source_closure(
    records: list[dict[str, Any]],
    readability_summary: dict[str, Any],
    raw_roots: Sequence[Path],
    missing_roots: Sequence[str],
) -> dict[str, Any]:
    decision_counts = Counter(item.get("decision", "unknown") for item in records)
    reason_counts = Counter(
        item.get("reason", "unknown")
        for item in records
        if item.get("decision") in {"excluded", "missing_root"}
    )
    included = [item for item in records if item.get("decision") == "included"]
    excluded = [item for item in records if item.get("decision") == "excluded"]
    missing = [item for item in records if item.get("decision") == "missing_root"]
    unaccounted = [
        item for item in records
        if item.get("decision") not in {"included", "excluded", "missing_root"}
    ]
    unreadable_count = (
        int(readability_summary.get("metadata_only_file_count", 0) or 0)
        + int(readability_summary.get("blocked_file_count", 0) or 0)
    )
    status = "closed_no_silent_skips"
    if missing or unaccounted or unreadable_count:
        status = "blocked_no_data_skip_contract"
    return {
        "schema_version": "accounting-source-closure-v1",
        "status": status,
        "raw_root_count": len(raw_roots),
        "raw_roots": [str(path) for path in raw_roots],
        "candidate_file_count": len(records) - len(missing),
        "included_file_count": len(included),
        "excluded_file_count": len(excluded),
        "missing_root_count": len(missing_roots),
        "unaccounted_file_count": len(unaccounted),
        "all_candidate_decisions_recorded": not unaccounted,
        "all_included_files_content_readable": unreadable_count == 0
        and int(readability_summary.get("readable_file_count", 0) or 0) == len(included),
        "readable_included_file_count": int(readability_summary.get("readable_file_count", 0) or 0),
        "unreadable_included_file_count": unreadable_count,
        "decision_counts": dict(sorted(decision_counts.items())),
        "reason_counts": dict(sorted(reason_counts.items())),
        "blocked_examples": [
            {
                "path": item.get("repo_relative_path") or item.get("path"),
                "reason": item.get("reason") or item.get("status"),
                "next_action": item.get("autonomous_next_action"),
            }
            for item in (missing + unaccounted + excluded)[:25]
            if item.get("decision") != "excluded" or not str(item.get("reason", "")).startswith("excluded_generated_cache_or_runtime_folder")
        ],
    }


def summarise_file_readability(records: list[dict[str, Any]]) -> dict[str, Any]:
    status_counts = Counter(item.get("status", "unknown") for item in records)
    extractor_counts = Counter(item.get("extractor", "unknown") or "unknown" for item in records)
    readable = [item for item in records if item.get("readable_by_aureon")]
    not_readable = [item for item in records if not item.get("readable_by_aureon")]
    metadata_only = [
        item
        for item in records
        if str(item.get("status", "")).startswith("metadata_only")
    ]
    blocked = [
        item
        for item in not_readable
        if not str(item.get("status", "")).startswith("metadata_only")
    ]
    file_count = len(records)
    accounted = len(readable) + len(metadata_only) + len(blocked)
    return {
        "schema_version": "accounting-file-readability-v1",
        "file_count": file_count,
        "readable_file_count": len(readable),
        "metadata_only_file_count": len(metadata_only),
        "blocked_file_count": len(blocked),
        "accounted_file_count": accounted,
        "readability_ratio": round(len(readable) / file_count, 6) if file_count else 1.0,
        "closed_loop_accounted_ratio": round(accounted / file_count, 6) if file_count else 1.0,
        "status_counts": dict(sorted(status_counts.items())),
        "extractor_counts": dict(sorted(extractor_counts.items())),
        "blocked_examples": [
            {
                "path": item.get("repo_relative_path") or item.get("path"),
                "status": item.get("status"),
                "next_action": item.get("autonomous_next_action"),
            }
            for item in not_readable[:25]
        ],
    }


def summarise_ingestion_records(records: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "standard_intake_seen": sum(1 for item in records if item.get("standard_intake_seen")),
        "combined_feed_seen": sum(1 for item in records if item.get("combined_feed_seen")),
        "evidence_inventory_seen": sum(1 for item in records if item.get("evidence_inventory_seen")),
        "fold_status_counts": dict(sorted(Counter(item.get("fold_status", "unknown") for item in records).items())),
        "rows_imported": sum(int(item.get("rows_imported") or 0) for item in records),
        "readable_by_aureon": sum(1 for item in records if item.get("readable_by_aureon")),
    }


def build_swarm_agent_roster() -> list[dict[str, str]]:
    return [
        {"agent": "raw_lookup_agent", "role": "sweep every raw-data file and hash it"},
        {"agent": "file_readability_agent", "role": "extract readable text, rows, pages, workbook sheets, image metadata, and blocked-file actions"},
        {"agent": "provider_lookup_agent", "role": "identify SumUp, Zempler, Revolut, HMRC, Companies House, and unknown sources"},
        {"agent": "payer_lookup_agent", "role": "check incoming payments over GBP 400 and CIS/construction signals against payer/public-register controls"},
        {"agent": "cis_vat_tax_basis_agent", "role": "link payer provenance to CIS suffered, VAT taxable turnover, and domestic reverse-charge workpapers"},
        {"agent": "ingestion_match_agent", "role": "compare raw files to combined bank data and period feed evidence"},
        {"agent": "evidence_folding_agent", "role": "map files to filing, tax, ledger, and review evidence"},
        {"agent": "phi_bridge_agent", "role": "score coherence with a deterministic Phi threshold"},
        {"agent": "end_user_observance_agent", "role": "produce confirmation language for the human operator"},
    ]


def record_benchmark(
    benchmark: list[dict[str, Any]],
    wave: str,
    started: float,
    *,
    file_count: int = 0,
    byte_count: int = 0,
    transaction_sources: int = 0,
    excluded_file_count: int = 0,
) -> None:
    duration = max(time.perf_counter() - started, 0.000001)
    benchmark.append(
        {
            "wave": wave,
            "duration_seconds": round(duration, 6),
            "file_count": int(file_count),
            "byte_count": int(byte_count),
            "excluded_file_count": int(excluded_file_count),
            "transaction_sources": int(transaction_sources),
            "files_per_second": round(file_count / duration, 3) if file_count else 0.0,
        }
    )


def detect_external_mutation_words(manifest: dict[str, Any]) -> list[str]:
    text = json.dumps(
        {
            "safe_boundaries": manifest.get("safe_boundaries"),
            "manual_next_step": manifest.get("manual_next_step"),
        },
        sort_keys=True,
    ).lower()
    return [
        word
        for word in EXTERNAL_MUTATION_WORDS
        if word in text and "manual" not in text[max(0, text.find(word) - 80): text.find(word) + 120]
    ]


def render_swarm_wave_scan_markdown(manifest: dict[str, Any]) -> str:
    waves = manifest.get("waves") or {}
    filesystem = waves.get("filesystem_discovery") or {}
    lookup = waves.get("provider_lookup") or {}
    ingestion = waves.get("ingestion_match") or {}
    folding = waves.get("evidence_folding") or {}
    consensus = waves.get("phi_swarm_consensus") or {}
    benchmark = manifest.get("benchmark") or {}
    lines = [
        "# Accounting Swarm Raw-Data Wave Scan",
        "",
        f"- Company: `{manifest.get('company_number')}` {manifest.get('company_name')}",
        f"- Period: `{manifest.get('period_start')}` to `{manifest.get('period_end')}`",
        f"- Status: `{manifest.get('status')}`",
        f"- Files scanned: `{filesystem.get('file_count', 0)}`",
        f"- Bytes scanned: `{filesystem.get('bytes', 0)}`",
        f"- Phi consensus: `{consensus.get('status', 'unknown')}` score `{consensus.get('score', 0)}` threshold `{consensus.get('threshold', 0)}`",
        f"- Benchmark: `{benchmark.get('total_duration_seconds', 0)}` seconds, `{benchmark.get('files_per_second', 0)}` files/sec",
        "",
        "## Raw Roots",
        "",
    ]
    for root in manifest.get("raw_roots") or []:
        lines.append(f"- `{root}`")
    lines.extend(["", "## No Data Skipped Source Closure", ""])
    closure = waves.get("source_closure") or manifest.get("source_closure") or {}
    lines.append(f"- `status`: {closure.get('status', 'unknown')}")
    lines.append(f"- `candidate_file_count`: {closure.get('candidate_file_count', 0)}")
    lines.append(f"- `included_file_count`: {closure.get('included_file_count', 0)}")
    lines.append(f"- `excluded_file_count`: {closure.get('excluded_file_count', 0)}")
    lines.append(f"- `missing_root_count`: {closure.get('missing_root_count', 0)}")
    lines.append(f"- `unaccounted_file_count`: {closure.get('unaccounted_file_count', 0)}")
    lines.extend(["", "## Provider Lookup", ""])
    for provider, count in (lookup.get("provider_counts") or {}).items():
        lines.append(f"- `{provider}`: {count}")
    lines.extend(["", "## Ingestion Match", ""])
    for key, value in ingestion.items():
        if key != "fold_status_counts":
            lines.append(f"- `{key}`: {value}")
    lines.extend(["", "## File Readability", ""])
    readability = waves.get("file_readability") or manifest.get("file_readability") or {}
    lines.append(f"- `readable_file_count`: {readability.get('readable_file_count', 0)}")
    lines.append(f"- `metadata_only_file_count`: {readability.get('metadata_only_file_count', 0)}")
    lines.append(f"- `blocked_file_count`: {readability.get('blocked_file_count', 0)}")
    lines.append(f"- `readability_ratio`: {readability.get('readability_ratio', 0)}")
    lines.extend(["", "## Fold Status", ""])
    for status, count in (folding.get("fold_status_counts") or {}).items():
        lines.append(f"- `{status}`: {count}")
    lines.extend(["", "## Internal Logic Chain", ""])
    for item in manifest.get("internal_logic_chain_checklist") or []:
        lines.append(f"- `{item.get('id')}`: {item.get('status')} - {item.get('evidence')}")
    lines.extend(
        [
            "",
            "## Safety",
            "",
            "- No Companies House filing was submitted.",
            "- No HMRC return was submitted.",
            "- No tax, penalty, or filing fee was paid.",
            "- No exchange or trading order was placed.",
            "",
        ]
    )
    return "\n".join(lines)


def render_logic_chain_markdown(payload: dict[str, Any]) -> str:
    consensus = payload.get("consensus") or {}
    lines = [
        "# Internal Accounting Logic Chain Checklist",
        "",
        f"- Generated: `{payload.get('generated_at')}`",
        f"- Company: `{payload.get('company_number')}`",
        f"- Period: `{payload.get('period_start')}` to `{payload.get('period_end')}`",
        f"- Phi consensus: `{consensus.get('status', 'unknown')}` score `{consensus.get('score', 0)}`",
        "",
        "## Checklist",
        "",
    ]
    for item in payload.get("checklist") or []:
        lines.append(f"- `{item.get('id')}`: {item.get('status')} - {item.get('evidence')}")
        for example in item.get("examples") or []:
            lines.append(f"  - `{example}`")
    lines.extend(
        [
            "",
            "## Safe Boundary",
            "",
            "- The chain may inspect, reconcile, generate local documents, and report missing evidence.",
            "- The chain may not submit filings, pay tax, authenticate to government services, or create external evidence.",
            "",
        ]
    )
    return "\n".join(lines)


def render_end_user_confirmation_markdown(payload: dict[str, Any]) -> str:
    confirmation = payload.get("confirmation") or {}
    lines = [
        "# End-User Confirmation",
        "",
        f"- Company: `{payload.get('company_number')}` {payload.get('company_name')}",
        f"- Period: `{payload.get('period_start')}` to `{payload.get('period_end')}`",
        f"- Status: `{payload.get('status')}`",
        "",
        "## What Aureon Confirmed",
        "",
    ]
    for item in confirmation.get("what_aureon_confirmed") or []:
        lines.append(f"- {item}")
    lines.extend(["", "## Human Actions Left", ""])
    for item in confirmation.get("human_actions_left") or []:
        lines.append(f"- {item}")
    attention = confirmation.get("attention_items") or []
    lines.extend(["", "## Attention Items", ""])
    if not attention:
        lines.append("- None from the swarm checklist.")
    for item in attention:
        lines.append(f"- `{item.get('id')}`: {item.get('status')} - {item.get('evidence')}")
    lines.extend(["", "## No External Actions Were Taken", ""])
    for item in confirmation.get("no_external_actions_taken") or []:
        lines.append(f"- {item}")
    lines.extend(["", "## Next Step", "", str(payload.get("manual_next_step") or ""), ""])
    return "\n".join(lines)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run accounting swarm raw-data wave scan.")
    parser.add_argument("--company-name", default=DEFAULT_COMPANY_NAME)
    parser.add_argument("--company-number", default=DEFAULT_COMPANY_NUMBER)
    parser.add_argument("--period-start", default=DEFAULT_PERIOD_START)
    parser.add_argument("--period-end", default=DEFAULT_PERIOD_END)
    parser.add_argument("--raw-data-dir", action="append", default=[])
    parser.add_argument("--no-default-roots", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    manifest = run_accounting_swarm_data_wave_scan(
        company_name=args.company_name,
        company_number=args.company_number,
        period_start=args.period_start,
        period_end=args.period_end,
        raw_data_roots=args.raw_data_dir,
        include_default_roots=not args.no_default_roots,
    )
    outputs = manifest.get("outputs") or {}
    print(outputs.get("swarm_raw_data_wave_scan_json", ""))
    print(outputs.get("swarm_raw_data_wave_scan_markdown", ""))
    return 0 if manifest.get("status") == "completed" else 1


if __name__ == "__main__":
    raise SystemExit(main())
