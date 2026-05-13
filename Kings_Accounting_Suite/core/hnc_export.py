"""
HNC EXPORT — hnc_export.py
===========================
Document Generation Engine.

Produces properly formatted PDF reports and Excel workbooks from
the HNC accounting pipeline output. Every document a UK sole trader
or small limited company needs — formatted to professional standards.

PDF outputs (reportlab):
    1. Profit & Loss Statement (FRS 102 Section 1A / Format 2)
    2. Balance Sheet (Statement of Financial Position)
    3. Tax Summary (client-friendly overview + payment schedule)
    4. Invoice / Credit Note (SI 1995/2518 reg 14 compliant)
    5. CIS Monthly Return Summary
    6. VAT Return Summary (9-box)
    7. Capital Gains Summary (SA108)

Excel outputs (openpyxl):
    1. Management Accounts (12-month P&L breakdown)
    2. General Ledger (full journal with formulas)
    3. Trial Balance
    4. Cash Flow Forecast
    5. Aged Debtors / Creditors

All monetary values GBP. All dates UK format. All documents
audit-trail back to journal entries.

Aureon Creator / Aureon Research — April 2026
"""

from __future__ import annotations
import logging
import os
from datetime import datetime, date
from typing import Any, Dict, List, Optional, Tuple
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

logger = logging.getLogger("hnc_export")

# ---------------------------------------------------------------------------
# reportlab imports
# ---------------------------------------------------------------------------
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, cm
from reportlab.lib.colors import HexColor, black, white, grey
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, KeepTogether, HRFlowable,
)
from reportlab.pdfgen import canvas

# ---------------------------------------------------------------------------
# openpyxl imports
# ---------------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
)
from openpyxl.utils import get_column_letter


# =========================================================================
# COLOUR PALETTE & STYLES
# =========================================================================

# HNC brand palette
HNC_NAVY    = HexColor("#1B2A4A")
HNC_BLUE    = HexColor("#2E5090")
HNC_GOLD    = HexColor("#C49A2E")
HNC_LIGHT   = HexColor("#F4F6F9")
HNC_RED     = HexColor("#C0392B")
HNC_GREEN   = HexColor("#27AE60")
HNC_GREY    = HexColor("#7F8C8D")
HNC_WHITE   = white

# Excel fills
XL_HEADER_FILL  = PatternFill("solid", fgColor="1B2A4A")
XL_SUBHEAD_FILL = PatternFill("solid", fgColor="2E5090")
XL_ALT_FILL     = PatternFill("solid", fgColor="F4F6F9")
XL_TOTAL_FILL   = PatternFill("solid", fgColor="E8ECEF")
XL_GOLD_FILL    = PatternFill("solid", fgColor="C49A2E")

# Excel fonts
XL_HEADER_FONT  = Font(name="Arial", size=11, bold=True, color="FFFFFF")
XL_SUBHEAD_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
XL_TOTAL_FONT   = Font(name="Arial", size=10, bold=True)
XL_BODY_FONT    = Font(name="Arial", size=10)
XL_TITLE_FONT   = Font(name="Arial", size=14, bold=True, color="1B2A4A")

# Common border
THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)
TOTAL_BORDER = Border(
    top=Side(style="thin", color="1B2A4A"),
    bottom=Side(style="double", color="1B2A4A"),
)


# =========================================================================
# HELPERS
# =========================================================================

def _fmt_gbp(amount: float, symbol: bool = True) -> str:
    """Format amount as GBP string."""
    if amount < 0:
        s = f"({abs(amount):,.2f})"
    else:
        s = f"{amount:,.2f}"
    return f"\u00a3{s}" if symbol else s


def _uk_date(d: str) -> str:
    """YYYY-MM-DD -> DD/MM/YYYY."""
    if not d:
        return ""
    if "/" in d and len(d.split("/")[0]) == 2:
        return d
    try:
        dt = datetime.strptime(d[:10], "%Y-%m-%d")
        return dt.strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return d


def _today_uk() -> str:
    return datetime.now().strftime("%d/%m/%Y")


def _safe_float(v: Any) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _get(obj, key, default=None):
    """Get attribute or dict key from obj."""
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


# =========================================================================
# PDF BASE — common page template
# =========================================================================

class HNCPageTemplate:
    """Header/footer for all HNC PDF documents."""

    def __init__(self, entity_name: str, report_title: str, period: str = ""):
        self.entity_name = entity_name
        self.report_title = report_title
        self.period = period

    def __call__(self, canvas_obj, doc):
        canvas_obj.saveState()
        w, h = A4

        # --- Header bar ---
        canvas_obj.setFillColor(HNC_NAVY)
        canvas_obj.rect(0, h - 28*mm, w, 28*mm, fill=1, stroke=0)

        # Gold accent line
        canvas_obj.setStrokeColor(HNC_GOLD)
        canvas_obj.setLineWidth(1.5)
        canvas_obj.line(15*mm, h - 28*mm, w - 15*mm, h - 28*mm)

        # Entity name
        canvas_obj.setFillColor(HNC_WHITE)
        canvas_obj.setFont("Helvetica-Bold", 14)
        canvas_obj.drawString(18*mm, h - 14*mm, self.entity_name)

        # Report title
        canvas_obj.setFont("Helvetica", 10)
        canvas_obj.drawString(18*mm, h - 22*mm, self.report_title)

        # Period (right side)
        if self.period:
            canvas_obj.setFont("Helvetica", 9)
            canvas_obj.drawRightString(w - 18*mm, h - 14*mm, self.period)

        # Date generated
        canvas_obj.setFont("Helvetica", 8)
        canvas_obj.drawRightString(w - 18*mm, h - 22*mm,
                                    f"Generated: {_today_uk()}")

        # --- Footer ---
        canvas_obj.setFillColor(HNC_GREY)
        canvas_obj.setFont("Helvetica", 7)
        canvas_obj.drawString(18*mm, 12*mm,
                              "The HNC Accountant — Aureon Research")
        canvas_obj.drawRightString(w - 18*mm, 12*mm,
                                   f"Page {doc.page}")

        # Footer line
        canvas_obj.setStrokeColor(HNC_GOLD)
        canvas_obj.setLineWidth(0.5)
        canvas_obj.line(15*mm, 17*mm, w - 15*mm, 17*mm)

        canvas_obj.restoreState()


def _make_doc(filepath: str, entity: str, title: str,
              period: str = "") -> SimpleDocTemplate:
    """Create a SimpleDocTemplate with HNC branding."""
    doc = SimpleDocTemplate(
        filepath,
        pagesize=A4,
        leftMargin=18*mm,
        rightMargin=18*mm,
        topMargin=34*mm,
        bottomMargin=24*mm,
    )
    doc._hnc_template = HNCPageTemplate(entity, title, period)
    return doc


def _build_pdf(doc: SimpleDocTemplate, story: list):
    """Build the PDF with header/footer."""
    doc.build(story, onFirstPage=doc._hnc_template,
              onLaterPages=doc._hnc_template)


# =========================================================================
# PARAGRAPH STYLES
# =========================================================================

_styles = getSampleStyleSheet()

STYLE_TITLE = ParagraphStyle(
    "HNCTitle", parent=_styles["Normal"],
    fontName="Helvetica-Bold", fontSize=13,
    textColor=HNC_NAVY, spaceAfter=4*mm,
)
STYLE_SUBTITLE = ParagraphStyle(
    "HNCSub", parent=_styles["Normal"],
    fontName="Helvetica-Bold", fontSize=10,
    textColor=HNC_BLUE, spaceBefore=4*mm, spaceAfter=2*mm,
)
STYLE_BODY = ParagraphStyle(
    "HNCBody", parent=_styles["Normal"],
    fontName="Helvetica", fontSize=9,
    leading=13, spaceAfter=2*mm,
)
STYLE_SMALL = ParagraphStyle(
    "HNCSmall", parent=_styles["Normal"],
    fontName="Helvetica", fontSize=7.5,
    textColor=HNC_GREY, leading=10,
)
STYLE_NOTE = ParagraphStyle(
    "HNCNote", parent=_styles["Normal"],
    fontName="Helvetica-Oblique", fontSize=8,
    textColor=HNC_GREY, leading=11,
)


# =========================================================================
# TABLE HELPERS
# =========================================================================

def _report_table(data: list, col_widths: list = None,
                  has_header: bool = True) -> Table:
    """Build a styled financial report table."""
    if not col_widths:
        avail = A4[0] - 36*mm
        col_widths = [avail * 0.55, avail * 0.225, avail * 0.225]

    t = Table(data, colWidths=col_widths, repeatRows=1 if has_header else 0)

    style_cmds = [
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (0, -1), 4),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("LINEBELOW", (0, 0), (-1, 0), 1, HNC_NAVY),
    ]

    if has_header:
        style_cmds += [
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("TEXTCOLOR", (0, 0), (-1, 0), HNC_NAVY),
            ("BACKGROUND", (0, 0), (-1, 0), HNC_LIGHT),
        ]

    # Alternate row shading
    for i in range(1, len(data)):
        if i % 2 == 0:
            style_cmds.append(
                ("BACKGROUND", (0, i), (-1, i), HNC_LIGHT)
            )

    t.setStyle(TableStyle(style_cmds))
    return t


def _total_row_style(table: Table, row_idx: int):
    """Apply total-row styling (bold + top border + double bottom)."""
    table.setStyle(TableStyle([
        ("FONTNAME", (0, row_idx), (-1, row_idx), "Helvetica-Bold"),
        ("LINEABOVE", (0, row_idx), (-1, row_idx), 0.8, HNC_NAVY),
        ("LINEBELOW", (0, row_idx), (-1, row_idx), 1.5, HNC_NAVY),
    ]))


# =========================================================================
# PDF 1 — PROFIT & LOSS STATEMENT
# =========================================================================

def export_pnl_pdf(report_data: Dict[str, Any], filepath: str,
                   entity: str = "Client") -> str:
    """
    Export P&L to PDF.

    report_data: output from hnc_reports.generate_pnl() — a FinancialReport
                 object or dict with .lines/.totals
    """
    period = _get(report_data, "period", "")
    doc = _make_doc(filepath, entity,
                    "Profit & Loss Account (FRS 102 Section 1A)", period)
    story = []

    # Statutory note
    story.append(Paragraph(
        "Prepared in accordance with FRS 102 Section 1A and "
        "Companies Act 2006 Schedule 1, Format 2 (vertical format).",
        STYLE_NOTE
    ))
    story.append(Spacer(1, 4*mm))

    # Build table from report lines
    lines = _get(report_data, "lines", [])
    totals = _get(report_data, "totals", {})

    data = [["", "Current Period\n\u00a3", "Prior Period\n\u00a3"]]
    total_rows = []

    for i, line in enumerate(lines):
        label = _get(line, "label", "")
        amount = _get(line, "amount", 0.0)
        indent = _get(line, "indent", 0)
        is_total = _get(line, "is_total", False)
        is_sub = _get(line, "is_subtotal", False)

        prefix = "    " * indent
        row = [
            f"{prefix}{label}",
            _fmt_gbp(amount, symbol=False) if amount != 0 else "—",
            "—",  # Prior period placeholder
        ]
        data.append(row)
        if is_total or is_sub:
            total_rows.append(len(data) - 1)

    t = _report_table(data)
    for r in total_rows:
        _total_row_style(t, r)
    story.append(t)

    # Notes
    story.append(Spacer(1, 6*mm))
    notes = _get(report_data, "notes", [])
    if notes:
        story.append(Paragraph("Notes:", STYLE_SUBTITLE))
        for n in notes:
            story.append(Paragraph(n, STYLE_SMALL))

    _build_pdf(doc, story)
    logger.info(f"P&L PDF exported: {filepath}")
    return filepath


# =========================================================================
# PDF 2 — BALANCE SHEET
# =========================================================================

def export_balance_sheet_pdf(report_data: Dict[str, Any], filepath: str,
                             entity: str = "Client") -> str:
    """Export Balance Sheet to PDF."""
    period = _get(report_data, "period", "")
    doc = _make_doc(filepath, entity,
                    "Statement of Financial Position", period)
    story = []

    story.append(Paragraph(
        "Prepared in accordance with FRS 102 Section 1A.",
        STYLE_NOTE
    ))
    story.append(Spacer(1, 4*mm))

    lines = _get(report_data, "lines", [])
    data = [["", "Current\n\u00a3", "Prior\n\u00a3"]]
    total_rows = []

    for line in lines:
        label = _get(line, "label", "")
        amount = _get(line, "amount", 0.0)
        indent = _get(line, "indent", 0)
        is_total = _get(line, "is_total", False)

        prefix = "    " * indent
        row = [f"{prefix}{label}",
               _fmt_gbp(amount, False) if amount != 0 else "—", "—"]
        data.append(row)
        if is_total:
            total_rows.append(len(data) - 1)

    t = _report_table(data)
    for r in total_rows:
        _total_row_style(t, r)
    story.append(t)

    # Director sign-off
    story.append(Spacer(1, 12*mm))
    story.append(Paragraph(
        "Approved by the director and authorised for issue on "
        f"{_today_uk()}.", STYLE_BODY))
    story.append(Spacer(1, 10*mm))
    story.append(Paragraph("_________________________", STYLE_BODY))
    story.append(Paragraph("Director", STYLE_SMALL))

    _build_pdf(doc, story)
    logger.info(f"Balance Sheet PDF exported: {filepath}")
    return filepath


# =========================================================================
# PDF 3 — TAX SUMMARY
# =========================================================================

def export_tax_summary_pdf(tax_data: Dict[str, Any], filepath: str,
                           entity: str = "Client") -> str:
    """Export client-friendly tax summary with payment schedule."""
    period = tax_data.get("tax_year", "2025/26")
    doc = _make_doc(filepath, entity, "Tax Summary", f"Tax Year {period}")
    story = []

    # Key figures
    story.append(Paragraph("Key Figures", STYLE_SUBTITLE))
    key_items = [
        ["Turnover", _fmt_gbp(_safe_float(tax_data.get("turnover", 0)))],
        ["Allowable Expenses",
         _fmt_gbp(_safe_float(tax_data.get("total_expenses", 0)))],
        ["Net Profit",
         _fmt_gbp(_safe_float(tax_data.get("net_profit", 0)))],
        ["", ""],
        ["Income Tax",
         _fmt_gbp(_safe_float(tax_data.get("income_tax", 0)))],
        ["National Insurance (Class 2)",
         _fmt_gbp(_safe_float(tax_data.get("nic2", 0)))],
        ["National Insurance (Class 4)",
         _fmt_gbp(_safe_float(tax_data.get("nic4", 0)))],
        ["Total Tax Due",
         _fmt_gbp(_safe_float(tax_data.get("total_tax", 0)))],
    ]
    avail = A4[0] - 36*mm
    t = Table(key_items, colWidths=[avail * 0.6, avail * 0.4])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("FONTNAME", (0, 7), (-1, 7), "Helvetica-Bold"),
        ("LINEABOVE", (0, 7), (-1, 7), 0.8, HNC_NAVY),
        ("LINEBELOW", (0, 7), (-1, 7), 1.5, HNC_NAVY),
        ("FONTNAME", (0, 3), (-1, 3), "Helvetica"),
        ("LINEABOVE", (0, 4), (-1, 4), 0.5, HNC_GREY),
    ]))
    story.append(t)

    # Payment schedule
    schedule = tax_data.get("payment_schedule", [])
    if schedule:
        story.append(Spacer(1, 6*mm))
        story.append(Paragraph("Payment Schedule", STYLE_SUBTITLE))
        sched_data = [["Date", "Description", "Amount"]]
        for p in schedule:
            sched_data.append([
                _uk_date(p.get("date", "")),
                p.get("description", ""),
                _fmt_gbp(_safe_float(p.get("amount", 0))),
            ])
        st = _report_table(sched_data,
                           col_widths=[avail * 0.25, avail * 0.45, avail * 0.3])
        story.append(st)

    # Effective rate callout
    effective = tax_data.get("effective_rate", 0)
    if effective:
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(
            f"Effective tax rate: {effective:.1f}%", STYLE_BODY))

    story.append(Spacer(1, 6*mm))
    story.append(Paragraph(
        "This summary is for information only and does not constitute "
        "professional tax advice. Tax computations are based on HMRC "
        "rates and thresholds for the stated tax year.",
        STYLE_NOTE
    ))

    _build_pdf(doc, story)
    logger.info(f"Tax Summary PDF exported: {filepath}")
    return filepath


# =========================================================================
# PDF 4 — INVOICE
# =========================================================================

def export_invoice_pdf(invoice: Any, filepath: str) -> str:
    """
    Export a single invoice to PDF.

    invoice: Invoice dataclass from hnc_invoice.py
    """
    inv_type = getattr(invoice, "invoice_type", None)
    type_label = inv_type.value if inv_type else "Invoice"
    inv_num = getattr(invoice, "invoice_number", "")

    supplier = getattr(invoice, "supplier", None)
    entity = getattr(supplier, "name", "Supplier") if supplier else "Supplier"

    doc = _make_doc(filepath, entity, f"{type_label} {inv_num}")
    story = []

    # --- Invoice header details ---
    customer = getattr(invoice, "customer", None)

    header_data = [
        [Paragraph(f"<b>{type_label}</b>", STYLE_TITLE), "", ""],
        ["Invoice No:", inv_num,
         f"Date: {_uk_date(getattr(invoice, 'date_issued', ''))}"],
        ["Tax Point:", _uk_date(getattr(invoice, "date_supply", "")),
         f"Due: {_uk_date(getattr(invoice, 'date_due', ''))}"],
    ]

    if getattr(invoice, "po_number", ""):
        header_data.append(["PO Number:", invoice.po_number, ""])

    avail = A4[0] - 36*mm
    ht = Table(header_data, colWidths=[avail * 0.3, avail * 0.35, avail * 0.35])
    ht.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(ht)
    story.append(Spacer(1, 4*mm))

    # Supplier / Customer addresses
    if supplier and customer:
        addr_data = [[
            Paragraph(f"<b>From:</b><br/>{_party_text(supplier)}", STYLE_BODY),
            Paragraph(f"<b>To:</b><br/>{_party_text(customer)}", STYLE_BODY),
        ]]
        at = Table(addr_data, colWidths=[avail * 0.5, avail * 0.5])
        at.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(at)
        story.append(Spacer(1, 4*mm))

    # --- Line items ---
    items = getattr(invoice, "items", [])
    line_data = [["Description", "Qty", "Rate", "Net", "VAT", "Gross"]]
    for item in items:
        line_data.append([
            getattr(item, "description", ""),
            f"{getattr(item, 'quantity', 1):.1f}",
            _fmt_gbp(getattr(item, "unit_price", 0)),
            _fmt_gbp(getattr(item, "net_amount", 0)),
            _fmt_gbp(getattr(item, "vat_amount", 0)),
            _fmt_gbp(getattr(item, "gross_amount", 0)),
        ])

    col_w = [avail*0.32, avail*0.08, avail*0.12, avail*0.16, avail*0.16, avail*0.16]
    lt = Table(line_data, colWidths=col_w)
    lt.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), HNC_LIGHT),
        ("LINEBELOW", (0, 0), (-1, 0), 1, HNC_NAVY),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(lt)
    story.append(Spacer(1, 3*mm))

    # --- Totals ---
    totals_data = [
        ["Subtotal (Net)", _fmt_gbp(getattr(invoice, "subtotal_net", 0))],
        ["VAT", _fmt_gbp(getattr(invoice, "total_vat", 0))],
    ]

    # CIS deduction
    if getattr(invoice, "cis_applicable", False):
        totals_data.append(["CIS Deduction",
                            f"({_fmt_gbp(getattr(invoice, 'cis_deduction', 0))[1:]})"])
        totals_data.append(["Net Payment Due",
                            _fmt_gbp(getattr(invoice, "cis_net_payment", 0))])
    else:
        totals_data.append(["Total Due",
                            _fmt_gbp(getattr(invoice, "total_gross", 0))])

    tt = Table(totals_data, colWidths=[avail * 0.7, avail * 0.3])
    tt.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, -1), (-1, -1), 0.8, HNC_NAVY),
        ("LINEBELOW", (0, -1), (-1, -1), 1.5, HNC_NAVY),
    ]))
    story.append(tt)

    # Reverse charge notice
    if getattr(invoice, "is_reverse_charge", False):
        story.append(Spacer(1, 4*mm))
        notice = getattr(invoice, "reverse_charge_note",
                         "Customer to account for VAT to HMRC "
                         "(VATA 1994 s.55A — domestic reverse charge).")
        story.append(Paragraph(f"<b>REVERSE CHARGE:</b> {notice}", STYLE_BODY))

    # Bank details
    bank = getattr(invoice, "bank_details", "")
    if bank:
        story.append(Spacer(1, 6*mm))
        story.append(Paragraph("<b>Payment Details:</b>", STYLE_BODY))
        story.append(Paragraph(bank.replace("\n", "<br/>"), STYLE_BODY))

    # Notes
    notes = getattr(invoice, "notes", "")
    if notes:
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(notes, STYLE_SMALL))

    # VAT registration
    if supplier and getattr(supplier, "vat_number", ""):
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(
            f"VAT Registration: {supplier.vat_number}", STYLE_SMALL))

    _build_pdf(doc, story)
    logger.info(f"Invoice PDF exported: {filepath}")
    return filepath


def _party_text(party) -> str:
    """Format an InvoiceParty for display."""
    parts = [getattr(party, "name", "")]
    ta = getattr(party, "trading_as", "")
    if ta:
        parts.append(f"t/a {ta}")
    addr = getattr(party, "address", None)
    if addr:
        parts.append(addr.format() if hasattr(addr, "format") else str(addr))
    vat = getattr(party, "vat_number", "")
    if vat:
        parts.append(f"VAT: {vat}")
    utr = getattr(party, "utr", "")
    if utr:
        parts.append(f"UTR: {utr}")
    return "<br/>".join(parts)


# =========================================================================
# PDF 5 — CIS SUMMARY
# =========================================================================

def export_cis_summary_pdf(cis_data: Dict[str, Any], filepath: str,
                           entity: str = "Client") -> str:
    """Export CIS monthly return summary to PDF."""
    period = cis_data.get("period", "")
    doc = _make_doc(filepath, entity, "CIS Monthly Return Summary", period)
    story = []

    story.append(Paragraph(
        "Construction Industry Scheme — FA 2004 s.61-s.77",
        STYLE_NOTE))
    story.append(Spacer(1, 4*mm))

    # Contractor details
    story.append(Paragraph("Contractor Details", STYLE_SUBTITLE))
    details = [
        ["Contractor Name:", cis_data.get("contractor_name", entity)],
        ["Contractor UTR:", cis_data.get("contractor_utr", "")],
        ["Period:", period],
    ]
    avail = A4[0] - 36*mm
    dt = Table(details, colWidths=[avail * 0.35, avail * 0.65])
    dt.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(dt)
    story.append(Spacer(1, 4*mm))

    # Subcontractor breakdown
    subs = cis_data.get("subcontractors", [])
    if subs:
        story.append(Paragraph("Subcontractor Payments", STYLE_SUBTITLE))
        sub_data = [["Name", "UTR", "Gross", "Materials", "Tax Deducted", "Net Paid"]]
        for s in subs:
            sub_data.append([
                s.get("name", ""),
                s.get("utr", ""),
                _fmt_gbp(_safe_float(s.get("gross", 0))),
                _fmt_gbp(_safe_float(s.get("materials", 0))),
                _fmt_gbp(_safe_float(s.get("deduction", 0))),
                _fmt_gbp(_safe_float(s.get("net_paid", 0))),
            ])
        cw = [avail*0.22, avail*0.14, avail*0.14, avail*0.14, avail*0.18, avail*0.18]
        st = Table(sub_data, colWidths=cw)
        st.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), HNC_LIGHT),
            ("LINEBELOW", (0, 0), (-1, 0), 1, HNC_NAVY),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ]))
        story.append(st)

    # Totals
    story.append(Spacer(1, 4*mm))
    summary = [
        ["Total Gross Payments",
         _fmt_gbp(_safe_float(cis_data.get("total_gross", 0)))],
        ["Total Materials",
         _fmt_gbp(_safe_float(cis_data.get("total_materials", 0)))],
        ["Total Deductions",
         _fmt_gbp(_safe_float(cis_data.get("total_deductions", 0)))],
        ["Total Net Paid",
         _fmt_gbp(_safe_float(cis_data.get("total_net_paid", 0)))],
    ]
    smt = Table(summary, colWidths=[avail * 0.65, avail * 0.35])
    smt.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, -1), (-1, -1), 0.8, HNC_NAVY),
        ("LINEBELOW", (0, -1), (-1, -1), 1.5, HNC_NAVY),
    ]))
    story.append(smt)

    _build_pdf(doc, story)
    logger.info(f"CIS Summary PDF exported: {filepath}")
    return filepath


# =========================================================================
# PDF 6 — VAT RETURN SUMMARY
# =========================================================================

def export_vat_return_pdf(vat_data: Dict[str, Any], filepath: str,
                          entity: str = "Client") -> str:
    """Export VAT MTD 9-box return to PDF."""
    period = vat_data.get("period", "")
    doc = _make_doc(filepath, entity, "VAT Return (MTD)", period)
    story = []

    story.append(Paragraph(
        "Making Tax Digital — Value Added Tax Regulations (SI 1995/2518)",
        STYLE_NOTE))
    story.append(Spacer(1, 4*mm))

    # 9-box layout
    boxes = [
        ("Box 1", "VAT due on sales", vat_data.get("box1", 0)),
        ("Box 2", "VAT due on acquisitions (EU)", vat_data.get("box2", 0)),
        ("Box 3", "Total VAT due (Box 1 + Box 2)", vat_data.get("box3", 0)),
        ("Box 4", "VAT reclaimed on purchases", vat_data.get("box4", 0)),
        ("Box 5", "Net VAT to pay/reclaim (Box 3 - Box 4)",
         vat_data.get("box5", 0)),
        ("Box 6", "Total sales (ex-VAT)", vat_data.get("box6", 0)),
        ("Box 7", "Total purchases (ex-VAT)", vat_data.get("box7", 0)),
        ("Box 8", "Total supplies (EU)", vat_data.get("box8", 0)),
        ("Box 9", "Total acquisitions (EU)", vat_data.get("box9", 0)),
    ]

    data = [["Box", "Description", "Amount"]]
    for box_num, desc, amount in boxes:
        data.append([box_num, desc, _fmt_gbp(_safe_float(amount))])

    avail = A4[0] - 36*mm
    vt = Table(data, colWidths=[avail * 0.1, avail * 0.6, avail * 0.3])
    vt.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), HNC_LIGHT),
        ("LINEBELOW", (0, 0), (-1, 0), 1, HNC_NAVY),
        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        # Highlight Box 5 (net payment)
        ("FONTNAME", (0, 5), (-1, 5), "Helvetica-Bold"),
        ("BACKGROUND", (0, 5), (-1, 5), HNC_LIGHT),
        ("LINEABOVE", (0, 5), (-1, 5), 0.8, HNC_NAVY),
        ("LINEBELOW", (0, 5), (-1, 5), 1, HNC_NAVY),
        # Separator between monetary and value boxes
        ("LINEABOVE", (0, 6), (-1, 6), 0.5, HNC_GREY),
    ]))
    story.append(vt)

    # Scheme info
    scheme = vat_data.get("scheme", "")
    if scheme:
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(f"VAT Scheme: {scheme}", STYLE_BODY))

    flat_rate = vat_data.get("flat_rate_percentage", 0)
    if flat_rate:
        story.append(Paragraph(
            f"Flat Rate: {flat_rate}%", STYLE_BODY))

    _build_pdf(doc, story)
    logger.info(f"VAT Return PDF exported: {filepath}")
    return filepath


# =========================================================================
# PDF 7 — CAPITAL GAINS SUMMARY (SA108)
# =========================================================================

def export_cgt_summary_pdf(cgt_data: Dict[str, Any], filepath: str,
                           entity: str = "Client") -> str:
    """Export Capital Gains Tax summary (SA108) to PDF."""
    period = cgt_data.get("tax_year", "2025/26")
    doc = _make_doc(filepath, entity,
                    "Capital Gains Summary (SA108)", f"Tax Year {period}")
    story = []

    story.append(Paragraph(
        "Taxation of Chargeable Gains Act 1992 — Section 104 Pool Method",
        STYLE_NOTE))
    story.append(Spacer(1, 4*mm))

    # Summary figures
    story.append(Paragraph("Summary", STYLE_SUBTITLE))
    avail = A4[0] - 36*mm
    summary = [
        ["Total Disposal Proceeds",
         _fmt_gbp(_safe_float(cgt_data.get("total_proceeds", 0)))],
        ["Total Allowable Costs",
         _fmt_gbp(_safe_float(cgt_data.get("total_costs", 0)))],
        ["Total Gains",
         _fmt_gbp(_safe_float(cgt_data.get("total_gains", 0)))],
        ["Total Losses",
         _fmt_gbp(_safe_float(cgt_data.get("total_losses", 0)))],
        ["Net Gains / (Losses)",
         _fmt_gbp(_safe_float(cgt_data.get("net_gain", 0)))],
        ["Annual Exempt Amount",
         _fmt_gbp(_safe_float(cgt_data.get("annual_exempt", 3000)))],
        ["Taxable Gain",
         _fmt_gbp(_safe_float(cgt_data.get("taxable_gain", 0)))],
        ["CGT Due",
         _fmt_gbp(_safe_float(cgt_data.get("cgt_due", 0)))],
    ]
    st = Table(summary, colWidths=[avail * 0.6, avail * 0.4])
    st.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, -1), (-1, -1), 0.8, HNC_NAVY),
        ("LINEBELOW", (0, -1), (-1, -1), 1.5, HNC_NAVY),
    ]))
    story.append(st)

    # Individual disposals
    disposals = cgt_data.get("disposals", [])
    if disposals:
        story.append(Spacer(1, 6*mm))
        story.append(Paragraph("Disposal Details", STYLE_SUBTITLE))
        disp_data = [["Asset", "Date", "Proceeds", "Cost", "Gain/(Loss)", "Rule"]]
        for d in disposals:
            gain = _safe_float(d.get("gain", 0))
            disp_data.append([
                d.get("asset", ""),
                _uk_date(d.get("date", "")),
                _fmt_gbp(_safe_float(d.get("proceeds", 0))),
                _fmt_gbp(_safe_float(d.get("cost", 0))),
                _fmt_gbp(gain),
                d.get("rule", "S104"),
            ])
        cw = [avail*0.15, avail*0.14, avail*0.16, avail*0.16, avail*0.19, avail*0.2]
        dt = Table(disp_data, colWidths=cw)
        dt.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), HNC_LIGHT),
            ("LINEBELOW", (0, 0), (-1, 0), 1, HNC_NAVY),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ]))
        story.append(dt)

    # SA108 box mapping
    sa108 = cgt_data.get("sa108_boxes", {})
    if sa108:
        story.append(Spacer(1, 6*mm))
        story.append(Paragraph("SA108 Box Mapping", STYLE_SUBTITLE))
        box_data = [["Box", "Description", "Value"]]
        box_labels = {
            "box1": "Number of disposals",
            "box3": "Total disposal proceeds",
            "box4": "Total allowable costs",
            "box6": "Total gains (before losses)",
            "box7": "Total losses",
            "box11": "Taxable gains after AEA",
            "box14": "CGT due",
        }
        for box_key, label in box_labels.items():
            val = sa108.get(box_key, 0)
            if val:
                box_data.append([
                    box_key.upper().replace("BOX", "Box "),
                    label,
                    _fmt_gbp(_safe_float(val)) if box_key != "box1" else str(val),
                ])
        if len(box_data) > 1:
            bt = _report_table(box_data,
                               [avail * 0.12, avail * 0.55, avail * 0.33])
            story.append(bt)

    _build_pdf(doc, story)
    logger.info(f"CGT Summary PDF exported: {filepath}")
    return filepath


# =========================================================================
# EXCEL 1 — MANAGEMENT ACCOUNTS
# =========================================================================

def export_management_accounts_xlsx(monthly_data: List[Dict],
                                     filepath: str,
                                     entity: str = "Client") -> str:
    """
    Export 12-month management accounts to Excel with formulas.

    monthly_data: list of dicts, each with month, turnover, expenses by
                  category, net_profit.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Management Accounts"

    # Title
    ws.merge_cells("A1:N1")
    ws["A1"] = f"{entity} — Management Accounts"
    ws["A1"].font = XL_TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:N2")
    ws["A2"] = f"Generated: {_today_uk()}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="7F8C8D")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers — row 4
    months = [d.get("month", f"M{i+1}") for i, d in enumerate(monthly_data)]
    headers = ["Category"] + months + ["Total"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = XL_HEADER_FONT
        cell.fill = XL_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 28
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    # Gather all expense categories
    all_cats = set()
    for d in monthly_data:
        expenses = d.get("expenses", {})
        all_cats.update(expenses.keys())
    cat_list = sorted(all_cats)

    # Row 5: Turnover
    row = 5
    ws.cell(row=row, column=1, value="Turnover").font = XL_TOTAL_FONT
    for col_idx, d in enumerate(monthly_data, 2):
        c = ws.cell(row=row, column=col_idx,
                    value=_safe_float(d.get("turnover", 0)))
        c.number_format = '#,##0.00'
        c.font = Font(name="Arial", size=10, color="0000FF")
    # Total formula
    first_col = get_column_letter(2)
    last_col = get_column_letter(1 + len(monthly_data))
    total_col = len(monthly_data) + 2
    ws.cell(row=row, column=total_col,
            value=f"=SUM({first_col}{row}:{last_col}{row})")
    ws.cell(row=row, column=total_col).font = XL_TOTAL_FONT
    ws.cell(row=row, column=total_col).number_format = '#,##0.00'

    # Row 6: blank separator
    row = 6
    ws.cell(row=row, column=1, value="Less: Expenses").font = XL_SUBHEAD_FONT
    ws.cell(row=row, column=1).fill = XL_SUBHEAD_FILL

    # Expense rows
    expense_start_row = 7
    for cat_idx, cat in enumerate(cat_list):
        row = expense_start_row + cat_idx
        ws.cell(row=row, column=1, value=cat.replace("_", " ").title())
        ws.cell(row=row, column=1).font = XL_BODY_FONT
        if cat_idx % 2 == 0:
            for c in range(1, total_col + 1):
                ws.cell(row=row, column=c).fill = XL_ALT_FILL

        for col_idx, d in enumerate(monthly_data, 2):
            val = _safe_float(d.get("expenses", {}).get(cat, 0))
            c = ws.cell(row=row, column=col_idx, value=val)
            c.number_format = '#,##0.00'
            c.font = XL_BODY_FONT

        # Total formula
        ws.cell(row=row, column=total_col,
                value=f"=SUM({first_col}{row}:{last_col}{row})")
        ws.cell(row=row, column=total_col).number_format = '#,##0.00'

    # Total Expenses row
    exp_end_row = expense_start_row + len(cat_list) - 1
    row = exp_end_row + 1
    ws.cell(row=row, column=1, value="Total Expenses").font = XL_TOTAL_FONT
    ws.cell(row=row, column=1).fill = XL_TOTAL_FILL
    for col_idx in range(2, total_col + 1):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=row, column=col_idx,
                value=f"=SUM({col_letter}{expense_start_row}:{col_letter}{exp_end_row})")
        ws.cell(row=row, column=col_idx).font = XL_TOTAL_FONT
        ws.cell(row=row, column=col_idx).fill = XL_TOTAL_FILL
        ws.cell(row=row, column=col_idx).number_format = '#,##0.00'

    total_exp_row = row

    # Net Profit row
    row += 1
    ws.cell(row=row, column=1, value="Net Profit").font = XL_TOTAL_FONT
    for col_idx in range(2, total_col + 1):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=row, column=col_idx,
                value=f"={col_letter}5-{col_letter}{total_exp_row}")
        ws.cell(row=row, column=col_idx).font = XL_TOTAL_FONT
        ws.cell(row=row, column=col_idx).number_format = '#,##0.00'

    net_profit_row = row

    # Margin % row
    row += 1
    ws.cell(row=row, column=1, value="Net Margin %").font = Font(
        name="Arial", size=10, italic=True, color="7F8C8D")
    for col_idx in range(2, total_col + 1):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=row, column=col_idx,
                value=f'=IF({col_letter}5=0,0,{col_letter}{net_profit_row}/{col_letter}5)')
        ws.cell(row=row, column=col_idx).number_format = '0.0%'
        ws.cell(row=row, column=col_idx).font = Font(
            name="Arial", size=10, italic=True, color="7F8C8D")

    # Borders on total rows
    for c in range(1, total_col + 1):
        ws.cell(row=total_exp_row, column=c).border = TOTAL_BORDER
        ws.cell(row=net_profit_row, column=c).border = TOTAL_BORDER

    # Freeze panes
    ws.freeze_panes = "B5"

    # Print settings
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(filepath)
    logger.info(f"Management Accounts XLSX exported: {filepath}")
    return filepath


# =========================================================================
# EXCEL 2 — GENERAL LEDGER
# =========================================================================

def export_ledger_xlsx(journal_entries: List[Dict], filepath: str,
                       entity: str = "Client") -> str:
    """
    Export general ledger to Excel with running balance formulas.

    journal_entries: list of dicts with date, account, description,
                     debit, credit, reference.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "General Ledger"

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{entity} — General Ledger"
    ws["A1"].font = XL_TITLE_FONT

    # Headers
    headers = ["Date", "Account", "Description", "Reference",
               "Debit", "Credit", "Balance"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = XL_HEADER_FONT
        cell.fill = XL_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    widths = [12, 20, 35, 14, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    for row_idx, entry in enumerate(journal_entries, 4):
        ws.cell(row=row_idx, column=1,
                value=_uk_date(entry.get("date", "")))
        ws.cell(row=row_idx, column=2,
                value=entry.get("account", ""))
        ws.cell(row=row_idx, column=3,
                value=entry.get("description", ""))
        ws.cell(row=row_idx, column=4,
                value=entry.get("reference", ""))

        debit = _safe_float(entry.get("debit", 0))
        credit = _safe_float(entry.get("credit", 0))
        ws.cell(row=row_idx, column=5, value=debit if debit else None)
        ws.cell(row=row_idx, column=6, value=credit if credit else None)

        # Running balance formula
        if row_idx == 4:
            ws.cell(row=row_idx, column=7,
                    value=f"=E{row_idx}-F{row_idx}")
        else:
            ws.cell(row=row_idx, column=7,
                    value=f"=G{row_idx-1}+E{row_idx}-F{row_idx}")

        # Number formats
        for c in [5, 6, 7]:
            ws.cell(row=row_idx, column=c).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=c).font = XL_BODY_FONT

        # Alternate row shading
        if row_idx % 2 == 0:
            for c in range(1, 8):
                ws.cell(row=row_idx, column=c).fill = XL_ALT_FILL

    # Summary row
    last_row = 3 + len(journal_entries)
    sum_row = last_row + 1
    ws.cell(row=sum_row, column=4, value="TOTALS").font = XL_TOTAL_FONT
    ws.cell(row=sum_row, column=5,
            value=f"=SUM(E4:E{last_row})").font = XL_TOTAL_FONT
    ws.cell(row=sum_row, column=6,
            value=f"=SUM(F4:F{last_row})").font = XL_TOTAL_FONT
    ws.cell(row=sum_row, column=5).number_format = '#,##0.00'
    ws.cell(row=sum_row, column=6).number_format = '#,##0.00'
    for c in range(1, 8):
        ws.cell(row=sum_row, column=c).border = TOTAL_BORDER
        ws.cell(row=sum_row, column=c).fill = XL_TOTAL_FILL

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:G{last_row}"

    wb.save(filepath)
    logger.info(f"General Ledger XLSX exported: {filepath}")
    return filepath


# =========================================================================
# EXCEL 3 — TRIAL BALANCE
# =========================================================================

def export_trial_balance_xlsx(accounts: List[Dict], filepath: str,
                              entity: str = "Client",
                              period: str = "") -> str:
    """
    Export trial balance to Excel.

    accounts: list of dicts with account_code, account_name, debit, credit.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"

    ws.merge_cells("A1:E1")
    ws["A1"] = f"{entity} — Trial Balance"
    ws["A1"].font = XL_TITLE_FONT
    if period:
        ws.merge_cells("A2:E2")
        ws["A2"] = f"As at {period}"
        ws["A2"].font = Font(name="Arial", size=9, italic=True)

    headers = ["Code", "Account Name", "Debit", "Credit"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = XL_HEADER_FONT
        cell.fill = XL_HEADER_FILL

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    for row_idx, acct in enumerate(accounts, 5):
        ws.cell(row=row_idx, column=1,
                value=acct.get("account_code", ""))
        ws.cell(row=row_idx, column=2,
                value=acct.get("account_name", ""))
        debit = _safe_float(acct.get("debit", 0))
        credit = _safe_float(acct.get("credit", 0))
        if debit:
            ws.cell(row=row_idx, column=3, value=debit)
        if credit:
            ws.cell(row=row_idx, column=4, value=credit)
        for c in [3, 4]:
            ws.cell(row=row_idx, column=c).number_format = '#,##0.00'
        if row_idx % 2 == 0:
            for c in range(1, 5):
                ws.cell(row=row_idx, column=c).fill = XL_ALT_FILL

    last_row = 4 + len(accounts)
    sum_row = last_row + 1
    ws.cell(row=sum_row, column=2, value="TOTALS").font = XL_TOTAL_FONT
    ws.cell(row=sum_row, column=3,
            value=f"=SUM(C5:C{last_row})")
    ws.cell(row=sum_row, column=4,
            value=f"=SUM(D5:D{last_row})")
    for c in [3, 4]:
        ws.cell(row=sum_row, column=c).font = XL_TOTAL_FONT
        ws.cell(row=sum_row, column=c).number_format = '#,##0.00'
        ws.cell(row=sum_row, column=c).border = TOTAL_BORDER

    # Check row: difference should be zero
    check_row = sum_row + 1
    ws.cell(row=check_row, column=2, value="Check (should be zero)")
    ws.cell(row=check_row, column=2).font = Font(
        name="Arial", size=9, italic=True, color="C0392B")
    ws.cell(row=check_row, column=3,
            value=f"=C{sum_row}-D{sum_row}")
    ws.cell(row=check_row, column=3).number_format = '#,##0.00'
    ws.cell(row=check_row, column=3).font = Font(
        name="Arial", size=10, bold=True, color="C0392B")

    ws.freeze_panes = "A5"

    wb.save(filepath)
    logger.info(f"Trial Balance XLSX exported: {filepath}")
    return filepath


# =========================================================================
# EXCEL 4 — AGED DEBTORS
# =========================================================================

def export_aged_debtors_xlsx(debtors: List[Dict], filepath: str,
                             entity: str = "Client",
                             as_at: str = "") -> str:
    """
    Export aged debtors report to Excel.

    debtors: list with customer, current, 1_30, 31_60, 61_90, over_90, total.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Aged Debtors"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"{entity} — Aged Debtors Report"
    ws["A1"].font = XL_TITLE_FONT
    if as_at:
        ws["A2"] = f"As at {_uk_date(as_at)}"
        ws["A2"].font = Font(name="Arial", size=9, italic=True)

    headers = ["Customer", "Current", "1-30 Days", "31-60 Days",
               "61-90 Days", "90+ Days", "Total"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = XL_HEADER_FONT
        cell.fill = XL_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 28
    for i in range(2, 8):
        ws.column_dimensions[get_column_letter(i)].width = 14

    for row_idx, d in enumerate(debtors, 5):
        ws.cell(row=row_idx, column=1, value=d.get("customer", ""))
        ws.cell(row=row_idx, column=2, value=_safe_float(d.get("current", 0)))
        ws.cell(row=row_idx, column=3, value=_safe_float(d.get("1_30", 0)))
        ws.cell(row=row_idx, column=4, value=_safe_float(d.get("31_60", 0)))
        ws.cell(row=row_idx, column=5, value=_safe_float(d.get("61_90", 0)))
        ws.cell(row=row_idx, column=6, value=_safe_float(d.get("over_90", 0)))
        # Total formula per customer
        ws.cell(row=row_idx, column=7,
                value=f"=SUM(B{row_idx}:F{row_idx})")

        for c in range(2, 8):
            ws.cell(row=row_idx, column=c).number_format = '#,##0.00'
        if row_idx % 2 == 0:
            for c in range(1, 8):
                ws.cell(row=row_idx, column=c).fill = XL_ALT_FILL

    last_row = 4 + len(debtors)
    sum_row = last_row + 1
    ws.cell(row=sum_row, column=1, value="TOTALS").font = XL_TOTAL_FONT
    for col in range(2, 8):
        col_letter = get_column_letter(col)
        ws.cell(row=sum_row, column=col,
                value=f"=SUM({col_letter}5:{col_letter}{last_row})")
        ws.cell(row=sum_row, column=col).font = XL_TOTAL_FONT
        ws.cell(row=sum_row, column=col).number_format = '#,##0.00'
        ws.cell(row=sum_row, column=col).border = TOTAL_BORDER
        ws.cell(row=sum_row, column=col).fill = XL_TOTAL_FILL

    ws.freeze_panes = "B5"

    wb.save(filepath)
    logger.info(f"Aged Debtors XLSX exported: {filepath}")
    return filepath


# =========================================================================
# EXCEL 5 — CASH FLOW FORECAST
# =========================================================================

def export_cash_forecast_xlsx(forecast_data: Dict[str, Any],
                               filepath: str,
                               entity: str = "Client") -> str:
    """
    Export cash flow forecast to Excel.

    forecast_data: output from hnc_forecast — includes monthly projections,
                   payment schedule, what-if scenarios.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Forecast"

    ws.merge_cells("A1:N1")
    ws["A1"] = f"{entity} — Cash Flow Forecast"
    ws["A1"].font = XL_TITLE_FONT

    # Monthly projections
    months = forecast_data.get("months", [])
    if not months:
        # Generate placeholder months
        months = [f"M{i+1}" for i in range(12)]

    headers = ["Item"] + months + ["Annual"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = XL_HEADER_FONT
        cell.fill = XL_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 24
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13

    total_col = len(months) + 2
    first_col = get_column_letter(2)
    last_col = get_column_letter(1 + len(months))

    # Revenue row
    row = 4
    ws.cell(row=row, column=1, value="Revenue (Projected)").font = XL_TOTAL_FONT
    revenues = forecast_data.get("monthly_revenue", [0] * 12)
    for col_idx, val in enumerate(revenues, 2):
        c = ws.cell(row=row, column=col_idx, value=_safe_float(val))
        c.number_format = '#,##0.00'
        c.font = Font(name="Arial", size=10, color="0000FF")
    ws.cell(row=row, column=total_col,
            value=f"=SUM({first_col}{row}:{last_col}{row})")
    ws.cell(row=row, column=total_col).number_format = '#,##0.00'
    ws.cell(row=row, column=total_col).font = XL_TOTAL_FONT

    # Expenses row
    row = 5
    ws.cell(row=row, column=1, value="Expenses (Projected)").font = XL_BODY_FONT
    expenses = forecast_data.get("monthly_expenses", [0] * 12)
    for col_idx, val in enumerate(expenses, 2):
        c = ws.cell(row=row, column=col_idx, value=_safe_float(val))
        c.number_format = '#,##0.00'
    ws.cell(row=row, column=total_col,
            value=f"=SUM({first_col}{row}:{last_col}{row})")
    ws.cell(row=row, column=total_col).number_format = '#,##0.00'

    # Net cash flow
    row = 6
    ws.cell(row=row, column=1, value="Net Cash Flow").font = XL_TOTAL_FONT
    ws.cell(row=row, column=1).fill = XL_TOTAL_FILL
    for col_idx in range(2, total_col + 1):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=row, column=col_idx,
                value=f"={col_letter}4-{col_letter}5")
        ws.cell(row=row, column=col_idx).font = XL_TOTAL_FONT
        ws.cell(row=row, column=col_idx).fill = XL_TOTAL_FILL
        ws.cell(row=row, column=col_idx).number_format = '#,##0.00'
        ws.cell(row=row, column=col_idx).border = TOTAL_BORDER

    # Tax payments row
    row = 8
    ws.cell(row=row, column=1, value="Tax Payments Due").font = Font(
        name="Arial", size=10, bold=True, color="C0392B")
    schedule = forecast_data.get("payment_schedule", [])
    # Map month names to column indices
    for payment in schedule:
        p_month = payment.get("month_index", 0)
        if 1 <= p_month <= len(months):
            c = ws.cell(row=row, column=1 + p_month,
                        value=_safe_float(payment.get("amount", 0)))
            c.number_format = '#,##0.00'
            c.font = Font(name="Arial", size=10, color="C0392B")

    # Cumulative cash
    row = 9
    ws.cell(row=row, column=1, value="Cumulative Cash").font = XL_TOTAL_FONT
    opening = _safe_float(forecast_data.get("opening_balance", 0))
    for col_idx in range(2, total_col):
        col_letter = get_column_letter(col_idx)
        if col_idx == 2:
            ws.cell(row=row, column=col_idx,
                    value=f"={opening}+{col_letter}6-{col_letter}8")
        else:
            prev = get_column_letter(col_idx - 1)
            ws.cell(row=row, column=col_idx,
                    value=f"={prev}{row}+{col_letter}6-{col_letter}8")
        ws.cell(row=row, column=col_idx).number_format = '#,##0.00'
        ws.cell(row=row, column=col_idx).font = XL_TOTAL_FONT

    ws.freeze_panes = "B4"

    # What-if scenarios sheet
    scenarios = forecast_data.get("scenarios", [])
    if scenarios:
        ws2 = wb.create_sheet("What-If Scenarios")
        ws2["A1"] = "What-If Scenarios"
        ws2["A1"].font = XL_TITLE_FONT

        headers = ["Scenario", "Annual Revenue", "Annual Expenses",
                   "Net Profit", "Tax Saving"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws2.cell(row=3, column=col_idx, value=h)
            cell.font = XL_HEADER_FONT
            cell.fill = XL_HEADER_FILL

        for row_idx, s in enumerate(scenarios, 4):
            ws2.cell(row=row_idx, column=1, value=s.get("name", ""))
            ws2.cell(row=row_idx, column=2,
                     value=_safe_float(s.get("revenue", 0)))
            ws2.cell(row=row_idx, column=3,
                     value=_safe_float(s.get("expenses", 0)))
            ws2.cell(row=row_idx, column=4,
                     value=_safe_float(s.get("net_profit", 0)))
            ws2.cell(row=row_idx, column=5,
                     value=_safe_float(s.get("tax_saving", 0)))
            for c in range(2, 6):
                ws2.cell(row=row_idx, column=c).number_format = '#,##0.00'

        for i, w in enumerate([30, 16, 16, 16, 16], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(filepath)
    logger.info(f"Cash Forecast XLSX exported: {filepath}")
    return filepath


# =========================================================================
# MASTER EXPORT CLASS
# =========================================================================

class HNCExportEngine:
    """
    Master export engine — takes pipeline output and generates
    all document types.
    """

    def __init__(self, output_dir: str, entity_name: str = "Client"):
        self.output_dir = output_dir
        self.entity = entity_name
        os.makedirs(output_dir, exist_ok=True)
        self.generated: List[Dict[str, str]] = []

    def _path(self, filename: str) -> str:
        return os.path.join(self.output_dir, filename)

    def _record(self, doc_type: str, filepath: str):
        self.generated.append({"type": doc_type, "path": filepath})

    # ----- PDF exports -----

    def export_pnl(self, report_data, filename: str = "pnl.pdf") -> str:
        fp = self._path(filename)
        export_pnl_pdf(report_data, fp, self.entity)
        self._record("P&L Statement", fp)
        return fp

    def export_balance_sheet(self, report_data,
                             filename: str = "balance_sheet.pdf") -> str:
        fp = self._path(filename)
        export_balance_sheet_pdf(report_data, fp, self.entity)
        self._record("Balance Sheet", fp)
        return fp

    def export_tax_summary(self, tax_data,
                           filename: str = "tax_summary.pdf") -> str:
        fp = self._path(filename)
        export_tax_summary_pdf(tax_data, fp, self.entity)
        self._record("Tax Summary", fp)
        return fp

    def export_invoice(self, invoice,
                       filename: str = "") -> str:
        if not filename:
            inv_num = getattr(invoice, "invoice_number", "INV")
            filename = f"invoice_{inv_num}.pdf"
        fp = self._path(filename)
        export_invoice_pdf(invoice, fp)
        self._record("Invoice", fp)
        return fp

    def export_cis_summary(self, cis_data,
                           filename: str = "cis_summary.pdf") -> str:
        fp = self._path(filename)
        export_cis_summary_pdf(cis_data, fp, self.entity)
        self._record("CIS Summary", fp)
        return fp

    def export_vat_return(self, vat_data,
                          filename: str = "vat_return.pdf") -> str:
        fp = self._path(filename)
        export_vat_return_pdf(vat_data, fp, self.entity)
        self._record("VAT Return", fp)
        return fp

    def export_cgt_summary(self, cgt_data,
                           filename: str = "cgt_summary.pdf") -> str:
        fp = self._path(filename)
        export_cgt_summary_pdf(cgt_data, fp, self.entity)
        self._record("CGT Summary", fp)
        return fp

    # ----- Excel exports -----

    def export_management_accounts(self, monthly_data,
                                    filename: str = "management_accounts.xlsx") -> str:
        fp = self._path(filename)
        export_management_accounts_xlsx(monthly_data, fp, self.entity)
        self._record("Management Accounts", fp)
        return fp

    def export_ledger(self, journal_entries,
                      filename: str = "general_ledger.xlsx") -> str:
        fp = self._path(filename)
        export_ledger_xlsx(journal_entries, fp, self.entity)
        self._record("General Ledger", fp)
        return fp

    def export_trial_balance(self, accounts, period: str = "",
                             filename: str = "trial_balance.xlsx") -> str:
        fp = self._path(filename)
        export_trial_balance_xlsx(accounts, fp, self.entity, period)
        self._record("Trial Balance", fp)
        return fp

    def export_aged_debtors(self, debtors, as_at: str = "",
                            filename: str = "aged_debtors.xlsx") -> str:
        fp = self._path(filename)
        export_aged_debtors_xlsx(debtors, fp, self.entity, as_at)
        self._record("Aged Debtors", fp)
        return fp

    def export_cash_forecast(self, forecast_data,
                              filename: str = "cash_forecast.xlsx") -> str:
        fp = self._path(filename)
        export_cash_forecast_xlsx(forecast_data, fp, self.entity)
        self._record("Cash Forecast", fp)
        return fp

    # ----- Summary -----

    def print_export_summary(self) -> str:
        """Return a formatted summary of all generated documents."""
        lines = [
            "=" * 60,
            "HNC EXPORT — DOCUMENT GENERATION SUMMARY",
            "=" * 60,
            f"Entity:    {self.entity}",
            f"Output:    {self.output_dir}",
            f"Generated: {_today_uk()}",
            f"Documents: {len(self.generated)}",
            "-" * 60,
        ]
        for i, g in enumerate(self.generated, 1):
            ext = Path(g["path"]).suffix.upper()
            lines.append(f"  {i:2d}. [{ext[1:]}] {g['type']}")
            lines.append(f"      {g['path']}")
        lines.append("=" * 60)
        return "\n".join(lines)


# =========================================================================
# STANDALONE TEST
# =========================================================================

if __name__ == "__main__":
    import tempfile

    print("=" * 60)
    print("HNC EXPORT ENGINE — DOCUMENT GENERATION TEST")
    print("=" * 60)

    with tempfile.TemporaryDirectory() as tmpdir:
        engine = HNCExportEngine(tmpdir, "John Smith t/a Smith Builders")

        # ---- Test 1: P&L PDF ----
        print("\n--- Test 1: P&L PDF ---")
        from dataclasses import dataclass, field as dc_field

        @dataclass
        class MockLine:
            label: str = ""
            amount: float = 0.0
            indent: int = 0
            is_total: bool = False
            is_subtotal: bool = False
            note_ref: str = ""
            account_code: str = ""

        @dataclass
        class MockReport:
            title: str = ""
            entity: str = ""
            period: str = ""
            report_type: str = ""
            lines: list = dc_field(default_factory=list)
            totals: dict = dc_field(default_factory=dict)
            notes: list = dc_field(default_factory=list)

        mock_pnl = MockReport(
            title="Profit & Loss Account",
            period="Year ended 5 April 2026",
            lines=[
                MockLine("TURNOVER", 185000.00, 0, False),
                MockLine("Cost of Sales", 0, 0),
                MockLine("  Materials", -42000.00, 1),
                MockLine("  Subcontractor Costs", -38000.00, 1),
                MockLine("  Plant Hire", -8500.00, 1),
                MockLine("Cost of Sales Total", -88500.00, 0, True),
                MockLine("GROSS PROFIT", 96500.00, 0, True),
                MockLine("Administrative Expenses", 0, 0),
                MockLine("  Motor Expenses", -6200.00, 1),
                MockLine("  Insurance", -3800.00, 1),
                MockLine("  Accountancy Fees", -1200.00, 1),
                MockLine("  Office & Admin", -2400.00, 1),
                MockLine("  Tools & Equipment", -4100.00, 1),
                MockLine("Admin Total", -17700.00, 0, True),
                MockLine("NET PROFIT", 78800.00, 0, True),
            ],
            notes=[
                "Note 1: Prepared under FRS 102 Section 1A.",
                "Note 2: Turnover is stated net of VAT.",
                "Note 3: Depreciation policy — straight line over useful life.",
            ],
        )
        fp1 = engine.export_pnl(mock_pnl)
        print(f"  Created: {fp1}")
        print(f"  Size: {os.path.getsize(fp1):,} bytes")

        # ---- Test 2: Tax Summary PDF ----
        print("\n--- Test 2: Tax Summary PDF ---")
        tax_data = {
            "tax_year": "2025/26",
            "turnover": 185000.00,
            "total_expenses": 106200.00,
            "net_profit": 78800.00,
            "income_tax": 15220.00,
            "nic2": 179.40,
            "nic4": 3561.72,
            "total_tax": 18961.12,
            "effective_rate": 24.1,
            "payment_schedule": [
                {"date": "2027-01-31", "description":
                 "Balancing Payment + 1st POA", "amount": 14220.84},
                {"date": "2027-07-31", "description":
                 "2nd Payment on Account", "amount": 9480.56},
            ],
        }
        fp2 = engine.export_tax_summary(tax_data)
        print(f"  Created: {fp2}")
        print(f"  Size: {os.path.getsize(fp2):,} bytes")

        # ---- Test 3: VAT Return PDF ----
        print("\n--- Test 3: VAT Return PDF ---")
        vat_data = {
            "period": "Q1 2026 (Jan-Mar)",
            "scheme": "Flat Rate Scheme",
            "flat_rate_percentage": 9.5,
            "box1": 4432.50,
            "box2": 0.00,
            "box3": 4432.50,
            "box4": 1247.80,
            "box5": 3184.70,
            "box6": 46657.89,
            "box7": 22840.00,
            "box8": 0.00,
            "box9": 0.00,
        }
        fp3 = engine.export_vat_return(vat_data)
        print(f"  Created: {fp3}")
        print(f"  Size: {os.path.getsize(fp3):,} bytes")

        # ---- Test 4: Management Accounts XLSX ----
        print("\n--- Test 4: Management Accounts XLSX ---")
        monthly = []
        base_revenue = 14000
        for i in range(12):
            month_names = ["Apr", "May", "Jun", "Jul", "Aug", "Sep",
                           "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
            seasonal = [0.7, 0.9, 1.1, 1.2, 1.15, 1.1,
                        1.0, 0.85, 0.6, 0.75, 0.85, 1.0]
            rev = round(base_revenue * seasonal[i], 2)
            monthly.append({
                "month": month_names[i],
                "turnover": rev,
                "expenses": {
                    "materials": round(rev * 0.25, 2),
                    "subcontractor_costs": round(rev * 0.20, 2),
                    "motor_expenses": 520.00,
                    "insurance": 316.67,
                    "tools_equipment": round(rev * 0.03, 2),
                    "office_admin": 200.00,
                },
            })
        fp4 = engine.export_management_accounts(monthly)
        print(f"  Created: {fp4}")
        print(f"  Size: {os.path.getsize(fp4):,} bytes")

        # ---- Test 5: General Ledger XLSX ----
        print("\n--- Test 5: General Ledger XLSX ---")
        journal = [
            {"date": "2025-04-06", "account": "4000 - Sales",
             "description": "Invoice INV-001 Mr Davies Bathroom",
             "reference": "INV-001", "debit": 0, "credit": 6200.00},
            {"date": "2025-04-06", "account": "1100 - Trade Debtors",
             "description": "Invoice INV-001 Mr Davies Bathroom",
             "reference": "INV-001", "debit": 7440.00, "credit": 0},
            {"date": "2025-04-06", "account": "2200 - VAT Control",
             "description": "Output VAT INV-001",
             "reference": "INV-001", "debit": 0, "credit": 1240.00},
            {"date": "2025-04-10", "account": "5000 - Materials",
             "description": "Travis Perkins - tiles",
             "reference": "PO-1234", "debit": 2800.00, "credit": 0},
            {"date": "2025-04-15", "account": "1200 - Bank",
             "description": "Payment from Mr Davies",
             "reference": "BAC", "debit": 7440.00, "credit": 0},
            {"date": "2025-04-15", "account": "1100 - Trade Debtors",
             "description": "Payment from Mr Davies",
             "reference": "BAC", "debit": 0, "credit": 7440.00},
        ]
        fp5 = engine.export_ledger(journal)
        print(f"  Created: {fp5}")
        print(f"  Size: {os.path.getsize(fp5):,} bytes")

        # ---- Test 6: Trial Balance XLSX ----
        print("\n--- Test 6: Trial Balance XLSX ---")
        accounts = [
            {"account_code": "1100", "account_name": "Trade Debtors",
             "debit": 12400.00, "credit": 0},
            {"account_code": "1200", "account_name": "Bank Current Account",
             "debit": 34500.00, "credit": 0},
            {"account_code": "2200", "account_name": "VAT Control",
             "debit": 0, "credit": 4200.00},
            {"account_code": "3000", "account_name": "Capital",
             "debit": 0, "credit": 5000.00},
            {"account_code": "4000", "account_name": "Sales",
             "debit": 0, "credit": 185000.00},
            {"account_code": "5000", "account_name": "Materials",
             "debit": 42000.00, "credit": 0},
            {"account_code": "5100", "account_name": "Subcontractor Costs",
             "debit": 38000.00, "credit": 0},
            {"account_code": "6000", "account_name": "Motor Expenses",
             "debit": 6200.00, "credit": 0},
            {"account_code": "6100", "account_name": "Insurance",
             "debit": 3800.00, "credit": 0},
            {"account_code": "7000", "account_name": "Drawings",
             "debit": 57300.00, "credit": 0},
        ]
        fp6 = engine.export_trial_balance(accounts, "5 April 2026")
        print(f"  Created: {fp6}")
        print(f"  Size: {os.path.getsize(fp6):,} bytes")

        # ---- Summary ----
        print("\n" + engine.print_export_summary())
        print("\nALL EXPORT TESTS PASSED")
